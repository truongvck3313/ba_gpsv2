import openpyxl
from collections import defaultdict
import chucnangkhac
import var
import re
import caseid







def ModuleTest():
    modetest = ''.join(re.findall(r'\d+', var.modetest))
    for i in modetest:
        print("so1", i)
        if i == "0":
            run_all(self='')

        else:
            moduletest = ''.join(re.findall(r'\d+', var.moduletest))
            print(type(moduletest))
            print(moduletest)
            for i in moduletest:
                print("so2", i)
                if i == "1":
                    login(self='')
                if i == "2":
                    monitor(self='')
                if i == "3":
                    route(self='')
                if i == "4":
                    administration(self='')
                if i == "5":
                    report(self='')
                if i == "6":
                    videoclip(self='')
                if i == "7":
                    image(self='')
                if i == "8":
                    utility(self='')
                if i == "9":
                    ai(self='')




def change_casenone():
    list_casefail = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                    sheet["G" + rownum] = "Fail"
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            wordbook.save(var.checklistpath)
            print("Số case trống mức1: ", count)


        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                    sheet["G" + rownum] = "Fail"
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            wordbook.save(var.checklistpath)
            print("Số case trống mức2: ", count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                    sheet["G" + rownum] = "Fail"
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            wordbook.save(var.checklistpath)
            print("Số case trống mức3: ", count)


        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                    sheet["G" + rownum] = "Fail"
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            wordbook.save(var.checklistpath)
            print("Số case trống mức4: ", count)


        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == None:
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                    sheet["G" + rownum] = "Fail"
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            wordbook.save(var.checklistpath)
            print("Số case trống mức 0 : ", count)







def check_casenone():
    var.writeData(var.path_luutamthoi, "Sheet1", 60, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 61, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 62, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 63, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 64, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 65, 2, "0")


    list_casefail = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 61, 2, count)


        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 62, 2, count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 63, 2, count)


        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 64, 2, count)


        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == None:
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 66, 2, count)
            var.writeData(var.path_luutamthoi, "Sheet1", 60, 2, count)

        else:
            muc1 = int(var.readData(var.path_luutamthoi, 'Sheet1', 61, 2))
            muc2 = int(var.readData(var.path_luutamthoi, 'Sheet1', 62, 2))
            muc3 = int(var.readData(var.path_luutamthoi, 'Sheet1', 63, 2))
            muc4 = int(var.readData(var.path_luutamthoi, 'Sheet1', 64, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var.writeData(var.path_luutamthoi, "Sheet1", 60, 2, sumarry_case_none)
            var.writeData(var.path_luutamthoi, "Sheet1", 66, 2, sumarry_case_none)



    rownum = 9
    #Đặc biệt:
    list_case_nghiemtrong = []
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        if sheet["L" + rownum].value == "x" and sheet["G" + rownum].value == "Fail":
            print(sheet["A" + rownum].value)
            case_fail = sheet["A" + rownum].value
            list_case_nghiemtrong.append(case_fail)
        rownum = int(rownum)
    print(list_case_nghiemtrong)
    count = len(list_case_nghiemtrong)
    print("Số case fail mức nghiêm trọng: ", count)
    var.writeData(var.path_luutamthoi, "Sheet1", 65, 2, count)





def check_casefail():
    var.writeData(var.path_luutamthoi, "Sheet1", 70, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 71, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 72, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 73, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 74, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 75, 2, "0")
    var.driver.set_page_load_timeout(20)

    list_casefail = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 71, 2, count)


        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 72, 2, count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 73, 2, count)


        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 74, 2, count)


        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == "Fail":
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 76, 2, count)
            var.writeData(var.path_luutamthoi, "Sheet1", 70, 2, count)

        else:
            muc1 = int(var.readData(var.path_luutamthoi, 'Sheet1', 71, 2))
            muc2 = int(var.readData(var.path_luutamthoi, 'Sheet1', 72, 2))
            muc3 = int(var.readData(var.path_luutamthoi, 'Sheet1', 73, 2))
            muc4 = int(var.readData(var.path_luutamthoi, 'Sheet1', 74, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var.writeData(var.path_luutamthoi, "Sheet1", 70, 2, sumarry_case_none)
            var.writeData(var.path_luutamthoi, "Sheet1", 76, 2, sumarry_case_none)



    rownum = 9
    #Đặc biệt:
    list_case_nghiemtrong = []
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        if sheet["L" + rownum].value == "x" and sheet["G" + rownum].value == "Fail":
            print(sheet["A" + rownum].value)
            case_fail = sheet["A" + rownum].value
            list_case_nghiemtrong.append(case_fail)
        rownum = int(rownum)
    print(list_case_nghiemtrong)
    count = len(list_case_nghiemtrong)
    print("Số case fail mức nghiêm trọng: ", count)
    var.writeData(var.path_luutamthoi, "Sheet1", 75, 2, count)


    if var.modetest == "0":
        case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 70, 2))
        var.writeData(var.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var.modetest == "1":
        case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 71, 2))
        var.writeData(var.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var.modetest == "2":
        case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 72, 2))
        var.writeData(var.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var.modetest == "3":
        case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 73, 2))
        var.writeData(var.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var.modetest == "4":
        case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 74, 2))
        var.writeData(var.path_luutamthoi, "Sheet1", 77, 2, case_fail)




def check_casepass():
    var.writeData(var.path_luutamthoi, "Sheet1", 70, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 71, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 72, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 73, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 74, 2, "0")
    var.writeData(var.path_luutamthoi, "Sheet1", 75, 2, "0")


    list_casefail = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 81, 2, count)


        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 82, 2, count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 83, 2, count)


        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 84, 2, count)


        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == "Pass":
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var.writeData(var.path_luutamthoi, "Sheet1", 86, 2, count)
            var.writeData(var.path_luutamthoi, "Sheet1", 80, 2, count)

        else:
            muc1 = int(var.readData(var.path_luutamthoi, 'Sheet1', 71, 2))
            muc2 = int(var.readData(var.path_luutamthoi, 'Sheet1', 72, 2))
            muc3 = int(var.readData(var.path_luutamthoi, 'Sheet1', 73, 2))
            muc4 = int(var.readData(var.path_luutamthoi, 'Sheet1', 74, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var.writeData(var.path_luutamthoi, "Sheet1", 80, 2, sumarry_case_none)
            var.writeData(var.path_luutamthoi, "Sheet1", 86, 2, sumarry_case_none)

        if var.modetest == "0":
            case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 80, 2))
            var.writeData(var.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var.modetest == "1":
            case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 81, 2))
            var.writeData(var.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var.modetest == "2":
            case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 82, 2))
            var.writeData(var.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var.modetest == "3":
            case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 83, 2))
            var.writeData(var.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var.modetest == "4":
            case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 84, 2))
            var.writeData(var.path_luutamthoi, "Sheet1", 87, 2, case_fail)




def retest_casenone(self):
    var.driver.set_page_load_timeout(20)

    list_casefail = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9

    modetest = ''.join(re.findall(r'\d+', var.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)


        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)

        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)

        for case in list_casefail:
            try:
                if case == 'Login01':
                    caseid.caseid_login01(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login02':
                    caseid.caseid_login02(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login03':
                    caseid.caseid_login03(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login04':
                    caseid.caseid_login04(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login05':
                    caseid.caseid_login05(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login08':
                    caseid.caseid_login08(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login09':
                    caseid.caseid_login09(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login10':
                    caseid.caseid_login10(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login11':
                    caseid.caseid_login11(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login12':
                    caseid.caseid_login12(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login13':
                    caseid.caseid_login13(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login14':
                    caseid.caseid_login14(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login15':
                    caseid.caseid_login15(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login16':
                    caseid.caseid_login16(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login17':
                    caseid.caseid_login17(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login18':
                    caseid.caseid_login18(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login19':
                    caseid.caseid_login19(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login20':
                    caseid.caseid_login20(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Login21':
                    caseid.caseid_login21(self)
            except:
                chucnangkhac.swich_tab_0()

            try:
                if case == 'GiamSat01':
                    caseid.caseid_giamsat01(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat02':
                    caseid.caseid_giamsat02(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat03':
                    caseid.caseid_giamsat03(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat04':
                    caseid.caseid_giamsat04(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat05':
                    caseid.caseid_giamsat05(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat06':
                    caseid.caseid_giamsat06(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat07':
                    caseid.caseid_giamsat07(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat08':
                    caseid.caseid_giamsat08(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat09':
                    caseid.caseid_giamsat09(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat10':
                    caseid.caseid_giamsat10(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat11':
                    caseid.caseid_giamsat11(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat12':
                    caseid.caseid_giamsat12(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat13':
                    caseid.caseid_giamsat13(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat14':
                    caseid.caseid_giamsat14(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat15':
                    caseid.caseid_giamsat15(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat16':
                    caseid.caseid_giamsat16(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat17':
                    caseid.caseid_giamsat17(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat18':
                    caseid.caseid_giamsat18(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat19':
                    caseid.caseid_giamsat19(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat20':
                    caseid.caseid_giamsat20(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat21':
                    caseid.caseid_giamsat21(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat22':
                    caseid.caseid_giamsat22(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat23':
                    caseid.caseid_giamsat23(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat24':
                    caseid.caseid_giamsat24(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat25':
                    caseid.caseid_giamsat25(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat26':
                    caseid.caseid_giamsat26(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat27':
                    caseid.caseid_giamsat27(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat28':
                    caseid.caseid_giamsat28(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat29':
                    caseid.caseid_giamsat29(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat30':
                    caseid.caseid_giamsat30(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat31':
                    caseid.caseid_giamsat31(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat32':
                    caseid.caseid_giamsat32(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat33':
                    caseid.caseid_giamsat33(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat34':
                    caseid.caseid_giamsat34(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat35':
                    caseid.caseid_giamsat35(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat36':
                    caseid.caseid_giamsat36(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat37':
                    caseid.caseid_giamsat37(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat38':
                    caseid.caseid_giamsat38(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat39':
                    caseid.caseid_giamsat39(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat40':
                    caseid.caseid_giamsat40(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat41':
                    caseid.caseid_giamsat41(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat42':
                    caseid.caseid_giamsat42(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat43':
                    caseid.caseid_giamsat43(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat44':
                    caseid.caseid_giamsat44(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat45':
                    caseid.caseid_giamsat45(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat46':
                    caseid.caseid_giamsat46(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat47':
                    caseid.caseid_giamsat47(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat48':
                    caseid.caseid_giamsat48(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat49':
                    caseid.caseid_giamsat49(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat50':
                    caseid.caseid_giamsat50(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat51':
                    caseid.caseid_giamsat51(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat52':
                    caseid.caseid_giamsat52(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat53':
                    caseid.caseid_giamsat53(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat54':
                    caseid.caseid_giamsat54(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat55':
                    caseid.caseid_giamsat55(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat56':
                    caseid.caseid_giamsat56(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat57':
                    caseid.caseid_giamsat57(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat58':
                    caseid.caseid_giamsat58(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat59':
                    caseid.caseid_giamsat59(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat60':
                    caseid.caseid_giamsat60(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat61':
                    caseid.caseid_giamsat61(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat62':
                    caseid.caseid_giamsat62(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat63':
                    caseid.caseid_giamsat63(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat64':
                    caseid.caseid_giamsat64(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat65':
                    caseid.caseid_giamsat65(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat66':
                    caseid.caseid_giamsat66(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat67':
                    caseid.caseid_giamsat67(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat68':
                    caseid.caseid_giamsat68(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat69':
                    caseid.caseid_giamsat69(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat70':
                    caseid.caseid_giamsat70(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat71':
                    caseid.caseid_giamsat71(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat72':
                    caseid.caseid_giamsat72(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat73':
                    caseid.caseid_giamsat73(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat74':
                    caseid.caseid_giamsat74(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat75':
                    caseid.caseid_giamsat75(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat76':
                    caseid.caseid_giamsat76(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat77':
                    caseid.caseid_giamsat77(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat78':
                    caseid.caseid_giamsat78(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat79':
                    caseid.caseid_giamsat79(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat80':
                    caseid.caseid_giamsat80(self)
            except:
                chucnangkhac.swich_tab_0()

            try:
                if case == 'GiamSat80_1':
                    caseid.caseid_giamsat80_1(self)
            except:
                chucnangkhac.swich_tab_0()

            try:
                if case == 'GiamSat81':
                    caseid.caseid_giamsat81(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat82':
                    caseid.caseid_giamsat82(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat83':
                    caseid.caseid_giamsat83(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat84':
                    caseid.caseid_giamsat84(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat85':
                    caseid.caseid_giamsat85(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat86':
                    caseid.caseid_giamsat86(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat87':
                    caseid.caseid_giamsat87(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat88':
                    caseid.caseid_giamsat88(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat89':
                    caseid.caseid_giamsat89(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat90':
                    caseid.caseid_giamsat90(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat91':
                    caseid.caseid_giamsat91(self)
                    caseid.caseid_giamsat92(self)
                    caseid.caseid_giamsat93(self)
                    caseid.caseid_giamsat94(self)
                    caseid.caseid_giamsat95(self)
                    caseid.caseid_giamsat96(self)
                    caseid.caseid_giamsat97(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat92':
                    caseid.caseid_giamsat91(self)
                    caseid.caseid_giamsat92(self)
                    caseid.caseid_giamsat93(self)
                    caseid.caseid_giamsat94(self)
                    caseid.caseid_giamsat95(self)
                    caseid.caseid_giamsat96(self)
                    caseid.caseid_giamsat97(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat93':
                    caseid.caseid_giamsat91(self)
                    caseid.caseid_giamsat92(self)
                    caseid.caseid_giamsat93(self)
                    caseid.caseid_giamsat94(self)
                    caseid.caseid_giamsat95(self)
                    caseid.caseid_giamsat96(self)
                    caseid.caseid_giamsat97(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat94':
                    caseid.caseid_giamsat91(self)
                    caseid.caseid_giamsat92(self)
                    caseid.caseid_giamsat93(self)
                    caseid.caseid_giamsat94(self)
                    caseid.caseid_giamsat95(self)
                    caseid.caseid_giamsat96(self)
                    caseid.caseid_giamsat97(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat95':
                    caseid.caseid_giamsat91(self)
                    caseid.caseid_giamsat92(self)
                    caseid.caseid_giamsat93(self)
                    caseid.caseid_giamsat94(self)
                    caseid.caseid_giamsat95(self)
                    caseid.caseid_giamsat96(self)
                    caseid.caseid_giamsat97(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat96':
                    caseid.caseid_giamsat91(self)
                    caseid.caseid_giamsat92(self)
                    caseid.caseid_giamsat93(self)
                    caseid.caseid_giamsat94(self)
                    caseid.caseid_giamsat95(self)
                    caseid.caseid_giamsat96(self)
                    caseid.caseid_giamsat97(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat97':
                    caseid.caseid_giamsat91(self)
                    caseid.caseid_giamsat92(self)
                    caseid.caseid_giamsat93(self)
                    caseid.caseid_giamsat94(self)
                    caseid.caseid_giamsat95(self)
                    caseid.caseid_giamsat96(self)
                    caseid.caseid_giamsat97(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat98':
                    caseid.caseid_giamsat98(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat99':
                    caseid.caseid_giamsat99(self)
                    caseid.caseid_giamsat100(self)
                    caseid.caseid_giamsat101(self)
                    caseid.caseid_giamsat102(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat100':
                    caseid.caseid_giamsat99(self)
                    caseid.caseid_giamsat100(self)
                    caseid.caseid_giamsat101(self)
                    caseid.caseid_giamsat102(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat101':
                    caseid.caseid_giamsat99(self)
                    caseid.caseid_giamsat100(self)
                    caseid.caseid_giamsat101(self)
                    caseid.caseid_giamsat102(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat102':
                    caseid.caseid_giamsat99(self)
                    caseid.caseid_giamsat100(self)
                    caseid.caseid_giamsat101(self)
                    caseid.caseid_giamsat102(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat103':
                    caseid.caseid_giamsat103(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat104':
                    caseid.caseid_giamsat103(self)
                    caseid.caseid_giamsat104(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat105':
                    caseid.caseid_giamsat105(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat106':
                    caseid.caseid_giamsat106(self)
                    caseid.caseid_giamsat107(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat107':
                    caseid.caseid_giamsat106(self)
                    caseid.caseid_giamsat107(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat108':
                    caseid.caseid_giamsat108(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat109':
                    caseid.caseid_giamsat108(self)
                    caseid.caseid_giamsat109(self)
                    caseid.caseid_giamsat110(self)
                    caseid.caseid_giamsat111(self)
                    caseid.caseid_giamsat112(self)
                    caseid.caseid_giamsat113(self)
                    caseid.caseid_giamsat114(self)
                    caseid.caseid_giamsat115(self)
                    caseid.caseid_giamsat116(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat110':
                    caseid.caseid_giamsat108(self)
                    caseid.caseid_giamsat109(self)
                    caseid.caseid_giamsat110(self)
                    caseid.caseid_giamsat111(self)
                    caseid.caseid_giamsat112(self)
                    caseid.caseid_giamsat113(self)
                    caseid.caseid_giamsat114(self)
                    caseid.caseid_giamsat115(self)
                    caseid.caseid_giamsat116(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat111':
                    caseid.caseid_giamsat108(self)
                    caseid.caseid_giamsat109(self)
                    caseid.caseid_giamsat110(self)
                    caseid.caseid_giamsat111(self)
                    caseid.caseid_giamsat112(self)
                    caseid.caseid_giamsat113(self)
                    caseid.caseid_giamsat114(self)
                    caseid.caseid_giamsat115(self)
                    caseid.caseid_giamsat116(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat112':
                    caseid.caseid_giamsat108(self)
                    caseid.caseid_giamsat109(self)
                    caseid.caseid_giamsat110(self)
                    caseid.caseid_giamsat111(self)
                    caseid.caseid_giamsat112(self)
                    caseid.caseid_giamsat113(self)
                    caseid.caseid_giamsat114(self)
                    caseid.caseid_giamsat115(self)
                    caseid.caseid_giamsat116(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat113':
                    caseid.caseid_giamsat108(self)
                    caseid.caseid_giamsat109(self)
                    caseid.caseid_giamsat110(self)
                    caseid.caseid_giamsat111(self)
                    caseid.caseid_giamsat112(self)
                    caseid.caseid_giamsat113(self)
                    caseid.caseid_giamsat114(self)
                    caseid.caseid_giamsat115(self)
                    caseid.caseid_giamsat116(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat114':
                    caseid.caseid_giamsat108(self)
                    caseid.caseid_giamsat109(self)
                    caseid.caseid_giamsat110(self)
                    caseid.caseid_giamsat111(self)
                    caseid.caseid_giamsat112(self)
                    caseid.caseid_giamsat113(self)
                    caseid.caseid_giamsat114(self)
                    caseid.caseid_giamsat115(self)
                    caseid.caseid_giamsat116(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat115':
                    caseid.caseid_giamsat108(self)
                    caseid.caseid_giamsat109(self)
                    caseid.caseid_giamsat110(self)
                    caseid.caseid_giamsat111(self)
                    caseid.caseid_giamsat112(self)
                    caseid.caseid_giamsat113(self)
                    caseid.caseid_giamsat114(self)
                    caseid.caseid_giamsat115(self)
                    caseid.caseid_giamsat116(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat116':
                    caseid.caseid_giamsat108(self)
                    caseid.caseid_giamsat109(self)
                    caseid.caseid_giamsat110(self)
                    caseid.caseid_giamsat111(self)
                    caseid.caseid_giamsat112(self)
                    caseid.caseid_giamsat113(self)
                    caseid.caseid_giamsat114(self)
                    caseid.caseid_giamsat115(self)
                    caseid.caseid_giamsat116(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat117':
                    caseid.caseid_giamsat117(self)
                    caseid.caseid_giamsat118(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat118':
                    caseid.caseid_giamsat117(self)
                    caseid.caseid_giamsat118(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat119':
                    caseid.caseid_giamsat119(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat120':
                    caseid.caseid_giamsat120(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat121':
                    caseid.caseid_giamsat121(self)
                    caseid.caseid_giamsat122(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat122':
                    caseid.caseid_giamsat121(self)
                    caseid.caseid_giamsat122(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat123':
                    caseid.caseid_giamsat123(self)
                    caseid.caseid_giamsat124(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat124':
                    caseid.caseid_giamsat123(self)
                    caseid.caseid_giamsat124(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat125':
                    caseid.caseid_giamsat125(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat126':
                    caseid.caseid_giamsat126(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat127':
                    caseid.caseid_giamsat127(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat128':
                    caseid.caseid_giamsat128(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat129':
                    caseid.caseid_giamsat129(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat130':
                    caseid.caseid_giamsat130(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat131':
                    caseid.caseid_giamsat131(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat132':
                    caseid.caseid_giamsat132(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat133':
                    caseid.caseid_giamsat133(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat134':
                    caseid.caseid_giamsat134(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat135':
                    caseid.caseid_giamsat135(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat136':
                    caseid.caseid_giamsat136(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat137':
                    caseid.caseid_giamsat137(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat138':
                    caseid.caseid_giamsat138(self)
                    caseid.caseid_giamsat139(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat139':
                    caseid.caseid_giamsat138(self)
                    caseid.caseid_giamsat139(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat140':
                    caseid.caseid_giamsat140(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat143':
                    caseid.caseid_giamsat143(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat144':
                    caseid.caseid_giamsat144(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat145':
                    caseid.caseid_giamsat145(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat146':
                    caseid.caseid_giamsat146(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat147':
                    caseid.caseid_giamsat147(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat148':
                    caseid.caseid_giamsat148(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat149':
                    caseid.caseid_giamsat149(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat150':
                    caseid.caseid_giamsat150(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat151':
                    caseid.caseid_giamsat151(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat153':
                    caseid.caseid_giamsat153(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat154':
                    caseid.caseid_giamsat154(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat155':
                    caseid.caseid_giamsat155(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat156':
                    caseid.caseid_giamsat156(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat157':
                    caseid.caseid_giamsat157(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat158':
                    caseid.caseid_giamsat158(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat159':
                    caseid.caseid_giamsat159(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat160':
                    caseid.caseid_giamsat160(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat161':
                    caseid.caseid_giamsat161(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat162':
                    caseid.caseid_giamsat162(self)
                    caseid.caseid_giamsat163(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat163':
                    caseid.caseid_giamsat162(self)
                    caseid.caseid_giamsat163(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat164':
                    caseid.caseid_giamsat164(self)
                    caseid.caseid_giamsat165(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat165':
                    caseid.caseid_giamsat164(self)
                    caseid.caseid_giamsat165(self)
            except:
                chucnangkhac.swich_tab_0()

            try:
                if case == 'GiamSat165_1':
                    caseid.caseid_giamsat165_1(self)
                    caseid.caseid_giamsat165_2(self)
            except:
                chucnangkhac.swich_tab_0()

            try:
                if case == 'GiamSat165_2':
                    caseid.caseid_giamsat165_1(self)
                    caseid.caseid_giamsat165_2(self)
            except:
                chucnangkhac.swich_tab_0()


            try:
                if case == 'GiamSat166':
                    caseid.caseid_giamsat166(self)
            except:
                chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat167':
            #         caseid.caseid_giamsat167(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat168':
            #         caseid.caseid_giamsat168(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat169':
            #         caseid.caseid_giamsat169(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat170':
            #         caseid.caseid_giamsat170(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat171':
            #         caseid.caseid_giamsat171(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat172':
            #         caseid.caseid_giamsat172(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat173':
            #         caseid.caseid_giamsat173(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat174':
            #         caseid.caseid_giamsat174(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat175':
            #         caseid.caseid_giamsat175(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat176':
            #         caseid.caseid_giamsat176(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat177':
            #         caseid.caseid_giamsat177(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat178':
            #         caseid.caseid_giamsat178(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat179':
            #         caseid.caseid_giamsat179(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat180':
            #         caseid.caseid_giamsat180(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat181':
            #         caseid.caseid_giamsat181(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat182':
            #         caseid.caseid_giamsat182(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat183':
            #         caseid.caseid_giamsat183(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat184':
            #         caseid.caseid_giamsat184(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat185':
            #         caseid.caseid_giamsat185(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat186':
            #         caseid.caseid_giamsat186(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat187':
            #         caseid.caseid_giamsat187(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat188':
            #         caseid.caseid_giamsat188(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat189':
            #         caseid.caseid_giamsat189(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat190':
            #         caseid.caseid_giamsat190(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat191':
            #         caseid.caseid_giamsat191(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat192':
            #         caseid.caseid_giamsat192(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            # try:
            #     if case == 'GiamSat193':
            #         caseid.caseid_giamsat193(self)
            # except:
            #     chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat194':
                    caseid.caseid_giamsat194(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat195':
                    caseid.caseid_giamsat195(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat196':
                    caseid.caseid_giamsat196(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat197':
                    caseid.caseid_giamsat197(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat198':
                    caseid.caseid_giamsat198(self)
                    caseid.caseid_giamsat201(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat199':
                    caseid.caseid_giamsat198(self)
                    caseid.caseid_giamsat199(self)
                    caseid.caseid_giamsat200(self)
                    caseid.caseid_giamsat201(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat200':
                    caseid.caseid_giamsat198(self)
                    caseid.caseid_giamsat199(self)
                    caseid.caseid_giamsat200(self)
                    caseid.caseid_giamsat201(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat201':
                    caseid.caseid_giamsat198(self)
                    caseid.caseid_giamsat201(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat202':
                    caseid.caseid_giamsat202(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat203':
                    caseid.caseid_giamsat203(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat204':
                    caseid.caseid_giamsat204(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat205':
                    caseid.caseid_giamsat205(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat206':
                    caseid.caseid_giamsat206(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat207':
                    caseid.caseid_giamsat207(self)
                    caseid.caseid_giamsat203(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat209':
                    caseid.caseid_giamsat209(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat210':
                    caseid.caseid_giamsat210(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat211':
                    caseid.caseid_giamsat211(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat212':
                    caseid.caseid_giamsat212(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat213':
                    caseid.caseid_giamsat213(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat214':
                    caseid.caseid_giamsat214(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat215':
                    caseid.caseid_giamsat215(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat216':
                    caseid.caseid_giamsat216(self)
                    caseid.caseid_giamsat217(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat217':
                    caseid.caseid_giamsat217(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat218':
                    caseid.caseid_giamsat218(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat219':
                    caseid.caseid_giamsat219(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat220':
                    caseid.caseid_giamsat220(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat221':
                    caseid.caseid_giamsat221(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat222':
                    caseid.caseid_giamsat222(self)
                    caseid.caseid_giamsat223(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat223':
                    caseid.caseid_giamsat222(self)
                    caseid.caseid_giamsat223(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat224':
                    caseid.caseid_giamsat224(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat225':
                    caseid.caseid_giamsat225(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat226':
                    caseid.caseid_giamsat224(self)
                    caseid.caseid_giamsat226(self)
                    caseid.caseid_giamsat227(self)
                    caseid.caseid_giamsat228(self)
                    caseid.caseid_giamsat229(self)
                    caseid.caseid_giamsat230(self)
                    caseid.caseid_giamsat231(self)
                    caseid.caseid_giamsat232(self)
                    caseid.caseid_giamsat233(self)
                    caseid.caseid_giamsat234(self)
                    caseid.caseid_giamsat235(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat227':
                    caseid.caseid_giamsat224(self)
                    caseid.caseid_giamsat226(self)
                    caseid.caseid_giamsat227(self)
                    caseid.caseid_giamsat228(self)
                    caseid.caseid_giamsat229(self)
                    caseid.caseid_giamsat230(self)
                    caseid.caseid_giamsat231(self)
                    caseid.caseid_giamsat232(self)
                    caseid.caseid_giamsat233(self)
                    caseid.caseid_giamsat234(self)
                    caseid.caseid_giamsat235(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat228':
                    caseid.caseid_giamsat224(self)
                    caseid.caseid_giamsat226(self)
                    caseid.caseid_giamsat227(self)
                    caseid.caseid_giamsat228(self)
                    caseid.caseid_giamsat229(self)
                    caseid.caseid_giamsat230(self)
                    caseid.caseid_giamsat231(self)
                    caseid.caseid_giamsat232(self)
                    caseid.caseid_giamsat233(self)
                    caseid.caseid_giamsat234(self)
                    caseid.caseid_giamsat235(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat229':
                    caseid.caseid_giamsat224(self)
                    caseid.caseid_giamsat226(self)
                    caseid.caseid_giamsat227(self)
                    caseid.caseid_giamsat228(self)
                    caseid.caseid_giamsat229(self)
                    caseid.caseid_giamsat230(self)
                    caseid.caseid_giamsat231(self)
                    caseid.caseid_giamsat232(self)
                    caseid.caseid_giamsat233(self)
                    caseid.caseid_giamsat234(self)
                    caseid.caseid_giamsat235(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat230':
                    caseid.caseid_giamsat224(self)
                    caseid.caseid_giamsat226(self)
                    caseid.caseid_giamsat227(self)
                    caseid.caseid_giamsat228(self)
                    caseid.caseid_giamsat229(self)
                    caseid.caseid_giamsat230(self)
                    caseid.caseid_giamsat231(self)
                    caseid.caseid_giamsat232(self)
                    caseid.caseid_giamsat233(self)
                    caseid.caseid_giamsat234(self)
                    caseid.caseid_giamsat235(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat231':
                    caseid.caseid_giamsat224(self)
                    caseid.caseid_giamsat226(self)
                    caseid.caseid_giamsat227(self)
                    caseid.caseid_giamsat228(self)
                    caseid.caseid_giamsat229(self)
                    caseid.caseid_giamsat230(self)
                    caseid.caseid_giamsat231(self)
                    caseid.caseid_giamsat232(self)
                    caseid.caseid_giamsat233(self)
                    caseid.caseid_giamsat234(self)
                    caseid.caseid_giamsat235(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat232':
                    caseid.caseid_giamsat224(self)
                    caseid.caseid_giamsat226(self)
                    caseid.caseid_giamsat227(self)
                    caseid.caseid_giamsat228(self)
                    caseid.caseid_giamsat229(self)
                    caseid.caseid_giamsat230(self)
                    caseid.caseid_giamsat231(self)
                    caseid.caseid_giamsat232(self)
                    caseid.caseid_giamsat233(self)
                    caseid.caseid_giamsat234(self)
                    caseid.caseid_giamsat235(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat233':
                    caseid.caseid_giamsat224(self)
                    caseid.caseid_giamsat226(self)
                    caseid.caseid_giamsat227(self)
                    caseid.caseid_giamsat228(self)
                    caseid.caseid_giamsat229(self)
                    caseid.caseid_giamsat230(self)
                    caseid.caseid_giamsat231(self)
                    caseid.caseid_giamsat232(self)
                    caseid.caseid_giamsat233(self)
                    caseid.caseid_giamsat234(self)
                    caseid.caseid_giamsat235(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat234':
                    caseid.caseid_giamsat224(self)
                    caseid.caseid_giamsat226(self)
                    caseid.caseid_giamsat227(self)
                    caseid.caseid_giamsat228(self)
                    caseid.caseid_giamsat229(self)
                    caseid.caseid_giamsat230(self)
                    caseid.caseid_giamsat231(self)
                    caseid.caseid_giamsat232(self)
                    caseid.caseid_giamsat233(self)
                    caseid.caseid_giamsat234(self)
                    caseid.caseid_giamsat235(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat235':
                    caseid.caseid_giamsat224(self)
                    caseid.caseid_giamsat226(self)
                    caseid.caseid_giamsat227(self)
                    caseid.caseid_giamsat228(self)
                    caseid.caseid_giamsat229(self)
                    caseid.caseid_giamsat230(self)
                    caseid.caseid_giamsat231(self)
                    caseid.caseid_giamsat232(self)
                    caseid.caseid_giamsat233(self)
                    caseid.caseid_giamsat234(self)
                    caseid.caseid_giamsat235(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat236':
                    caseid.caseid_giamsat236(self)
                    caseid.caseid_giamsat237(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat237':
                    caseid.caseid_giamsat237(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat238':
                    caseid.caseid_giamsat238(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat239':
                    caseid.caseid_giamsat239(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat241':
                    caseid.caseid_giamsat241(self)
                    caseid.caseid_giamsat242(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat242':
                    caseid.caseid_giamsat242(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat243':
                    caseid.caseid_giamsat243(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat244':
                    caseid.caseid_giamsat244(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat245':
                    caseid.caseid_giamsat245(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat246':
                    caseid.caseid_giamsat246(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat247':
                    caseid.caseid_giamsat247(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat248':
                    caseid.caseid_giamsat248(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat249':
                    caseid.caseid_giamsat249(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat250':
                    caseid.caseid_giamsat250(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat251':
                    caseid.caseid_giamsat251(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat252':
                    caseid.caseid_giamsat252(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat253':
                    caseid.caseid_giamsat253(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat254':
                    caseid.caseid_giamsat254(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat255':
                    caseid.caseid_giamsat255(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat256':
                    caseid.caseid_giamsat256(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat258':
                    caseid.caseid_giamsat258(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat259':
                    caseid.caseid_giamsat259(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'GiamSat260':
                    caseid.caseid_giamsat260(self)
            except:
                chucnangkhac.swich_tab_0()

            try:
                if case == 'Route01':
                    caseid.caseid_route01(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route02':
                    caseid.caseid_route02(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route03':
                    caseid.caseid_route03(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route04':
                    caseid.caseid_route04(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route05':
                    caseid.caseid_route05(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route06':
                    caseid.caseid_route06(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route07':
                    caseid.caseid_route07(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route08':
                    caseid.caseid_route08(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route09':
                    caseid.caseid_route09(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route10':
                    caseid.caseid_route10(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route11':
                    caseid.caseid_route11(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route12':
                    caseid.caseid_route12(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route13':
                    caseid.caseid_route13(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route14':
                    caseid.caseid_route14(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route15':
                    caseid.caseid_route15(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route16':
                    caseid.caseid_route16(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route17':
                    caseid.caseid_route17(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route18':
                    caseid.caseid_route18(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route19':
                    caseid.caseid_route19(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route20':
                    caseid.caseid_route20(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route21':
                    caseid.caseid_route21(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route22':
                    caseid.caseid_route22(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Route23':
                    caseid.caseid_route23(self)
            except:
                chucnangkhac.swich_tab_0()

            try:
                if case == 'Admin01':
                    caseid.caseid_admin01(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin02':
                    caseid.caseid_admin02(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin03':
                    caseid.caseid_admin03(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin04':
                    caseid.caseid_admin04(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin05':
                    caseid.caseid_admin05(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin06':
                    caseid.caseid_admin06(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin07':
                    caseid.caseid_admin07(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin08':
                    caseid.caseid_admin08(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin09':
                    caseid.caseid_admin09(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin10':
                    caseid.caseid_admin10(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin11':
                    caseid.caseid_admin11(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin12':
                    caseid.caseid_admin12(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin13':
                    caseid.caseid_admin13(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin14':
                    caseid.caseid_admin14(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin15':
                    caseid.caseid_admin15(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin16':
                    caseid.caseid_admin16(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin17':
                    caseid.caseid_admin17(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin18':
                    caseid.caseid_admin18(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin19':
                    caseid.caseid_admin19(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin20':
                    caseid.caseid_admin20(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin21':
                    caseid.caseid_admin21(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin22':
                    caseid.caseid_admin22(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin23':
                    caseid.caseid_admin22(self)
                    caseid.caseid_admin23(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin24':
                    caseid.caseid_admin22(self)
                    caseid.caseid_admin23(self)
                    caseid.caseid_admin24(self)
                    caseid.caseid_admin25(self)
                    caseid.caseid_admin26(self)
                    caseid.caseid_admin27(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin25':
                    caseid.caseid_admin22(self)
                    caseid.caseid_admin23(self)
                    caseid.caseid_admin24(self)
                    caseid.caseid_admin25(self)
                    caseid.caseid_admin26(self)
                    caseid.caseid_admin27(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin26':
                    caseid.caseid_admin22(self)
                    caseid.caseid_admin23(self)
                    caseid.caseid_admin24(self)
                    caseid.caseid_admin25(self)
                    caseid.caseid_admin26(self)
                    caseid.caseid_admin27(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin27':
                    caseid.caseid_admin22(self)
                    caseid.caseid_admin23(self)
                    caseid.caseid_admin24(self)
                    caseid.caseid_admin25(self)
                    caseid.caseid_admin26(self)
                    caseid.caseid_admin27(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin28':
                    caseid.caseid_admin28(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin29':
                    caseid.caseid_admin29(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin30':
                    caseid.caseid_admin30(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin31':
                    caseid.caseid_admin31(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin32':
                    caseid.caseid_admin32(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin33':
                    caseid.caseid_admin33(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin34':
                    caseid.caseid_admin34(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Admin35':
                    caseid.caseid_admin35(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User01':
                    caseid.caseid_user01(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User02':
                    caseid.caseid_user02(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User03':
                    caseid.caseid_user03(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User04':
                    caseid.caseid_user04(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User05':
                    caseid.caseid_user05(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User06':
                    caseid.caseid_user06(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User07':
                    caseid.caseid_user07(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User08':
                    caseid.caseid_user08(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User09':
                    caseid.caseid_user09(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User10':
                    caseid.caseid_user10(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User11':
                    caseid.caseid_user11(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User12':
                    caseid.caseid_user12(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User12':
                    caseid.caseid_user12(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User13':
                    caseid.caseid_user13(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User14':
                    caseid.caseid_user14(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User15':
                    caseid.caseid_user15(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User16':
                    caseid.caseid_user16(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'User17':
                    caseid.caseid_user17(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report01':
                    caseid.caseid_report01(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report02':
                    caseid.caseid_report02(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report03':
                    caseid.caseid_report03(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report04':
                    caseid.caseid_report04(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report05':
                    caseid.caseid_report05(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report06':
                    caseid.caseid_report06(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report07':
                    caseid.caseid_report07(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report08':
                    caseid.caseid_report08(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report09':
                    caseid.caseid_report09(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report10':
                    caseid.caseid_report10(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report11':
                    caseid.caseid_report11(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report12':
                    caseid.caseid_report12(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report13':
                    caseid.caseid_report13(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report14':
                    caseid.caseid_report14(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report15':
                    caseid.caseid_report15(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report16':
                    caseid.caseid_report16(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report17':
                    caseid.caseid_report17(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report18':
                    caseid.caseid_report18(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report19':
                    caseid.caseid_report19(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report20':
                    caseid.caseid_report20(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report21':
                    caseid.caseid_report21(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report22':
                    caseid.caseid_report22(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report23':
                    caseid.caseid_report23(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report24':
                    caseid.caseid_report24(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report25':
                    caseid.caseid_report25(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report26':
                    caseid.caseid_report26(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report27':
                    caseid.caseid_report27(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report28':
                    caseid.caseid_report28(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report29':
                    caseid.caseid_report29(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report30':
                    caseid.caseid_report30(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report31':
                    caseid.caseid_report31(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report32':
                    caseid.caseid_report32(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report33':
                    caseid.caseid_report33(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report34':
                    caseid.caseid_report34(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report35':
                    caseid.caseid_report35(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report36':
                    caseid.caseid_report36(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report37':
                    caseid.caseid_report37(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report38':
                    caseid.caseid_report38(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report39':
                    caseid.caseid_report39(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report40':
                    caseid.caseid_report40(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report41':
                    caseid.caseid_report41(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report42':
                    caseid.caseid_report42(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report43':
                    caseid.caseid_report43(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report44':
                    caseid.caseid_report44(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report45':
                    caseid.caseid_report45(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report46':
                    caseid.caseid_report46(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report47':
                    caseid.caseid_report47(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report48':
                    caseid.caseid_report48(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report49':
                    caseid.caseid_report49(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report50':
                    caseid.caseid_report50(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report51':
                    caseid.caseid_report51(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report52':
                    caseid.caseid_report52(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report53':
                    caseid.caseid_report53(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report54':
                    caseid.caseid_report54(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report55':
                    caseid.caseid_report55(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report56':
                    caseid.caseid_report56(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report57':
                    caseid.caseid_report57(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report58':
                    caseid.caseid_report58(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report59':
                    caseid.caseid_report59(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report60':
                    caseid.caseid_report60(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report61':
                    caseid.caseid_report61(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report62':
                    caseid.caseid_report62(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report63':
                    caseid.caseid_report63(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report64':
                    caseid.caseid_report64(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report65':
                    caseid.caseid_report65(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report66':
                    caseid.caseid_report66(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report67':
                    caseid.caseid_report67(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report68':
                    caseid.caseid_report68(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report69':
                    caseid.caseid_report69(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Report70':
                    caseid.caseid_report70(self)
            except:
                chucnangkhac.swich_tab_0()

            try:
                if case == 'Video01':
                    caseid.caseid_video01(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Video02':
                    caseid.caseid_video02(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Video03':
                    caseid.caseid_video03(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Video04':
                    caseid.caseid_video04(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Video05':
                    caseid.caseid_video05(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Video06':
                    caseid.caseid_video06(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Video07':
                    caseid.caseid_video07(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Video08':
                    caseid.caseid_video08(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Video09':
                    caseid.caseid_video09(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Video10':
                    caseid.caseid_video10(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Video11':
                    caseid.caseid_video11(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Video12':
                    caseid.caseid_video12(self)
            except:
                chucnangkhac.swich_tab_0()

            try:
                if case == 'Image01':
                    caseid.caseid_image01(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image02':
                    caseid.caseid_image02(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image03':
                    caseid.caseid_image03(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image04':
                    caseid.caseid_image04(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image05':
                    caseid.caseid_image05(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image06':
                    caseid.caseid_image06(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image07':
                    caseid.caseid_image07(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image08':
                    caseid.caseid_image08(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image09':
                    caseid.caseid_image09(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image10':
                    caseid.caseid_image10(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image11':
                    caseid.caseid_image11(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image12':
                    caseid.caseid_image12(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image13':
                    caseid.caseid_image13(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image14':
                    caseid.caseid_image14(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image15':
                    caseid.caseid_image15(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image16':
                    caseid.caseid_image16(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image17':
                    caseid.caseid_image17(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image18':
                    caseid.caseid_image18(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image19':
                    caseid.caseid_image19(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Image20':
                    caseid.caseid_image20(self)
            except:
                chucnangkhac.swich_tab_0()

            try:
                if case == 'Utility01':
                    caseid.caseid_utility01(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility02':
                    caseid.caseid_utility02(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility03':
                    caseid.caseid_utility03(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility04':
                    caseid.caseid_utility04(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility05':
                    caseid.caseid_utility05(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility06':
                    caseid.caseid_utility06(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility07':
                    caseid.caseid_utility07(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility08':
                    caseid.caseid_utility08(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility09':
                    caseid.caseid_utility09(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility10':
                    caseid.caseid_utility10(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility11':
                    caseid.caseid_utility11(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility12':
                    caseid.caseid_utility12(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility13':
                    caseid.caseid_utility13(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility14':
                    caseid.caseid_utility14(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility15':
                    caseid.caseid_utility15(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility16':
                    caseid.caseid_utility16(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility17':
                    caseid.caseid_utility17(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility18':
                    caseid.caseid_utility18(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility19':
                    caseid.caseid_utility19(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility20':
                    caseid.caseid_utility20(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility21':
                    caseid.caseid_utility21(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility22':
                    caseid.caseid_utility22(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Utility23':
                    caseid.caseid_utility23(self)
            except:
                chucnangkhac.swich_tab_0()

            try:
                if case == 'Ai01':
                    caseid.caseid_ai01(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Ai02':
                    caseid.caseid_ai02(self)
            except:
                chucnangkhac.swich_tab_0()
            try:
                if case == 'Ai03':
                    caseid.caseid_ai03(self)
            except:
                chucnangkhac.swich_tab_0()




def retest_casefail(self):
    var.driver.set_page_load_timeout(20)

    list_casefail = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        if sheet["G"+rownum].value == "Fail":
            print(sheet["A"+rownum].value)
            case_fail = sheet["A"+rownum].value
            list_casefail.append(case_fail)
        rownum = int(rownum)
    print(list_casefail)
    for case in list_casefail:
        try:
            if case == 'Login01':
                caseid.caseid_login01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login02':
                caseid.caseid_login02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login03':
                caseid.caseid_login03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login04':
                caseid.caseid_login04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login05':
                caseid.caseid_login05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login08':
                caseid.caseid_login08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login09':
                caseid.caseid_login09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login10':
                caseid.caseid_login10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login11':
                caseid.caseid_login11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login12':
                caseid.caseid_login12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login13':
                caseid.caseid_login13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login14':
                caseid.caseid_login14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login15':
                caseid.caseid_login15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login16':
                caseid.caseid_login16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login17':
                caseid.caseid_login17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login18':
                caseid.caseid_login18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login19':
                caseid.caseid_login19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login20':
                caseid.caseid_login20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login21':
                caseid.caseid_login21(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'GiamSat01':
                caseid.caseid_giamsat01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat02':
                caseid.caseid_giamsat02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat03':
                caseid.caseid_giamsat03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat04':
                caseid.caseid_giamsat04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat05':
                caseid.caseid_giamsat05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat06':
                caseid.caseid_giamsat06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat07':
                caseid.caseid_giamsat07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat08':
                caseid.caseid_giamsat08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat09':
                caseid.caseid_giamsat09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat10':
                caseid.caseid_giamsat10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat11':
                caseid.caseid_giamsat11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat12':
                caseid.caseid_giamsat12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat13':
                caseid.caseid_giamsat13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat14':
                caseid.caseid_giamsat14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat15':
                caseid.caseid_giamsat15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat16':
                caseid.caseid_giamsat16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat17':
                caseid.caseid_giamsat17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat18':
                caseid.caseid_giamsat18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat19':
                caseid.caseid_giamsat19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat20':
                caseid.caseid_giamsat20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat21':
                caseid.caseid_giamsat21(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat22':
                caseid.caseid_giamsat22(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat23':
                caseid.caseid_giamsat23(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat24':
                caseid.caseid_giamsat24(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat25':
                caseid.caseid_giamsat25(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat26':
                caseid.caseid_giamsat26(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat27':
                caseid.caseid_giamsat27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat28':
                caseid.caseid_giamsat28(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat29':
                caseid.caseid_giamsat29(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat30':
                caseid.caseid_giamsat30(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat31':
                caseid.caseid_giamsat31(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat32':
                caseid.caseid_giamsat32(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat33':
                caseid.caseid_giamsat33(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat34':
                caseid.caseid_giamsat34(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat35':
                caseid.caseid_giamsat35(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat36':
                caseid.caseid_giamsat36(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat37':
                caseid.caseid_giamsat37(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat38':
                caseid.caseid_giamsat38(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat39':
                caseid.caseid_giamsat39(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat40':
                caseid.caseid_giamsat40(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat41':
                caseid.caseid_giamsat41(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat42':
                caseid.caseid_giamsat42(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat43':
                caseid.caseid_giamsat43(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat44':
                caseid.caseid_giamsat44(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat45':
                caseid.caseid_giamsat45(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat46':
                caseid.caseid_giamsat46(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat47':
                caseid.caseid_giamsat47(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat48':
                caseid.caseid_giamsat48(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat49':
                caseid.caseid_giamsat49(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat50':
                caseid.caseid_giamsat50(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat51':
                caseid.caseid_giamsat51(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat52':
                caseid.caseid_giamsat52(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat53':
                caseid.caseid_giamsat53(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat54':
                caseid.caseid_giamsat54(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat55':
                caseid.caseid_giamsat55(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat56':
                caseid.caseid_giamsat56(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat57':
                caseid.caseid_giamsat57(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat58':
                caseid.caseid_giamsat58(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat59':
                caseid.caseid_giamsat59(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat60':
                caseid.caseid_giamsat60(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat61':
                caseid.caseid_giamsat61(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat62':
                caseid.caseid_giamsat62(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat63':
                caseid.caseid_giamsat63(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat64':
                caseid.caseid_giamsat64(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat65':
                caseid.caseid_giamsat65(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat66':
                caseid.caseid_giamsat66(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat67':
                caseid.caseid_giamsat67(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat68':
                caseid.caseid_giamsat68(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat69':
                caseid.caseid_giamsat69(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat70':
                caseid.caseid_giamsat70(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat71':
                caseid.caseid_giamsat71(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat72':
                caseid.caseid_giamsat72(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat73':
                caseid.caseid_giamsat73(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat74':
                caseid.caseid_giamsat74(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat75':
                caseid.caseid_giamsat75(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat76':
                caseid.caseid_giamsat76(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat77':
                caseid.caseid_giamsat77(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat78':
                caseid.caseid_giamsat78(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat79':
                caseid.caseid_giamsat79(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat80':
                caseid.caseid_giamsat80(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'GiamSat80_1':
                caseid.caseid_giamsat80_1(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'GiamSat81':
                caseid.caseid_giamsat81(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat82':
                caseid.caseid_giamsat82(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat83':
                caseid.caseid_giamsat83(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat84':
                caseid.caseid_giamsat84(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat85':
                caseid.caseid_giamsat85(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat86':
                caseid.caseid_giamsat86(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat87':
                caseid.caseid_giamsat87(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat88':
                caseid.caseid_giamsat88(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat89':
                caseid.caseid_giamsat89(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat90':
                caseid.caseid_giamsat90(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat91':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat92':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat93':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat94':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat95':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat96':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat97':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat98':
                caseid.caseid_giamsat98(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat99':
                caseid.caseid_giamsat99(self)
                caseid.caseid_giamsat100(self)
                caseid.caseid_giamsat101(self)
                caseid.caseid_giamsat102(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat100':
                caseid.caseid_giamsat99(self)
                caseid.caseid_giamsat100(self)
                caseid.caseid_giamsat101(self)
                caseid.caseid_giamsat102(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat101':
                caseid.caseid_giamsat99(self)
                caseid.caseid_giamsat100(self)
                caseid.caseid_giamsat101(self)
                caseid.caseid_giamsat102(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat102':
                caseid.caseid_giamsat99(self)
                caseid.caseid_giamsat100(self)
                caseid.caseid_giamsat101(self)
                caseid.caseid_giamsat102(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat103':
                caseid.caseid_giamsat103(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat104':
                caseid.caseid_giamsat103(self)
                caseid.caseid_giamsat104(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat105':
                caseid.caseid_giamsat105(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat106':
                caseid.caseid_giamsat106(self)
                caseid.caseid_giamsat107(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat107':
                caseid.caseid_giamsat106(self)
                caseid.caseid_giamsat107(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat108':
                caseid.caseid_giamsat108(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat109':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat110':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat111':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat112':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat113':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat114':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat115':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat116':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat117':
                caseid.caseid_giamsat117(self)
                caseid.caseid_giamsat118(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat118':
                caseid.caseid_giamsat117(self)
                caseid.caseid_giamsat118(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat119':
                caseid.caseid_giamsat119(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat120':
                caseid.caseid_giamsat120(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat121':
                caseid.caseid_giamsat121(self)
                caseid.caseid_giamsat122(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat122':
                caseid.caseid_giamsat121(self)
                caseid.caseid_giamsat122(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat123':
                caseid.caseid_giamsat123(self)
                caseid.caseid_giamsat124(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat124':
                caseid.caseid_giamsat123(self)
                caseid.caseid_giamsat124(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat125':
                caseid.caseid_giamsat125(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat126':
                caseid.caseid_giamsat126(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat127':
                caseid.caseid_giamsat127(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat128':
                caseid.caseid_giamsat128(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat129':
                caseid.caseid_giamsat129(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat130':
                caseid.caseid_giamsat130(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat131':
                caseid.caseid_giamsat131(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat132':
                caseid.caseid_giamsat132(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat133':
                caseid.caseid_giamsat133(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat134':
                caseid.caseid_giamsat134(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat135':
                caseid.caseid_giamsat135(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat136':
                caseid.caseid_giamsat136(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat137':
                caseid.caseid_giamsat137(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat138':
                caseid.caseid_giamsat138(self)
                caseid.caseid_giamsat139(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat139':
                caseid.caseid_giamsat138(self)
                caseid.caseid_giamsat139(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat140':
                caseid.caseid_giamsat140(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat143':
                caseid.caseid_giamsat143(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat144':
                caseid.caseid_giamsat144(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat145':
                caseid.caseid_giamsat145(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat146':
                caseid.caseid_giamsat146(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat147':
                caseid.caseid_giamsat147(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat148':
                caseid.caseid_giamsat148(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat149':
                caseid.caseid_giamsat149(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat150':
                caseid.caseid_giamsat150(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat151':
                caseid.caseid_giamsat151(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat153':
                caseid.caseid_giamsat153(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat154':
                caseid.caseid_giamsat154(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat155':
                caseid.caseid_giamsat155(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat156':
                caseid.caseid_giamsat156(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat157':
                caseid.caseid_giamsat157(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat158':
                caseid.caseid_giamsat158(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat159':
                caseid.caseid_giamsat159(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat160':
                caseid.caseid_giamsat160(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat161':
                caseid.caseid_giamsat161(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat162':
                caseid.caseid_giamsat162(self)
                caseid.caseid_giamsat163(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat163':
                caseid.caseid_giamsat162(self)
                caseid.caseid_giamsat163(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat164':
                caseid.caseid_giamsat164(self)
                caseid.caseid_giamsat165(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat165':
                caseid.caseid_giamsat164(self)
                caseid.caseid_giamsat165(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'GiamSat165_1':
                caseid.caseid_giamsat165_1(self)
                caseid.caseid_giamsat165_2(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'GiamSat165_2':
                caseid.caseid_giamsat165_1(self)
                caseid.caseid_giamsat165_2(self)
        except:
            chucnangkhac.swich_tab_0()




        try:
            if case == 'GiamSat166':
                caseid.caseid_giamsat166(self)
        except:
            chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat167':
        #         caseid.caseid_giamsat167(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat168':
        #         caseid.caseid_giamsat168(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat169':
        #         caseid.caseid_giamsat169(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat170':
        #         caseid.caseid_giamsat170(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat171':
        #         caseid.caseid_giamsat171(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat172':
        #         caseid.caseid_giamsat172(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat173':
        #         caseid.caseid_giamsat173(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat174':
        #         caseid.caseid_giamsat174(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat175':
        #         caseid.caseid_giamsat175(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat176':
        #         caseid.caseid_giamsat176(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat177':
        #         caseid.caseid_giamsat177(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat178':
        #         caseid.caseid_giamsat178(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat179':
        #         caseid.caseid_giamsat179(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat180':
        #         caseid.caseid_giamsat180(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat181':
        #         caseid.caseid_giamsat181(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat182':
        #         caseid.caseid_giamsat182(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat183':
        #         caseid.caseid_giamsat183(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat184':
        #         caseid.caseid_giamsat184(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat185':
        #         caseid.caseid_giamsat185(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat186':
        #         caseid.caseid_giamsat186(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat187':
        #         caseid.caseid_giamsat187(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat188':
        #         caseid.caseid_giamsat188(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat189':
        #         caseid.caseid_giamsat189(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat190':
        #         caseid.caseid_giamsat190(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat191':
        #         caseid.caseid_giamsat191(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat192':
        #         caseid.caseid_giamsat192(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        # try:
        #     if case == 'GiamSat193':
        #         caseid.caseid_giamsat193(self)
        # except:
        #     chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat194':
                caseid.caseid_giamsat194(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat195':
                caseid.caseid_giamsat195(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat196':
                caseid.caseid_giamsat196(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat197':
                caseid.caseid_giamsat197(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat198':
                caseid.caseid_giamsat198(self)
                caseid.caseid_giamsat201(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat199':
                caseid.caseid_giamsat198(self)
                caseid.caseid_giamsat199(self)
                caseid.caseid_giamsat200(self)
                caseid.caseid_giamsat201(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat200':
                caseid.caseid_giamsat198(self)
                caseid.caseid_giamsat199(self)
                caseid.caseid_giamsat200(self)
                caseid.caseid_giamsat201(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat201':
                caseid.caseid_giamsat198(self)
                caseid.caseid_giamsat201(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat202':
                caseid.caseid_giamsat202(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat203':
                caseid.caseid_giamsat203(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat204':
                caseid.caseid_giamsat204(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat205':
                caseid.caseid_giamsat205(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat206':
                caseid.caseid_giamsat206(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat207':
                caseid.caseid_giamsat207(self)
                caseid.caseid_giamsat203(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat209':
                caseid.caseid_giamsat209(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat210':
                caseid.caseid_giamsat210(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat211':
                caseid.caseid_giamsat211(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat212':
                caseid.caseid_giamsat212(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat213':
                caseid.caseid_giamsat213(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat214':
                caseid.caseid_giamsat214(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat215':
                caseid.caseid_giamsat215(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat216':
                caseid.caseid_giamsat216(self)
                caseid.caseid_giamsat217(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat217':
                caseid.caseid_giamsat217(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat218':
                caseid.caseid_giamsat218(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat219':
                caseid.caseid_giamsat219(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat220':
                caseid.caseid_giamsat220(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat221':
                caseid.caseid_giamsat221(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat222':
                caseid.caseid_giamsat222(self)
                caseid.caseid_giamsat223(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat223':
                caseid.caseid_giamsat222(self)
                caseid.caseid_giamsat223(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat224':
                caseid.caseid_giamsat224(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat225':
                caseid.caseid_giamsat225(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat226':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat227':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat228':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat229':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat230':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat231':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat232':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat233':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat234':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat235':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat236':
                caseid.caseid_giamsat236(self)
                caseid.caseid_giamsat237(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat237':
                caseid.caseid_giamsat237(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat238':
                caseid.caseid_giamsat238(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat239':
                caseid.caseid_giamsat239(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat241':
                caseid.caseid_giamsat241(self)
                caseid.caseid_giamsat242(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat242':
                caseid.caseid_giamsat242(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat243':
                caseid.caseid_giamsat243(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat244':
                caseid.caseid_giamsat244(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat245':
                caseid.caseid_giamsat245(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat246':
                caseid.caseid_giamsat246(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat247':
                caseid.caseid_giamsat247(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat248':
                caseid.caseid_giamsat248(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat249':
                caseid.caseid_giamsat249(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat250':
                caseid.caseid_giamsat250(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat251':
                caseid.caseid_giamsat251(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat252':
                caseid.caseid_giamsat252(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat253':
                caseid.caseid_giamsat253(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat254':
                caseid.caseid_giamsat254(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat255':
                caseid.caseid_giamsat255(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat256':
                caseid.caseid_giamsat256(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat258':
                caseid.caseid_giamsat258(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat259':
                caseid.caseid_giamsat259(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat260':
                caseid.caseid_giamsat260(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'Route01':
                caseid.caseid_route01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route02':
                caseid.caseid_route02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route03':
                caseid.caseid_route03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route04':
                caseid.caseid_route04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route05':
                caseid.caseid_route05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route06':
                caseid.caseid_route06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route07':
                caseid.caseid_route07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route08':
                caseid.caseid_route08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route09':
                caseid.caseid_route09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route10':
                caseid.caseid_route10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route11':
                caseid.caseid_route11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route12':
                caseid.caseid_route12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route13':
                caseid.caseid_route13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route14':
                caseid.caseid_route14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route15':
                caseid.caseid_route15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route16':
                caseid.caseid_route16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route17':
                caseid.caseid_route17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route18':
                caseid.caseid_route18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route19':
                caseid.caseid_route19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route20':
                caseid.caseid_route20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route21':
                caseid.caseid_route21(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route22':
                caseid.caseid_route22(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route23':
                caseid.caseid_route23(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin01':
                caseid.caseid_admin01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin02':
                caseid.caseid_admin02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin03':
                caseid.caseid_admin03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin04':
                caseid.caseid_admin04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin05':
                caseid.caseid_admin05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin06':
                caseid.caseid_admin06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin07':
                caseid.caseid_admin07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin08':
                caseid.caseid_admin08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin09':
                caseid.caseid_admin09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin10':
                caseid.caseid_admin10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin11':
                caseid.caseid_admin11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin12':
                caseid.caseid_admin12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin13':
                caseid.caseid_admin13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin14':
                caseid.caseid_admin14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin15':
                caseid.caseid_admin15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin16':
                caseid.caseid_admin16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin17':
                caseid.caseid_admin17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin18':
                caseid.caseid_admin18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin19':
                caseid.caseid_admin19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin20':
                caseid.caseid_admin20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin21':
                caseid.caseid_admin21(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin22':
                caseid.caseid_admin22(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin23':
                caseid.caseid_admin22(self)
                caseid.caseid_admin23(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin24':
                caseid.caseid_admin22(self)
                caseid.caseid_admin23(self)
                caseid.caseid_admin24(self)
                caseid.caseid_admin25(self)
                caseid.caseid_admin26(self)
                caseid.caseid_admin27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin25':
                caseid.caseid_admin22(self)
                caseid.caseid_admin23(self)
                caseid.caseid_admin24(self)
                caseid.caseid_admin25(self)
                caseid.caseid_admin26(self)
                caseid.caseid_admin27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin26':
                caseid.caseid_admin22(self)
                caseid.caseid_admin23(self)
                caseid.caseid_admin24(self)
                caseid.caseid_admin25(self)
                caseid.caseid_admin26(self)
                caseid.caseid_admin27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin27':
                caseid.caseid_admin22(self)
                caseid.caseid_admin23(self)
                caseid.caseid_admin24(self)
                caseid.caseid_admin25(self)
                caseid.caseid_admin26(self)
                caseid.caseid_admin27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin28':
                caseid.caseid_admin28(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin29':
                caseid.caseid_admin29(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin30':
                caseid.caseid_admin30(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin31':
                caseid.caseid_admin31(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin32':
                caseid.caseid_admin32(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin33':
                caseid.caseid_admin33(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin34':
                caseid.caseid_admin34(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin35':
                caseid.caseid_admin35(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User01':
                caseid.caseid_user01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User02':
                caseid.caseid_user02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User03':
                caseid.caseid_user03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User04':
                caseid.caseid_user04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User05':
                caseid.caseid_user05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User06':
                caseid.caseid_user06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User07':
                caseid.caseid_user07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User08':
                caseid.caseid_user08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User09':
                caseid.caseid_user09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User10':
                caseid.caseid_user10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User11':
                caseid.caseid_user11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User12':
                caseid.caseid_user12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User12':
                caseid.caseid_user12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User13':
                caseid.caseid_user13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User14':
                caseid.caseid_user14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User15':
                caseid.caseid_user15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User16':
                caseid.caseid_user16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User17':
                caseid.caseid_user17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report01':
                caseid.caseid_report01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report02':
                caseid.caseid_report02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report03':
                caseid.caseid_report03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report04':
                caseid.caseid_report04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report05':
                caseid.caseid_report05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report06':
                caseid.caseid_report06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report07':
                caseid.caseid_report07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report08':
                caseid.caseid_report08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report09':
                caseid.caseid_report09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report10':
                caseid.caseid_report10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report11':
                caseid.caseid_report11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report12':
                caseid.caseid_report12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report13':
                caseid.caseid_report13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report14':
                caseid.caseid_report14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report15':
                caseid.caseid_report15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report16':
                caseid.caseid_report16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report17':
                caseid.caseid_report17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report18':
                caseid.caseid_report18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report19':
                caseid.caseid_report19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report20':
                caseid.caseid_report20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report21':
                caseid.caseid_report21(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report22':
                caseid.caseid_report22(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report23':
                caseid.caseid_report23(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report24':
                caseid.caseid_report24(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report25':
                caseid.caseid_report25(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report26':
                caseid.caseid_report26(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report27':
                caseid.caseid_report27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report28':
                caseid.caseid_report28(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report29':
                caseid.caseid_report29(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report30':
                caseid.caseid_report30(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report31':
                caseid.caseid_report31(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report32':
                caseid.caseid_report32(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report33':
                caseid.caseid_report33(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report34':
                caseid.caseid_report34(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report35':
                caseid.caseid_report35(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report36':
                caseid.caseid_report36(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report37':
                caseid.caseid_report37(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report38':
                caseid.caseid_report38(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report39':
                caseid.caseid_report39(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report40':
                caseid.caseid_report40(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report41':
                caseid.caseid_report41(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report42':
                caseid.caseid_report42(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report43':
                caseid.caseid_report43(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report44':
                caseid.caseid_report44(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report45':
                caseid.caseid_report45(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report46':
                caseid.caseid_report46(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report47':
                caseid.caseid_report47(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report48':
                caseid.caseid_report48(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report49':
                caseid.caseid_report49(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report50':
                caseid.caseid_report50(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report51':
                caseid.caseid_report51(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report52':
                caseid.caseid_report52(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report53':
                caseid.caseid_report53(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report54':
                caseid.caseid_report54(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report55':
                caseid.caseid_report55(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report56':
                caseid.caseid_report56(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report57':
                caseid.caseid_report57(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report58':
                caseid.caseid_report58(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report59':
                caseid.caseid_report59(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report60':
                caseid.caseid_report60(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report61':
                caseid.caseid_report61(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report62':
                caseid.caseid_report62(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report63':
                caseid.caseid_report63(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report64':
                caseid.caseid_report64(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report65':
                caseid.caseid_report65(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report66':
                caseid.caseid_report66(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report67':
                caseid.caseid_report67(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report68':
                caseid.caseid_report68(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report69':
                caseid.caseid_report69(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report70':
                caseid.caseid_report70(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'Video01':
                caseid.caseid_video01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video02':
                caseid.caseid_video02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video03':
                caseid.caseid_video03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video04':
                caseid.caseid_video04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video05':
                caseid.caseid_video05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video06':
                caseid.caseid_video06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video07':
                caseid.caseid_video07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video08':
                caseid.caseid_video08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video09':
                caseid.caseid_video09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video10':
                caseid.caseid_video10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video11':
                caseid.caseid_video11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video12':
                caseid.caseid_video12(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'Image01':
                caseid.caseid_image01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image02':
                caseid.caseid_image02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image03':
                caseid.caseid_image03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image04':
                caseid.caseid_image04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image05':
                caseid.caseid_image05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image06':
                caseid.caseid_image06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image07':
                caseid.caseid_image07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image08':
                caseid.caseid_image08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image09':
                caseid.caseid_image09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image10':
                caseid.caseid_image10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image11':
                caseid.caseid_image11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image12':
                caseid.caseid_image12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image13':
                caseid.caseid_image13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image14':
                caseid.caseid_image14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image15':
                caseid.caseid_image15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image16':
                caseid.caseid_image16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image17':
                caseid.caseid_image17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image18':
                caseid.caseid_image18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image19':
                caseid.caseid_image19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image20':
                caseid.caseid_image20(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'Utility01':
                caseid.caseid_utility01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility02':
                caseid.caseid_utility02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility03':
                caseid.caseid_utility03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility04':
                caseid.caseid_utility04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility05':
                caseid.caseid_utility05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility06':
                caseid.caseid_utility06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility07':
                caseid.caseid_utility07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility08':
                caseid.caseid_utility08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility09':
                caseid.caseid_utility09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility10':
                caseid.caseid_utility10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility11':
                caseid.caseid_utility11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility12':
                caseid.caseid_utility12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility13':
                caseid.caseid_utility13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility14':
                caseid.caseid_utility14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility15':
                caseid.caseid_utility15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility16':
                caseid.caseid_utility16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility17':
                caseid.caseid_utility17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility18':
                caseid.caseid_utility18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility19':
                caseid.caseid_utility19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility20':
                caseid.caseid_utility20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility21':
                caseid.caseid_utility21(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility22':
                caseid.caseid_utility22(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility23':
                caseid.caseid_utility23(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'Ai01':
                caseid.caseid_ai01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Ai02':
                caseid.caseid_ai02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Ai03':
                caseid.caseid_ai03(self)
        except:
            chucnangkhac.swich_tab_0()





# def retest_serious():
#     list_casefail = []
#     grouped_by_tag = defaultdict(list)  # Nhóm theo tag
#
#     wordbook = openpyxl.load_workbook(var.checklistpath)
#     sheet = wordbook.get_sheet_by_name("Checklist")
#     rownum = 9
#     while rownum < 1200:
#         rownum += 1
#         row_str = str(rownum)
#
#         if sheet["L" + row_str].value == "x" and sheet["G" + row_str].value == "Fail":
#             case_fail = sheet["A" + row_str].value
#             event_name = sheet["B" + row_str].value
#             account = sheet["C" + row_str].value
#             step = sheet["D" + row_str].value
#             desired_result = sheet["E" + row_str].value
#             actual_result = sheet["F" + row_str].value
#             status = sheet["G" + row_str].value
#             phone_tag = sheet["N" + row_str].value
#             if phone_tag and '(' in phone_tag and ')' in phone_tag:
#                 phone, tag = phone_tag.split('(')
#                 tag = tag.rstrip(')')
#                 phone = phone.strip()
#                 print(phone)
#                 print(tag + "\n")
#
#                 # Nhóm theo tag
#                 grouped_by_tag[tag].append({
#                     'case_fail': case_fail,
#                     'event_name': event_name,
#                     'account': account,
#                     'step': step,
#                     'desired_result': desired_result,
#                     'actual_result': actual_result,
#                     'status': status,
#                     'phone': phone})
#                 list_casefail.append(case_fail)
#     print(list_casefail)
#     print("Số case fail nghiêm trọng: ", len(list_casefail))
#
#     # In kết quả nhóm theo tag
#     print("\n====== NHÓM THEO TAG ======")
#     n = 89
#     for tag, cases in grouped_by_tag.items():
#         print(f"\nTag: {tag} - {len(cases)} case - ")
#         n+= 1
#         var.writeData(var.path_luutamthoi, "Sheet1", n, 2, tag)
#         var.writeData(var.path_luutamthoi, "Sheet1", n, 4, len(cases))
#         for case in cases:
#             message_bug = (f" -Mã: {case['case_fail']}\n"
#                   f" -Tên sự kiện: {case['event_name']}\n"
#                   f" -Tài khoản: {case['account']}\n"
#                   f" -Các bước thao tác: {case['step']}\n"
#                   f" -Kết quả mong muốn: {case['desired_result']}\n"
#                   f" -Kết quả thực tế: {case['actual_result']}\n"
#                   f" -Trạng thái: {case['status']}\n---------------------------------------------------------\n")
#             var.writeData_append(var.path_luutamthoi, "Sheet1", n, 6, f"{case['case_fail']}, ")
#             var.writeData(var.path_luutamthoi, "Sheet1", n, 3, case['phone'])
#             var.writeData_append(var.path_luutamthoi, "Sheet1", n, 5, message_bug)
#             print(message_bug)



def retest_serious():
    var.writeData(var.path_luutamthoi, "Sheet1", 90, 2, "")
    var.writeData(var.path_luutamthoi, "Sheet1", 90, 3, "")
    var.writeData(var.path_luutamthoi, "Sheet1", 91, 2, 0)
    var.writeData(var.path_luutamthoi, "Sheet1", 92, 2, "")

    list_casefail = []
    grouped_by_casefail = defaultdict(list)  # Nhóm theo case_fail

    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while rownum < 1200:
        rownum += 1
        row_str = str(rownum)

        if sheet["L" + row_str].value == "x" and sheet["G" + row_str].value == "Fail":
            case_fail = sheet["A" + row_str].value
            event_name = sheet["B" + row_str].value
            account = sheet["C" + row_str].value
            step = sheet["D" + row_str].value
            desired_result = sheet["E" + row_str].value
            actual_result = sheet["F" + row_str].value
            status = sheet["G" + row_str].value

            grouped_by_casefail[case_fail].append({
                'event_name': event_name,
                'account': account,
                'step': step,
                'desired_result': desired_result,
                'actual_result': actual_result,
                'status': status
            })

            list_casefail.append(case_fail)

    print(list_casefail)
    print("Số case fail nghiêm trọng: ", len(list_casefail))

    print("\n====== NHÓM THEO CASE FAIL ======")
    for case_fail, items in grouped_by_casefail.items():
        print(f"\nCase Fail: {case_fail} - Số dòng liên quan: {len(items)}")

        # Gom tên case_fail (chỉ 1 vì mỗi nhóm là duy nhất)
        casefail_str = case_fail

        # Gom tất cả các bước step lại
        all_steps = "\n".join([item['step'] for item in items if item['step']])

        # Gộp nội dung chi tiết lỗi
        message_bug = ""
        for item in items:
            message_bug += (f" -Mã: {case_fail}\n"
                            f" -Tên sự kiện: {item['event_name']}\n"
                            f" -Tài khoản: {item['account']}\n"
                            f" -Các bước thao tác: {item['step']}\n"
                            f" -Kết quả mong muốn: {item['desired_result']}\n"
                            f" -Kết quả thực tế: {item['actual_result']}\n"
                            f" -Trạng thái: {item['status']}\n"
                            f"---------------------------------------------------------\n")
            var.writeData_append(var.path_luutamthoi, "Sheet1", 90, 3, f"{case_fail} {item['event_name']}, \n")


        # Ghi vào Excel
        var.writeData_append(var.path_luutamthoi, "Sheet1", 90, 2, f"{casefail_str}, ")  # Tên case_fail

        count_casefail = str(var.readData(var.path_luutamthoi, 'Sheet1', 90, 2))
        parts = count_casefail.split(',')
        cleaned = [p.strip() for p in parts if p.strip()]
        count = len(cleaned)

        var.writeData(var.path_luutamthoi, "Sheet1", 91, 2, count)  # Số dòng liên quan
        var.writeData_append(var.path_luutamthoi, "Sheet1", 92, 2, message_bug)  # Ghi toàn bộ message
        print(message_bug)








def run_all(self):
    var.driver.set_page_load_timeout(20)
    list_0 = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        print(sheet["G" + rownum].value)
        print(sheet["H" + rownum].value)
        if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[ "J" + rownum].value == "x"
            or sheet["K" + rownum].value == "x") and sheet["A" + rownum].value != None:
            print(sheet["A" + rownum].value)
            case_0 = sheet["A" + rownum].value
            list_0.append(case_0)
        rownum = int(rownum)
    print(list_0)
    for case in list_0:
        try:
            if case == 'Login01':
                caseid.caseid_login01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login02':
                caseid.caseid_login02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login03':
                caseid.caseid_login03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login04':
                caseid.caseid_login04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login05':
                caseid.caseid_login05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login08':
                caseid.caseid_login08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login09':
                caseid.caseid_login09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login10':
                caseid.caseid_login10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login11':
                caseid.caseid_login11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login12':
                caseid.caseid_login12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login13':
                caseid.caseid_login13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login14':
                caseid.caseid_login14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login15':
                caseid.caseid_login15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login16':
                caseid.caseid_login16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login17':
                caseid.caseid_login17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login18':
                caseid.caseid_login18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login19':
                caseid.caseid_login19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login20':
                caseid.caseid_login20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Login21':
                caseid.caseid_login21(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'GiamSat01':
                caseid.caseid_giamsat01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat02':
                caseid.caseid_giamsat02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat03':
                caseid.caseid_giamsat03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat04':
                caseid.caseid_giamsat04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat05':
                caseid.caseid_giamsat05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat06':
                caseid.caseid_giamsat06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat07':
                caseid.caseid_giamsat07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat08':
                caseid.caseid_giamsat08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat09':
                caseid.caseid_giamsat09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat10':
                caseid.caseid_giamsat10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat11':
                caseid.caseid_giamsat11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat12':
                caseid.caseid_giamsat12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat13':
                caseid.caseid_giamsat13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat14':
                caseid.caseid_giamsat14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat15':
                caseid.caseid_giamsat15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat16':
                caseid.caseid_giamsat16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat17':
                caseid.caseid_giamsat17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat18':
                caseid.caseid_giamsat18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat19':
                caseid.caseid_giamsat19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat20':
                caseid.caseid_giamsat20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat21':
                caseid.caseid_giamsat21(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat22':
                caseid.caseid_giamsat22(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat23':
                caseid.caseid_giamsat23(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat24':
                caseid.caseid_giamsat24(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat25':
                caseid.caseid_giamsat25(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat26':
                caseid.caseid_giamsat26(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat27':
                caseid.caseid_giamsat27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat28':
                caseid.caseid_giamsat28(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat29':
                caseid.caseid_giamsat29(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat30':
                caseid.caseid_giamsat30(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat31':
                caseid.caseid_giamsat31(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat32':
                caseid.caseid_giamsat32(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat33':
                caseid.caseid_giamsat33(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat34':
                caseid.caseid_giamsat34(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat35':
                caseid.caseid_giamsat35(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat36':
                caseid.caseid_giamsat36(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat37':
                caseid.caseid_giamsat37(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat38':
                caseid.caseid_giamsat38(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat39':
                caseid.caseid_giamsat39(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat40':
                caseid.caseid_giamsat40(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat41':
                caseid.caseid_giamsat41(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat42':
                caseid.caseid_giamsat42(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat43':
                caseid.caseid_giamsat43(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat44':
                caseid.caseid_giamsat44(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat45':
                caseid.caseid_giamsat45(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat46':
                caseid.caseid_giamsat46(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat47':
                caseid.caseid_giamsat47(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat48':
                caseid.caseid_giamsat48(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat49':
                caseid.caseid_giamsat49(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat50':
                caseid.caseid_giamsat50(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat51':
                caseid.caseid_giamsat51(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat52':
                caseid.caseid_giamsat52(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat53':
                caseid.caseid_giamsat53(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat54':
                caseid.caseid_giamsat54(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat55':
                caseid.caseid_giamsat55(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat56':
                caseid.caseid_giamsat56(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat57':
                caseid.caseid_giamsat57(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat58':
                caseid.caseid_giamsat58(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat59':
                caseid.caseid_giamsat59(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat60':
                caseid.caseid_giamsat60(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat61':
                caseid.caseid_giamsat61(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat62':
                caseid.caseid_giamsat62(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat63':
                caseid.caseid_giamsat63(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat64':
                caseid.caseid_giamsat64(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat65':
                caseid.caseid_giamsat65(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat66':
                caseid.caseid_giamsat66(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat67':
                caseid.caseid_giamsat67(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat68':
                caseid.caseid_giamsat68(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat69':
                caseid.caseid_giamsat69(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat70':
                caseid.caseid_giamsat70(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat71':
                caseid.caseid_giamsat71(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat72':
                caseid.caseid_giamsat72(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat73':
                caseid.caseid_giamsat73(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat74':
                caseid.caseid_giamsat74(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat75':
                caseid.caseid_giamsat75(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat76':
                caseid.caseid_giamsat76(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat77':
                caseid.caseid_giamsat77(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat78':
                caseid.caseid_giamsat78(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat79':
                caseid.caseid_giamsat79(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat80':
                caseid.caseid_giamsat80(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'GiamSat80_1':
                caseid.caseid_giamsat80_1(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'GiamSat81':
                caseid.caseid_giamsat81(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat82':
                caseid.caseid_giamsat82(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat83':
                caseid.caseid_giamsat83(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat84':
                caseid.caseid_giamsat84(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat85':
                caseid.caseid_giamsat85(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat86':
                caseid.caseid_giamsat86(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat87':
                caseid.caseid_giamsat87(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat88':
                caseid.caseid_giamsat88(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat89':
                caseid.caseid_giamsat89(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat90':
                caseid.caseid_giamsat90(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat91':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat92':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat93':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat94':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat95':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat96':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat97':
                caseid.caseid_giamsat91(self)
                caseid.caseid_giamsat92(self)
                caseid.caseid_giamsat93(self)
                caseid.caseid_giamsat94(self)
                caseid.caseid_giamsat95(self)
                caseid.caseid_giamsat96(self)
                caseid.caseid_giamsat97(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat98':
                caseid.caseid_giamsat98(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat99':
                caseid.caseid_giamsat99(self)
                caseid.caseid_giamsat100(self)
                caseid.caseid_giamsat101(self)
                caseid.caseid_giamsat102(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat100':
                caseid.caseid_giamsat99(self)
                caseid.caseid_giamsat100(self)
                caseid.caseid_giamsat101(self)
                caseid.caseid_giamsat102(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat101':
                caseid.caseid_giamsat99(self)
                caseid.caseid_giamsat100(self)
                caseid.caseid_giamsat101(self)
                caseid.caseid_giamsat102(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat102':
                caseid.caseid_giamsat99(self)
                caseid.caseid_giamsat100(self)
                caseid.caseid_giamsat101(self)
                caseid.caseid_giamsat102(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat103':
                caseid.caseid_giamsat103(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat104':
                caseid.caseid_giamsat103(self)
                caseid.caseid_giamsat104(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat105':
                caseid.caseid_giamsat105(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat106':
                caseid.caseid_giamsat106(self)
                caseid.caseid_giamsat107(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat107':
                caseid.caseid_giamsat106(self)
                caseid.caseid_giamsat107(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat108':
                caseid.caseid_giamsat108(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat109':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat110':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat111':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat112':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat113':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat114':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat115':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat116':
                caseid.caseid_giamsat108(self)
                caseid.caseid_giamsat109(self)
                caseid.caseid_giamsat110(self)
                caseid.caseid_giamsat111(self)
                caseid.caseid_giamsat112(self)
                caseid.caseid_giamsat113(self)
                caseid.caseid_giamsat114(self)
                caseid.caseid_giamsat115(self)
                caseid.caseid_giamsat116(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat117':
                caseid.caseid_giamsat117(self)
                caseid.caseid_giamsat118(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat118':
                caseid.caseid_giamsat117(self)
                caseid.caseid_giamsat118(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat119':
                caseid.caseid_giamsat119(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat120':
                caseid.caseid_giamsat120(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat121':
                caseid.caseid_giamsat121(self)
                caseid.caseid_giamsat122(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat122':
                caseid.caseid_giamsat121(self)
                caseid.caseid_giamsat122(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat123':
                caseid.caseid_giamsat123(self)
                caseid.caseid_giamsat124(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat124':
                caseid.caseid_giamsat123(self)
                caseid.caseid_giamsat124(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat125':
                caseid.caseid_giamsat125(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat126':
                caseid.caseid_giamsat126(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat127':
                caseid.caseid_giamsat127(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat128':
                caseid.caseid_giamsat128(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat129':
                caseid.caseid_giamsat129(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat130':
                caseid.caseid_giamsat130(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat131':
                caseid.caseid_giamsat131(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat132':
                caseid.caseid_giamsat132(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat133':
                caseid.caseid_giamsat133(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat134':
                caseid.caseid_giamsat134(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat135':
                caseid.caseid_giamsat135(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat136':
                caseid.caseid_giamsat136(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat137':
                caseid.caseid_giamsat137(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat138':
                caseid.caseid_giamsat138(self)
                caseid.caseid_giamsat139(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat139':
                caseid.caseid_giamsat138(self)
                caseid.caseid_giamsat139(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat140':
                caseid.caseid_giamsat140(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat143':
                caseid.caseid_giamsat143(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat144':
                caseid.caseid_giamsat144(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat145':
                caseid.caseid_giamsat145(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat146':
                caseid.caseid_giamsat146(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat147':
                caseid.caseid_giamsat147(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat148':
                caseid.caseid_giamsat148(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat149':
                caseid.caseid_giamsat149(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat150':
                caseid.caseid_giamsat150(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat151':
                caseid.caseid_giamsat151(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat153':
                caseid.caseid_giamsat153(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat154':
                caseid.caseid_giamsat154(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat155':
                caseid.caseid_giamsat155(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat156':
                caseid.caseid_giamsat156(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat157':
                caseid.caseid_giamsat157(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat158':
                caseid.caseid_giamsat158(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat159':
                caseid.caseid_giamsat159(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat160':
                caseid.caseid_giamsat160(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat161':
                caseid.caseid_giamsat161(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat162':
                caseid.caseid_giamsat162(self)
                caseid.caseid_giamsat163(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat163':
                caseid.caseid_giamsat162(self)
                caseid.caseid_giamsat163(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat164':
                caseid.caseid_giamsat164(self)
                caseid.caseid_giamsat165(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat165':
                caseid.caseid_giamsat164(self)
                caseid.caseid_giamsat165(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'GiamSat165_1':
                caseid.caseid_giamsat165_1(self)
                caseid.caseid_giamsat165_2(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'GiamSat165_2':
                caseid.caseid_giamsat165_1(self)
                caseid.caseid_giamsat165_2(self)
        except:
            chucnangkhac.swich_tab_0()



        try:
            if case == 'GiamSat166':
                caseid.caseid_giamsat166(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat167':
                caseid.caseid_giamsat167(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat168':
                caseid.caseid_giamsat168(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat169':
                caseid.caseid_giamsat169(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat170':
                caseid.caseid_giamsat170(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat171':
                caseid.caseid_giamsat171(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat172':
                caseid.caseid_giamsat172(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat173':
                caseid.caseid_giamsat173(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat174':
                caseid.caseid_giamsat174(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat175':
                caseid.caseid_giamsat175(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat176':
                caseid.caseid_giamsat176(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat177':
                caseid.caseid_giamsat177(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat178':
                caseid.caseid_giamsat178(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat179':
                caseid.caseid_giamsat179(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat180':
                caseid.caseid_giamsat180(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat181':
                caseid.caseid_giamsat181(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat182':
                caseid.caseid_giamsat182(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat183':
                caseid.caseid_giamsat183(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat184':
                caseid.caseid_giamsat184(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat185':
                caseid.caseid_giamsat185(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat186':
                caseid.caseid_giamsat186(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat187':
                caseid.caseid_giamsat187(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat188':
                caseid.caseid_giamsat188(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat189':
                caseid.caseid_giamsat189(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat190':
                caseid.caseid_giamsat190(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat191':
                caseid.caseid_giamsat191(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat192':
                caseid.caseid_giamsat192(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat193':
                caseid.caseid_giamsat193(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat194':
                caseid.caseid_giamsat194(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat195':
                caseid.caseid_giamsat195(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat196':
                caseid.caseid_giamsat196(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat197':
                caseid.caseid_giamsat197(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat198':
                caseid.caseid_giamsat198(self)
                caseid.caseid_giamsat201(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat199':
                caseid.caseid_giamsat198(self)
                caseid.caseid_giamsat199(self)
                caseid.caseid_giamsat200(self)
                caseid.caseid_giamsat201(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat200':
                caseid.caseid_giamsat198(self)
                caseid.caseid_giamsat199(self)
                caseid.caseid_giamsat200(self)
                caseid.caseid_giamsat201(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat201':
                caseid.caseid_giamsat198(self)
                caseid.caseid_giamsat201(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat202':
                caseid.caseid_giamsat202(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat203':
                caseid.caseid_giamsat203(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat204':
                caseid.caseid_giamsat204(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat205':
                caseid.caseid_giamsat205(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat206':
                caseid.caseid_giamsat206(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat207':
                caseid.caseid_giamsat207(self)
                caseid.caseid_giamsat203(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat209':
                caseid.caseid_giamsat209(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat210':
                caseid.caseid_giamsat210(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat211':
                caseid.caseid_giamsat211(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat212':
                caseid.caseid_giamsat212(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat213':
                caseid.caseid_giamsat213(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat214':
                caseid.caseid_giamsat214(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat215':
                caseid.caseid_giamsat215(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat216':
                caseid.caseid_giamsat216(self)
                caseid.caseid_giamsat217(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat217':
                caseid.caseid_giamsat217(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat218':
                caseid.caseid_giamsat218(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat219':
                caseid.caseid_giamsat219(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat220':
                caseid.caseid_giamsat220(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat221':
                caseid.caseid_giamsat221(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat222':
                caseid.caseid_giamsat222(self)
                caseid.caseid_giamsat223(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat223':
                caseid.caseid_giamsat222(self)
                caseid.caseid_giamsat223(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat224':
                caseid.caseid_giamsat224(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat225':
                caseid.caseid_giamsat225(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat226':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat227':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat228':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat229':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat230':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat231':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat232':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat233':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat234':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat235':
                caseid.caseid_giamsat224(self)
                caseid.caseid_giamsat226(self)
                caseid.caseid_giamsat227(self)
                caseid.caseid_giamsat228(self)
                caseid.caseid_giamsat229(self)
                caseid.caseid_giamsat230(self)
                caseid.caseid_giamsat231(self)
                caseid.caseid_giamsat232(self)
                caseid.caseid_giamsat233(self)
                caseid.caseid_giamsat234(self)
                caseid.caseid_giamsat235(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat236':
                caseid.caseid_giamsat236(self)
                caseid.caseid_giamsat237(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat237':
                caseid.caseid_giamsat237(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat238':
                caseid.caseid_giamsat238(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat239':
                caseid.caseid_giamsat239(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat241':
                caseid.caseid_giamsat241(self)
                caseid.caseid_giamsat242(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat242':
                caseid.caseid_giamsat242(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat243':
                caseid.caseid_giamsat243(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat244':
                caseid.caseid_giamsat244(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat245':
                caseid.caseid_giamsat245(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat246':
                caseid.caseid_giamsat246(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat247':
                caseid.caseid_giamsat247(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat248':
                caseid.caseid_giamsat248(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat249':
                caseid.caseid_giamsat249(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat250':
                caseid.caseid_giamsat250(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat251':
                caseid.caseid_giamsat251(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat252':
                caseid.caseid_giamsat252(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat253':
                caseid.caseid_giamsat253(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat254':
                caseid.caseid_giamsat254(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat255':
                caseid.caseid_giamsat255(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat256':
                caseid.caseid_giamsat256(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat258':
                caseid.caseid_giamsat258(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat259':
                caseid.caseid_giamsat259(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'GiamSat260':
                caseid.caseid_giamsat260(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'Route01':
                caseid.caseid_route01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route02':
                caseid.caseid_route02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route03':
                caseid.caseid_route03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route04':
                caseid.caseid_route04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route05':
                caseid.caseid_route05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route06':
                caseid.caseid_route06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route07':
                caseid.caseid_route07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route08':
                caseid.caseid_route08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route09':
                caseid.caseid_route09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route10':
                caseid.caseid_route10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route11':
                caseid.caseid_route11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route12':
                caseid.caseid_route12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route13':
                caseid.caseid_route13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route14':
                caseid.caseid_route14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route15':
                caseid.caseid_route15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route16':
                caseid.caseid_route16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route17':
                caseid.caseid_route17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route18':
                caseid.caseid_route18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route19':
                caseid.caseid_route19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route20':
                caseid.caseid_route20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route21':
                caseid.caseid_route21(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route22':
                caseid.caseid_route22(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Route23':
                caseid.caseid_route23(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'Admin01':
                caseid.caseid_admin01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin02':
                caseid.caseid_admin02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin03':
                caseid.caseid_admin03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin04':
                caseid.caseid_admin04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin05':
                caseid.caseid_admin05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin06':
                caseid.caseid_admin06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin07':
                caseid.caseid_admin07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin08':
                caseid.caseid_admin08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin09':
                caseid.caseid_admin09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin10':
                caseid.caseid_admin10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin11':
                caseid.caseid_admin11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin12':
                caseid.caseid_admin12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin13':
                caseid.caseid_admin13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin14':
                caseid.caseid_admin14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin15':
                caseid.caseid_admin15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin16':
                caseid.caseid_admin16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin17':
                caseid.caseid_admin17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin18':
                caseid.caseid_admin18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin19':
                caseid.caseid_admin19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin20':
                caseid.caseid_admin20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin21':
                caseid.caseid_admin21(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin22':
                caseid.caseid_admin22(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin23':
                caseid.caseid_admin22(self)
                caseid.caseid_admin23(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin24':
                caseid.caseid_admin22(self)
                caseid.caseid_admin23(self)
                caseid.caseid_admin24(self)
                caseid.caseid_admin25(self)
                caseid.caseid_admin26(self)
                caseid.caseid_admin27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin25':
                caseid.caseid_admin22(self)
                caseid.caseid_admin23(self)
                caseid.caseid_admin24(self)
                caseid.caseid_admin25(self)
                caseid.caseid_admin26(self)
                caseid.caseid_admin27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin26':
                caseid.caseid_admin22(self)
                caseid.caseid_admin23(self)
                caseid.caseid_admin24(self)
                caseid.caseid_admin25(self)
                caseid.caseid_admin26(self)
                caseid.caseid_admin27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin27':
                caseid.caseid_admin22(self)
                caseid.caseid_admin23(self)
                caseid.caseid_admin24(self)
                caseid.caseid_admin25(self)
                caseid.caseid_admin26(self)
                caseid.caseid_admin27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin28':
                caseid.caseid_admin28(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin29':
                caseid.caseid_admin29(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin30':
                caseid.caseid_admin30(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin31':
                caseid.caseid_admin31(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin32':
                caseid.caseid_admin32(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin33':
                caseid.caseid_admin33(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin34':
                caseid.caseid_admin34(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Admin35':
                caseid.caseid_admin35(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User01':
                caseid.caseid_user01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User02':
                caseid.caseid_user02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User03':
                caseid.caseid_user03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User04':
                caseid.caseid_user04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User05':
                caseid.caseid_user05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User06':
                caseid.caseid_user06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User07':
                caseid.caseid_user07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User08':
                caseid.caseid_user08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User09':
                caseid.caseid_user09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User10':
                caseid.caseid_user10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User11':
                caseid.caseid_user11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User12':
                caseid.caseid_user12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User12':
                caseid.caseid_user12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User13':
                caseid.caseid_user13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User14':
                caseid.caseid_user14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User15':
                caseid.caseid_user15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User16':
                caseid.caseid_user16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'User17':
                caseid.caseid_user17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report01':
                caseid.caseid_report01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report02':
                caseid.caseid_report02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report03':
                caseid.caseid_report03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report04':
                caseid.caseid_report04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report05':
                caseid.caseid_report05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report06':
                caseid.caseid_report06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report07':
                caseid.caseid_report07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report08':
                caseid.caseid_report08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report09':
                caseid.caseid_report09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report10':
                caseid.caseid_report10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report11':
                caseid.caseid_report11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report12':
                caseid.caseid_report12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report13':
                caseid.caseid_report13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report14':
                caseid.caseid_report14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report15':
                caseid.caseid_report15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report16':
                caseid.caseid_report16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report17':
                caseid.caseid_report17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report18':
                caseid.caseid_report18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report19':
                caseid.caseid_report19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report20':
                caseid.caseid_report20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report21':
                caseid.caseid_report21(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report22':
                caseid.caseid_report22(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report23':
                caseid.caseid_report23(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report24':
                caseid.caseid_report24(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report25':
                caseid.caseid_report25(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report26':
                caseid.caseid_report26(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report27':
                caseid.caseid_report27(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report28':
                caseid.caseid_report28(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report29':
                caseid.caseid_report29(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report30':
                caseid.caseid_report30(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report31':
                caseid.caseid_report31(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report32':
                caseid.caseid_report32(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report33':
                caseid.caseid_report33(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report34':
                caseid.caseid_report34(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report35':
                caseid.caseid_report35(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report36':
                caseid.caseid_report36(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report37':
                caseid.caseid_report37(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report38':
                caseid.caseid_report38(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report39':
                caseid.caseid_report39(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report40':
                caseid.caseid_report40(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report41':
                caseid.caseid_report41(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report42':
                caseid.caseid_report42(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report43':
                caseid.caseid_report43(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report44':
                caseid.caseid_report44(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report45':
                caseid.caseid_report45(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report46':
                caseid.caseid_report46(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report47':
                caseid.caseid_report47(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report48':
                caseid.caseid_report48(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report49':
                caseid.caseid_report49(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report50':
                caseid.caseid_report50(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report51':
                caseid.caseid_report51(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report52':
                caseid.caseid_report52(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report53':
                caseid.caseid_report53(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report54':
                caseid.caseid_report54(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report55':
                caseid.caseid_report55(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report56':
                caseid.caseid_report56(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report57':
                caseid.caseid_report57(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report58':
                caseid.caseid_report58(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report59':
                caseid.caseid_report59(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report60':
                caseid.caseid_report60(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report61':
                caseid.caseid_report61(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report62':
                caseid.caseid_report62(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report63':
                caseid.caseid_report63(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report64':
                caseid.caseid_report64(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report65':
                caseid.caseid_report65(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report66':
                caseid.caseid_report66(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report67':
                caseid.caseid_report67(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report68':
                caseid.caseid_report68(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report69':
                caseid.caseid_report69(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Report70':
                caseid.caseid_report70(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'Video01':
                caseid.caseid_video01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video02':
                caseid.caseid_video02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video03':
                caseid.caseid_video03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video04':
                caseid.caseid_video04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video05':
                caseid.caseid_video05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video06':
                caseid.caseid_video06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video07':
                caseid.caseid_video07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video08':
                caseid.caseid_video08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video09':
                caseid.caseid_video09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video10':
                caseid.caseid_video10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video11':
                caseid.caseid_video11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Video12':
                caseid.caseid_video12(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'Image01':
                caseid.caseid_image01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image02':
                caseid.caseid_image02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image03':
                caseid.caseid_image03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image04':
                caseid.caseid_image04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image05':
                caseid.caseid_image05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image06':
                caseid.caseid_image06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image07':
                caseid.caseid_image07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image08':
                caseid.caseid_image08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image09':
                caseid.caseid_image09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image10':
                caseid.caseid_image10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image11':
                caseid.caseid_image11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image12':
                caseid.caseid_image12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image13':
                caseid.caseid_image13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image14':
                caseid.caseid_image14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image15':
                caseid.caseid_image15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image16':
                caseid.caseid_image16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image17':
                caseid.caseid_image17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image18':
                caseid.caseid_image18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image19':
                caseid.caseid_image19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Image20':
                caseid.caseid_image20(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'Utility01':
                caseid.caseid_utility01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility02':
                caseid.caseid_utility02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility03':
                caseid.caseid_utility03(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility04':
                caseid.caseid_utility04(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility05':
                caseid.caseid_utility05(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility06':
                caseid.caseid_utility06(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility07':
                caseid.caseid_utility07(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility08':
                caseid.caseid_utility08(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility09':
                caseid.caseid_utility09(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility10':
                caseid.caseid_utility10(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility11':
                caseid.caseid_utility11(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility12':
                caseid.caseid_utility12(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility13':
                caseid.caseid_utility13(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility14':
                caseid.caseid_utility14(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility15':
                caseid.caseid_utility15(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility16':
                caseid.caseid_utility16(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility17':
                caseid.caseid_utility17(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility18':
                caseid.caseid_utility18(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility19':
                caseid.caseid_utility19(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility20':
                caseid.caseid_utility20(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility21':
                caseid.caseid_utility21(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility22':
                caseid.caseid_utility22(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Utility23':
                caseid.caseid_utility23(self)
        except:
            chucnangkhac.swich_tab_0()

        try:
            if case == 'Ai01':
                caseid.caseid_ai01(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Ai02':
                caseid.caseid_ai02(self)
        except:
            chucnangkhac.swich_tab_0()
        try:
            if case == 'Ai03':
                caseid.caseid_ai03(self)
        except:
            chucnangkhac.swich_tab_0()



def login(self):
    var.driver.set_page_load_timeout(30)
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 10
    while (rownum < 35):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("mcu do 1", list_mucdo1)
    print("mcu do 2", list_mucdo2)
    print("mcu do 3", list_mucdo3)
    print("mcu do 4", list_mucdo4)


    modetest = ''.join(re.findall(r'\d+', var.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Login01':
                        caseid.caseid_login01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid.caseid_login02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid.caseid_login03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid.caseid_login04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid.caseid_login05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid.caseid_login08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid.caseid_login09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid.caseid_login10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid.caseid_login11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid.caseid_login12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login13':
                        caseid.caseid_login13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login14':
                        caseid.caseid_login14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login15':
                        caseid.caseid_login15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login16':
                        caseid.caseid_login16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login17':
                        caseid.caseid_login17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login18':
                        caseid.caseid_login18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login19':
                        caseid.caseid_login19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login20':
                        caseid.caseid_login20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login21':
                        caseid.caseid_login21(self)
                except:
                    chucnangkhac.swich_tab_0()
        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Login01':
                        caseid.caseid_login01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid.caseid_login02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid.caseid_login03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid.caseid_login04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid.caseid_login05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid.caseid_login08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid.caseid_login09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid.caseid_login10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid.caseid_login11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid.caseid_login12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login13':
                        caseid.caseid_login13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login14':
                        caseid.caseid_login14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login15':
                        caseid.caseid_login15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login16':
                        caseid.caseid_login16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login17':
                        caseid.caseid_login17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login18':
                        caseid.caseid_login18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login19':
                        caseid.caseid_login19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login20':
                        caseid.caseid_login20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login21':
                        caseid.caseid_login21(self)
                except:
                    chucnangkhac.swich_tab_0()
        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Login01':
                        caseid.caseid_login01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid.caseid_login02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid.caseid_login03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid.caseid_login04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid.caseid_login05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid.caseid_login08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid.caseid_login09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid.caseid_login10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid.caseid_login11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid.caseid_login12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login13':
                        caseid.caseid_login13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login14':
                        caseid.caseid_login14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login15':
                        caseid.caseid_login15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login16':
                        caseid.caseid_login16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login17':
                        caseid.caseid_login17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login18':
                        caseid.caseid_login18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login19':
                        caseid.caseid_login19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login20':
                        caseid.caseid_login20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login21':
                        caseid.caseid_login21(self)
                except:
                    chucnangkhac.swich_tab_0()
        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Login01':
                        caseid.caseid_login01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid.caseid_login02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid.caseid_login03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid.caseid_login04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid.caseid_login05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid.caseid_login08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid.caseid_login09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid.caseid_login10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid.caseid_login11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid.caseid_login12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login13':
                        caseid.caseid_login13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login14':
                        caseid.caseid_login14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login15':
                        caseid.caseid_login15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login16':
                        caseid.caseid_login16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login17':
                        caseid.caseid_login17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login18':
                        caseid.caseid_login18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login19':
                        caseid.caseid_login19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login20':
                        caseid.caseid_login20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Login21':
                        caseid.caseid_login21(self)
                except:
                    chucnangkhac.swich_tab_0()        #0



#2 giám sát
def monitor(self):
    var.driver.set_page_load_timeout(20)
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 36
    while (rownum < 429):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)

    modetest = ''.join(re.findall(r'\d+', var.modetest))
    for i in modetest:
        if i == "1":
            print(list_mucdo1)
            for case in list_mucdo1:
                try:
                    if case == 'GiamSat01':
                        caseid.caseid_giamsat01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat02':
                        caseid.caseid_giamsat02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat03':
                        caseid.caseid_giamsat03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat04':
                        caseid.caseid_giamsat04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat05':
                        caseid.caseid_giamsat05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat06':
                        caseid.caseid_giamsat06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat07':
                        caseid.caseid_giamsat07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat08':
                        caseid.caseid_giamsat08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat09':
                        caseid.caseid_giamsat09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat10':
                        caseid.caseid_giamsat10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat11':
                        caseid.caseid_giamsat11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat12':
                        caseid.caseid_giamsat12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat13':
                        caseid.caseid_giamsat13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat14':
                        caseid.caseid_giamsat14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat15':
                        caseid.caseid_giamsat15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat16':
                        caseid.caseid_giamsat16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat17':
                        caseid.caseid_giamsat17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat18':
                        caseid.caseid_giamsat18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat19':
                        caseid.caseid_giamsat19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat20':
                        caseid.caseid_giamsat20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat21':
                        caseid.caseid_giamsat21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat22':
                        caseid.caseid_giamsat22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat23':
                        caseid.caseid_giamsat23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat24':
                        caseid.caseid_giamsat24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat25':
                        caseid.caseid_giamsat25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat26':
                        caseid.caseid_giamsat26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat27':
                        caseid.caseid_giamsat27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat28':
                        caseid.caseid_giamsat28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat29':
                        caseid.caseid_giamsat29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat30':
                        caseid.caseid_giamsat30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat31':
                        caseid.caseid_giamsat31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat32':
                        caseid.caseid_giamsat32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat33':
                        caseid.caseid_giamsat33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat34':
                        caseid.caseid_giamsat34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat35':
                        caseid.caseid_giamsat35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat36':
                        caseid.caseid_giamsat36(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat37':
                        caseid.caseid_giamsat37(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat38':
                        caseid.caseid_giamsat38(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat39':
                        caseid.caseid_giamsat39(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat40':
                        caseid.caseid_giamsat40(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat41':
                        caseid.caseid_giamsat41(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat42':
                        caseid.caseid_giamsat42(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat43':
                        caseid.caseid_giamsat43(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat44':
                        caseid.caseid_giamsat44(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat45':
                        caseid.caseid_giamsat45(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat46':
                        caseid.caseid_giamsat46(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat47':
                        caseid.caseid_giamsat47(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat48':
                        caseid.caseid_giamsat48(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat49':
                        caseid.caseid_giamsat49(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat50':
                        caseid.caseid_giamsat50(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat51':
                        caseid.caseid_giamsat51(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat52':
                        caseid.caseid_giamsat52(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat53':
                        caseid.caseid_giamsat53(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat54':
                        caseid.caseid_giamsat54(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat55':
                        caseid.caseid_giamsat55(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat56':
                        caseid.caseid_giamsat56(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat57':
                        caseid.caseid_giamsat57(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat58':
                        caseid.caseid_giamsat58(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat59':
                        caseid.caseid_giamsat59(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat60':
                        caseid.caseid_giamsat60(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat61':
                        caseid.caseid_giamsat61(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat62':
                        caseid.caseid_giamsat62(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat63':
                        caseid.caseid_giamsat63(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat64':
                        caseid.caseid_giamsat64(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat65':
                        caseid.caseid_giamsat65(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat66':
                        caseid.caseid_giamsat66(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat67':
                        caseid.caseid_giamsat67(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat68':
                        caseid.caseid_giamsat68(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat69':
                        caseid.caseid_giamsat69(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat70':
                        caseid.caseid_giamsat70(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat71':
                        caseid.caseid_giamsat71(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat72':
                        caseid.caseid_giamsat72(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat73':
                        caseid.caseid_giamsat73(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat74':
                        caseid.caseid_giamsat74(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat75':
                        caseid.caseid_giamsat75(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat76':
                        caseid.caseid_giamsat76(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat77':
                        caseid.caseid_giamsat77(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat78':
                        caseid.caseid_giamsat78(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat79':
                        caseid.caseid_giamsat79(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat80':
                        caseid.caseid_giamsat80(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat80_1':
                        caseid.caseid_giamsat80_1(self)
                except:
                    chucnangkhac.swich_tab_0()


                try:
                    if case == 'GiamSat81':
                        caseid.caseid_giamsat81(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat82':
                        caseid.caseid_giamsat82(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat83':
                        caseid.caseid_giamsat83(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat84':
                        caseid.caseid_giamsat84(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat85':
                        caseid.caseid_giamsat85(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat86':
                        caseid.caseid_giamsat86(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat87':
                        caseid.caseid_giamsat87(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat88':
                        caseid.caseid_giamsat88(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat89':
                        caseid.caseid_giamsat89(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat90':
                        caseid.caseid_giamsat90(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat91':
                        caseid.caseid_giamsat91(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat92':
                        caseid.caseid_giamsat92(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat93':
                        caseid.caseid_giamsat93(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat94':
                        caseid.caseid_giamsat94(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat95':
                        caseid.caseid_giamsat95(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat96':
                        caseid.caseid_giamsat96(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat97':
                        caseid.caseid_giamsat97(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat98':
                        caseid.caseid_giamsat98(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat99':
                        caseid.caseid_giamsat99(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat100':
                        caseid.caseid_giamsat100(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat101':
                        caseid.caseid_giamsat101(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat102':
                        caseid.caseid_giamsat102(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat103':
                        caseid.caseid_giamsat103(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat104':
                        caseid.caseid_giamsat104(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat105':
                        caseid.caseid_giamsat105(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat106':
                        caseid.caseid_giamsat106(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat107':
                        caseid.caseid_giamsat107(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat108':
                        caseid.caseid_giamsat108(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat109':
                        caseid.caseid_giamsat109(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat110':
                        caseid.caseid_giamsat110(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat111':
                        caseid.caseid_giamsat111(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat112':
                        caseid.caseid_giamsat112(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat113':
                        caseid.caseid_giamsat113(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat114':
                        caseid.caseid_giamsat114(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat115':
                        caseid.caseid_giamsat115(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat116':
                        caseid.caseid_giamsat116(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat117':
                        caseid.caseid_giamsat117(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat118':
                        caseid.caseid_giamsat118(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat119':
                        caseid.caseid_giamsat119(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat120':
                        caseid.caseid_giamsat120(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat121':
                        caseid.caseid_giamsat121(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat122':
                        caseid.caseid_giamsat122(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat123':
                        caseid.caseid_giamsat123(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat124':
                        caseid.caseid_giamsat124(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat125':
                        caseid.caseid_giamsat125(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat126':
                        caseid.caseid_giamsat126(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat127':
                        caseid.caseid_giamsat127(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat128':
                        caseid.caseid_giamsat128(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat129':
                        caseid.caseid_giamsat129(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat130':
                        caseid.caseid_giamsat130(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat131':
                        caseid.caseid_giamsat131(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat132':
                        caseid.caseid_giamsat132(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat133':
                        caseid.caseid_giamsat133(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat134':
                        caseid.caseid_giamsat134(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat135':
                        caseid.caseid_giamsat135(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat136':
                        caseid.caseid_giamsat136(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat137':
                        caseid.caseid_giamsat137(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat138':
                        caseid.caseid_giamsat138(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat139':
                        caseid.caseid_giamsat139(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat140':
                        caseid.caseid_giamsat140(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat143':
                        caseid.caseid_giamsat143(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat144':
                        caseid.caseid_giamsat144(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat145':
                        caseid.caseid_giamsat145(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat146':
                        caseid.caseid_giamsat146(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat147':
                        caseid.caseid_giamsat147(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat148':
                        caseid.caseid_giamsat148(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat149':
                        caseid.caseid_giamsat149(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat150':
                        caseid.caseid_giamsat150(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat151':
                        caseid.caseid_giamsat151(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat153':
                        caseid.caseid_giamsat153(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat154':
                        caseid.caseid_giamsat154(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat155':
                        caseid.caseid_giamsat155(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat156':
                        caseid.caseid_giamsat156(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat157':
                        caseid.caseid_giamsat157(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat158':
                        caseid.caseid_giamsat158(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat159':
                        caseid.caseid_giamsat159(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat160':
                        caseid.caseid_giamsat160(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat161':
                        caseid.caseid_giamsat161(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat162':
                        caseid.caseid_giamsat162(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat163':
                        caseid.caseid_giamsat163(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat164':
                        caseid.caseid_giamsat164(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat165':
                        caseid.caseid_giamsat165(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat165_1':
                        caseid.caseid_giamsat165_1(self)
                        caseid.caseid_giamsat165_2(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat165_2':
                        caseid.caseid_giamsat165_1(self)
                        caseid.caseid_giamsat165_2(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat166':
                        caseid.caseid_giamsat166(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat167':
                        caseid.caseid_giamsat167(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat168':
                        caseid.caseid_giamsat168(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat169':
                        caseid.caseid_giamsat169(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat170':
                        caseid.caseid_giamsat170(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat171':
                        caseid.caseid_giamsat171(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat172':
                        caseid.caseid_giamsat172(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat173':
                        caseid.caseid_giamsat173(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat174':
                        caseid.caseid_giamsat174(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat175':
                        caseid.caseid_giamsat175(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat176':
                        caseid.caseid_giamsat176(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat177':
                        caseid.caseid_giamsat177(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat178':
                        caseid.caseid_giamsat178(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat179':
                        caseid.caseid_giamsat179(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat180':
                        caseid.caseid_giamsat180(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat181':
                        caseid.caseid_giamsat181(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat182':
                        caseid.caseid_giamsat182(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat183':
                        caseid.caseid_giamsat183(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat184':
                        caseid.caseid_giamsat184(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat185':
                        caseid.caseid_giamsat185(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat186':
                        caseid.caseid_giamsat186(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat187':
                        caseid.caseid_giamsat187(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat188':
                        caseid.caseid_giamsat188(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat189':
                        caseid.caseid_giamsat189(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat190':
                        caseid.caseid_giamsat190(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat191':
                        caseid.caseid_giamsat191(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat192':
                        caseid.caseid_giamsat192(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat193':
                        caseid.caseid_giamsat193(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat194':
                        caseid.caseid_giamsat194(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat195':
                        caseid.caseid_giamsat195(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat196':
                        caseid.caseid_giamsat196(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat197':
                        caseid.caseid_giamsat197(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat198':
                        caseid.caseid_giamsat198(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat199':
                        caseid.caseid_giamsat199(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat200':
                        caseid.caseid_giamsat200(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat201':
                        caseid.caseid_giamsat201(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat202':
                        caseid.caseid_giamsat202(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat203':
                        caseid.caseid_giamsat203(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat204':
                        caseid.caseid_giamsat204(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat205':
                        caseid.caseid_giamsat205(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat206':
                        caseid.caseid_giamsat206(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat207':
                        caseid.caseid_giamsat207(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat209':
                        caseid.caseid_giamsat209(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat210':
                        caseid.caseid_giamsat210(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat211':
                        caseid.caseid_giamsat211(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat212':
                        caseid.caseid_giamsat212(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat213':
                        caseid.caseid_giamsat213(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat214':
                        caseid.caseid_giamsat214(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat215':
                        caseid.caseid_giamsat215(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat216':
                        caseid.caseid_giamsat216(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat217':
                        caseid.caseid_giamsat217(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat218':
                        caseid.caseid_giamsat218(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat219':
                        caseid.caseid_giamsat219(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat220':
                        caseid.caseid_giamsat220(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat221':
                        caseid.caseid_giamsat221(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat222':
                        caseid.caseid_giamsat222(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat223':
                        caseid.caseid_giamsat223(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat224':
                        caseid.caseid_giamsat224(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat225':
                        caseid.caseid_giamsat225(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat226':
                        caseid.caseid_giamsat226(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat227':
                        caseid.caseid_giamsat227(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat228':
                        caseid.caseid_giamsat228(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat229':
                        caseid.caseid_giamsat229(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat230':
                        caseid.caseid_giamsat230(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat231':
                        caseid.caseid_giamsat231(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat232':
                        caseid.caseid_giamsat232(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat233':
                        caseid.caseid_giamsat233(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat234':
                        caseid.caseid_giamsat234(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat235':
                        caseid.caseid_giamsat235(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat236':
                        caseid.caseid_giamsat236(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat237':
                        caseid.caseid_giamsat237(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat238':
                        caseid.caseid_giamsat238(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat239':
                        caseid.caseid_giamsat239(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat241':
                        caseid.caseid_giamsat241(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat242':
                        caseid.caseid_giamsat242(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat243':
                        caseid.caseid_giamsat243(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat244':
                        caseid.caseid_giamsat244(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat245':
                        caseid.caseid_giamsat245(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat246':
                        caseid.caseid_giamsat246(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat247':
                        caseid.caseid_giamsat247(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat248':
                        caseid.caseid_giamsat248(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat249':
                        caseid.caseid_giamsat249(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat250':
                        caseid.caseid_giamsat250(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat251':
                        caseid.caseid_giamsat251(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat252':
                        caseid.caseid_giamsat252(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat253':
                        caseid.caseid_giamsat253(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat254':
                        caseid.caseid_giamsat254(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat255':
                        caseid.caseid_giamsat255(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat256':
                        caseid.caseid_giamsat256(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat258':
                        caseid.caseid_giamsat258(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat259':
                        caseid.caseid_giamsat259(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat260':
                        caseid.caseid_giamsat260(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "2":
            print(list_mucdo2)
            for case in list_mucdo2:
                try:
                    if case == 'GiamSat01':
                        caseid.caseid_giamsat01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat02':
                        caseid.caseid_giamsat02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat03':
                        caseid.caseid_giamsat03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat04':
                        caseid.caseid_giamsat04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat05':
                        caseid.caseid_giamsat05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat06':
                        caseid.caseid_giamsat06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat07':
                        caseid.caseid_giamsat07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat08':
                        caseid.caseid_giamsat08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat09':
                        caseid.caseid_giamsat09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat10':
                        caseid.caseid_giamsat10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat11':
                        caseid.caseid_giamsat11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat12':
                        caseid.caseid_giamsat12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat13':
                        caseid.caseid_giamsat13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat14':
                        caseid.caseid_giamsat14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat15':
                        caseid.caseid_giamsat15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat16':
                        caseid.caseid_giamsat16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat17':
                        caseid.caseid_giamsat17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat18':
                        caseid.caseid_giamsat18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat19':
                        caseid.caseid_giamsat19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat20':
                        caseid.caseid_giamsat20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat21':
                        caseid.caseid_giamsat21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat22':
                        caseid.caseid_giamsat22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat23':
                        caseid.caseid_giamsat23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat24':
                        caseid.caseid_giamsat24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat25':
                        caseid.caseid_giamsat25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat26':
                        caseid.caseid_giamsat26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat27':
                        caseid.caseid_giamsat27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat28':
                        caseid.caseid_giamsat28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat29':
                        caseid.caseid_giamsat29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat30':
                        caseid.caseid_giamsat30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat31':
                        caseid.caseid_giamsat31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat32':
                        caseid.caseid_giamsat32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat33':
                        caseid.caseid_giamsat33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat34':
                        caseid.caseid_giamsat34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat35':
                        caseid.caseid_giamsat35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat36':
                        caseid.caseid_giamsat36(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat37':
                        caseid.caseid_giamsat37(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat38':
                        caseid.caseid_giamsat38(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat39':
                        caseid.caseid_giamsat39(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat40':
                        caseid.caseid_giamsat40(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat41':
                        caseid.caseid_giamsat41(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat42':
                        caseid.caseid_giamsat42(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat43':
                        caseid.caseid_giamsat43(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat44':
                        caseid.caseid_giamsat44(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat45':
                        caseid.caseid_giamsat45(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat46':
                        caseid.caseid_giamsat46(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat47':
                        caseid.caseid_giamsat47(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat48':
                        caseid.caseid_giamsat48(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat49':
                        caseid.caseid_giamsat49(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat50':
                        caseid.caseid_giamsat50(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat51':
                        caseid.caseid_giamsat51(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat52':
                        caseid.caseid_giamsat52(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat53':
                        caseid.caseid_giamsat53(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat54':
                        caseid.caseid_giamsat54(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat55':
                        caseid.caseid_giamsat55(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat56':
                        caseid.caseid_giamsat56(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat57':
                        caseid.caseid_giamsat57(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat58':
                        caseid.caseid_giamsat58(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat59':
                        caseid.caseid_giamsat59(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat60':
                        caseid.caseid_giamsat60(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat61':
                        caseid.caseid_giamsat61(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat62':
                        caseid.caseid_giamsat62(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat63':
                        caseid.caseid_giamsat63(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat64':
                        caseid.caseid_giamsat64(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat65':
                        caseid.caseid_giamsat65(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat66':
                        caseid.caseid_giamsat66(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat67':
                        caseid.caseid_giamsat67(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat68':
                        caseid.caseid_giamsat68(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat69':
                        caseid.caseid_giamsat69(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat70':
                        caseid.caseid_giamsat70(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat71':
                        caseid.caseid_giamsat71(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat72':
                        caseid.caseid_giamsat72(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat73':
                        caseid.caseid_giamsat73(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat74':
                        caseid.caseid_giamsat74(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat75':
                        caseid.caseid_giamsat75(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat76':
                        caseid.caseid_giamsat76(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat77':
                        caseid.caseid_giamsat77(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat78':
                        caseid.caseid_giamsat78(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat79':
                        caseid.caseid_giamsat79(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat80':
                        caseid.caseid_giamsat80(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat80_1':
                        caseid.caseid_giamsat80_1(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat81':
                        caseid.caseid_giamsat81(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat82':
                        caseid.caseid_giamsat82(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat83':
                        caseid.caseid_giamsat83(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat84':
                        caseid.caseid_giamsat84(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat85':
                        caseid.caseid_giamsat85(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat86':
                        caseid.caseid_giamsat86(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat87':
                        caseid.caseid_giamsat87(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat88':
                        caseid.caseid_giamsat88(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat89':
                        caseid.caseid_giamsat89(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat90':
                        caseid.caseid_giamsat90(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat91':
                        caseid.caseid_giamsat91(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat92':
                        caseid.caseid_giamsat92(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat93':
                        caseid.caseid_giamsat93(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat94':
                        caseid.caseid_giamsat94(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat95':
                        caseid.caseid_giamsat95(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat96':
                        caseid.caseid_giamsat96(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat97':
                        caseid.caseid_giamsat97(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat98':
                        caseid.caseid_giamsat98(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat99':
                        caseid.caseid_giamsat99(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat100':
                        caseid.caseid_giamsat100(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat101':
                        caseid.caseid_giamsat101(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat102':
                        caseid.caseid_giamsat102(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat103':
                        caseid.caseid_giamsat103(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat104':
                        caseid.caseid_giamsat104(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat105':
                        caseid.caseid_giamsat105(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat106':
                        caseid.caseid_giamsat106(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat107':
                        caseid.caseid_giamsat107(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat108':
                        caseid.caseid_giamsat108(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat109':
                        caseid.caseid_giamsat109(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat110':
                        caseid.caseid_giamsat110(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat111':
                        caseid.caseid_giamsat111(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat112':
                        caseid.caseid_giamsat112(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat113':
                        caseid.caseid_giamsat113(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat114':
                        caseid.caseid_giamsat114(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat115':
                        caseid.caseid_giamsat115(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat116':
                        caseid.caseid_giamsat116(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat117':
                        caseid.caseid_giamsat117(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat118':
                        caseid.caseid_giamsat118(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat119':
                        caseid.caseid_giamsat119(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat120':
                        caseid.caseid_giamsat120(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat121':
                        caseid.caseid_giamsat121(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat122':
                        caseid.caseid_giamsat122(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat123':
                        caseid.caseid_giamsat123(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat124':
                        caseid.caseid_giamsat124(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat125':
                        caseid.caseid_giamsat125(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat126':
                        caseid.caseid_giamsat126(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat127':
                        caseid.caseid_giamsat127(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat128':
                        caseid.caseid_giamsat128(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat129':
                        caseid.caseid_giamsat129(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat130':
                        caseid.caseid_giamsat130(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat131':
                        caseid.caseid_giamsat131(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat132':
                        caseid.caseid_giamsat132(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat133':
                        caseid.caseid_giamsat133(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat134':
                        caseid.caseid_giamsat134(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat135':
                        caseid.caseid_giamsat135(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat136':
                        caseid.caseid_giamsat136(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat137':
                        caseid.caseid_giamsat137(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat138':
                        caseid.caseid_giamsat138(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat139':
                        caseid.caseid_giamsat139(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat140':
                        caseid.caseid_giamsat140(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat143':
                        caseid.caseid_giamsat143(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat144':
                        caseid.caseid_giamsat144(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat145':
                        caseid.caseid_giamsat145(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat146':
                        caseid.caseid_giamsat146(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat147':
                        caseid.caseid_giamsat147(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat148':
                        caseid.caseid_giamsat148(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat149':
                        caseid.caseid_giamsat149(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat150':
                        caseid.caseid_giamsat150(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat151':
                        caseid.caseid_giamsat151(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat153':
                        caseid.caseid_giamsat153(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat154':
                        caseid.caseid_giamsat154(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat155':
                        caseid.caseid_giamsat155(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat156':
                        caseid.caseid_giamsat156(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat157':
                        caseid.caseid_giamsat157(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat158':
                        caseid.caseid_giamsat158(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat159':
                        caseid.caseid_giamsat159(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat160':
                        caseid.caseid_giamsat160(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat161':
                        caseid.caseid_giamsat161(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat162':
                        caseid.caseid_giamsat162(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat163':
                        caseid.caseid_giamsat163(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat164':
                        caseid.caseid_giamsat164(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat165':
                        caseid.caseid_giamsat165(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat165_1':
                        caseid.caseid_giamsat165_1(self)
                        caseid.caseid_giamsat165_2(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat165_2':
                        caseid.caseid_giamsat165_1(self)
                        caseid.caseid_giamsat165_2(self)
                except:
                    chucnangkhac.swich_tab_0()



                try:
                    if case == 'GiamSat166':
                        caseid.caseid_giamsat166(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat167':
                        caseid.caseid_giamsat167(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat168':
                        caseid.caseid_giamsat168(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat169':
                        caseid.caseid_giamsat169(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat170':
                        caseid.caseid_giamsat170(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat171':
                        caseid.caseid_giamsat171(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat172':
                        caseid.caseid_giamsat172(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat173':
                        caseid.caseid_giamsat173(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat174':
                        caseid.caseid_giamsat174(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat175':
                        caseid.caseid_giamsat175(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat176':
                        caseid.caseid_giamsat176(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat177':
                        caseid.caseid_giamsat177(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat178':
                        caseid.caseid_giamsat178(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat179':
                        caseid.caseid_giamsat179(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat180':
                        caseid.caseid_giamsat180(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat181':
                        caseid.caseid_giamsat181(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat182':
                        caseid.caseid_giamsat182(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat183':
                        caseid.caseid_giamsat183(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat184':
                        caseid.caseid_giamsat184(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat185':
                        caseid.caseid_giamsat185(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat186':
                        caseid.caseid_giamsat186(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat187':
                        caseid.caseid_giamsat187(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat188':
                        caseid.caseid_giamsat188(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat189':
                        caseid.caseid_giamsat189(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat190':
                        caseid.caseid_giamsat190(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat191':
                        caseid.caseid_giamsat191(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat192':
                        caseid.caseid_giamsat192(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat193':
                        caseid.caseid_giamsat193(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat194':
                        caseid.caseid_giamsat194(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat195':
                        caseid.caseid_giamsat195(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat196':
                        caseid.caseid_giamsat196(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat197':
                        caseid.caseid_giamsat197(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat198':
                        caseid.caseid_giamsat198(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat199':
                        caseid.caseid_giamsat199(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat200':
                        caseid.caseid_giamsat200(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat201':
                        caseid.caseid_giamsat201(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat202':
                        caseid.caseid_giamsat202(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat203':
                        caseid.caseid_giamsat203(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat204':
                        caseid.caseid_giamsat204(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat205':
                        caseid.caseid_giamsat205(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat206':
                        caseid.caseid_giamsat206(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat207':
                        caseid.caseid_giamsat207(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat209':
                        caseid.caseid_giamsat209(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat210':
                        caseid.caseid_giamsat210(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat211':
                        caseid.caseid_giamsat211(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat212':
                        caseid.caseid_giamsat212(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat213':
                        caseid.caseid_giamsat213(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat214':
                        caseid.caseid_giamsat214(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat215':
                        caseid.caseid_giamsat215(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat216':
                        caseid.caseid_giamsat216(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat217':
                        caseid.caseid_giamsat217(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat218':
                        caseid.caseid_giamsat218(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat219':
                        caseid.caseid_giamsat219(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat220':
                        caseid.caseid_giamsat220(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat221':
                        caseid.caseid_giamsat221(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat222':
                        caseid.caseid_giamsat222(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat223':
                        caseid.caseid_giamsat223(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat224':
                        caseid.caseid_giamsat224(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat225':
                        caseid.caseid_giamsat225(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat226':
                        caseid.caseid_giamsat226(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat227':
                        caseid.caseid_giamsat227(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat228':
                        caseid.caseid_giamsat228(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat229':
                        caseid.caseid_giamsat229(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat230':
                        caseid.caseid_giamsat230(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat231':
                        caseid.caseid_giamsat231(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat232':
                        caseid.caseid_giamsat232(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat233':
                        caseid.caseid_giamsat233(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat234':
                        caseid.caseid_giamsat234(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat235':
                        caseid.caseid_giamsat235(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat236':
                        caseid.caseid_giamsat236(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat237':
                        caseid.caseid_giamsat237(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat238':
                        caseid.caseid_giamsat238(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat239':
                        caseid.caseid_giamsat239(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat241':
                        caseid.caseid_giamsat241(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat242':
                        caseid.caseid_giamsat242(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat243':
                        caseid.caseid_giamsat243(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat244':
                        caseid.caseid_giamsat244(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat245':
                        caseid.caseid_giamsat245(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat246':
                        caseid.caseid_giamsat246(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat247':
                        caseid.caseid_giamsat247(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat248':
                        caseid.caseid_giamsat248(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat249':
                        caseid.caseid_giamsat249(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat250':
                        caseid.caseid_giamsat250(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat251':
                        caseid.caseid_giamsat251(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat252':
                        caseid.caseid_giamsat252(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat253':
                        caseid.caseid_giamsat253(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat254':
                        caseid.caseid_giamsat254(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat255':
                        caseid.caseid_giamsat255(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat256':
                        caseid.caseid_giamsat256(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat258':
                        caseid.caseid_giamsat258(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat259':
                        caseid.caseid_giamsat259(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat260':
                        caseid.caseid_giamsat260(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "3":
            print(list_mucdo3)
            for case in list_mucdo3:
                try:
                    if case == 'GiamSat01':
                        caseid.caseid_giamsat01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat02':
                        caseid.caseid_giamsat02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat03':
                        caseid.caseid_giamsat03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat04':
                        caseid.caseid_giamsat04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat05':
                        caseid.caseid_giamsat05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat06':
                        caseid.caseid_giamsat06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat07':
                        caseid.caseid_giamsat07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat08':
                        caseid.caseid_giamsat08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat09':
                        caseid.caseid_giamsat09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat10':
                        caseid.caseid_giamsat10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat11':
                        caseid.caseid_giamsat11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat12':
                        caseid.caseid_giamsat12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat13':
                        caseid.caseid_giamsat13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat14':
                        caseid.caseid_giamsat14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat15':
                        caseid.caseid_giamsat15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat16':
                        caseid.caseid_giamsat16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat17':
                        caseid.caseid_giamsat17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat18':
                        caseid.caseid_giamsat18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat19':
                        caseid.caseid_giamsat19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat20':
                        caseid.caseid_giamsat20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat21':
                        caseid.caseid_giamsat21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat22':
                        caseid.caseid_giamsat22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat23':
                        caseid.caseid_giamsat23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat24':
                        caseid.caseid_giamsat24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat25':
                        caseid.caseid_giamsat25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat26':
                        caseid.caseid_giamsat26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat27':
                        caseid.caseid_giamsat27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat28':
                        caseid.caseid_giamsat28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat29':
                        caseid.caseid_giamsat29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat30':
                        caseid.caseid_giamsat30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat31':
                        caseid.caseid_giamsat31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat32':
                        caseid.caseid_giamsat32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat33':
                        caseid.caseid_giamsat33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat34':
                        caseid.caseid_giamsat34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat35':
                        caseid.caseid_giamsat35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat36':
                        caseid.caseid_giamsat36(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat37':
                        caseid.caseid_giamsat37(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat38':
                        caseid.caseid_giamsat38(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat39':
                        caseid.caseid_giamsat39(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat40':
                        caseid.caseid_giamsat40(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat41':
                        caseid.caseid_giamsat41(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat42':
                        caseid.caseid_giamsat42(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat43':
                        caseid.caseid_giamsat43(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat44':
                        caseid.caseid_giamsat44(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat45':
                        caseid.caseid_giamsat45(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat46':
                        caseid.caseid_giamsat46(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat47':
                        caseid.caseid_giamsat47(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat48':
                        caseid.caseid_giamsat48(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat49':
                        caseid.caseid_giamsat49(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat50':
                        caseid.caseid_giamsat50(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat51':
                        caseid.caseid_giamsat51(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat52':
                        caseid.caseid_giamsat52(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat53':
                        caseid.caseid_giamsat53(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat54':
                        caseid.caseid_giamsat54(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat55':
                        caseid.caseid_giamsat55(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat56':
                        caseid.caseid_giamsat56(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat57':
                        caseid.caseid_giamsat57(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat58':
                        caseid.caseid_giamsat58(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat59':
                        caseid.caseid_giamsat59(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat60':
                        caseid.caseid_giamsat60(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat61':
                        caseid.caseid_giamsat61(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat62':
                        caseid.caseid_giamsat62(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat63':
                        caseid.caseid_giamsat63(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat64':
                        caseid.caseid_giamsat64(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat65':
                        caseid.caseid_giamsat65(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat66':
                        caseid.caseid_giamsat66(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat67':
                        caseid.caseid_giamsat67(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat68':
                        caseid.caseid_giamsat68(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat69':
                        caseid.caseid_giamsat69(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat70':
                        caseid.caseid_giamsat70(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat71':
                        caseid.caseid_giamsat71(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat72':
                        caseid.caseid_giamsat72(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat73':
                        caseid.caseid_giamsat73(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat74':
                        caseid.caseid_giamsat74(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat75':
                        caseid.caseid_giamsat75(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat76':
                        caseid.caseid_giamsat76(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat77':
                        caseid.caseid_giamsat77(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat78':
                        caseid.caseid_giamsat78(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat79':
                        caseid.caseid_giamsat79(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat80':
                        caseid.caseid_giamsat80(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat80_1':
                        caseid.caseid_giamsat80_1(self)
                except:
                    chucnangkhac.swich_tab_0()


                try:
                    if case == 'GiamSat81':
                        caseid.caseid_giamsat81(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat82':
                        caseid.caseid_giamsat82(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat83':
                        caseid.caseid_giamsat83(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat84':
                        caseid.caseid_giamsat84(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat85':
                        caseid.caseid_giamsat85(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat86':
                        caseid.caseid_giamsat86(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat87':
                        caseid.caseid_giamsat87(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat88':
                        caseid.caseid_giamsat88(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat89':
                        caseid.caseid_giamsat89(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat90':
                        caseid.caseid_giamsat90(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat91':
                        caseid.caseid_giamsat91(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat92':
                        caseid.caseid_giamsat92(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat93':
                        caseid.caseid_giamsat93(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat94':
                        caseid.caseid_giamsat94(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat95':
                        caseid.caseid_giamsat95(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat96':
                        caseid.caseid_giamsat96(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat97':
                        caseid.caseid_giamsat97(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat98':
                        caseid.caseid_giamsat98(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat99':
                        caseid.caseid_giamsat99(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat100':
                        caseid.caseid_giamsat100(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat101':
                        caseid.caseid_giamsat101(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat102':
                        caseid.caseid_giamsat102(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat103':
                        caseid.caseid_giamsat103(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat104':
                        caseid.caseid_giamsat104(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat105':
                        caseid.caseid_giamsat105(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat106':
                        caseid.caseid_giamsat106(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat107':
                        caseid.caseid_giamsat107(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat108':
                        caseid.caseid_giamsat108(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat109':
                        caseid.caseid_giamsat109(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat110':
                        caseid.caseid_giamsat110(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat111':
                        caseid.caseid_giamsat111(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat112':
                        caseid.caseid_giamsat112(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat113':
                        caseid.caseid_giamsat113(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat114':
                        caseid.caseid_giamsat114(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat115':
                        caseid.caseid_giamsat115(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat116':
                        caseid.caseid_giamsat116(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat117':
                        caseid.caseid_giamsat117(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat118':
                        caseid.caseid_giamsat118(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat119':
                        caseid.caseid_giamsat119(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat120':
                        caseid.caseid_giamsat120(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat121':
                        caseid.caseid_giamsat121(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat122':
                        caseid.caseid_giamsat122(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat123':
                        caseid.caseid_giamsat123(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat124':
                        caseid.caseid_giamsat124(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat125':
                        caseid.caseid_giamsat125(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat126':
                        caseid.caseid_giamsat126(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat127':
                        caseid.caseid_giamsat127(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat128':
                        caseid.caseid_giamsat128(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat129':
                        caseid.caseid_giamsat129(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat130':
                        caseid.caseid_giamsat130(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat131':
                        caseid.caseid_giamsat131(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat132':
                        caseid.caseid_giamsat132(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat133':
                        caseid.caseid_giamsat133(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat134':
                        caseid.caseid_giamsat134(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat135':
                        caseid.caseid_giamsat135(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat136':
                        caseid.caseid_giamsat136(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat137':
                        caseid.caseid_giamsat137(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat138':
                        caseid.caseid_giamsat138(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat139':
                        caseid.caseid_giamsat139(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat140':
                        caseid.caseid_giamsat140(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat143':
                        caseid.caseid_giamsat143(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat144':
                        caseid.caseid_giamsat144(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat145':
                        caseid.caseid_giamsat145(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat146':
                        caseid.caseid_giamsat146(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat147':
                        caseid.caseid_giamsat147(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat148':
                        caseid.caseid_giamsat148(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat149':
                        caseid.caseid_giamsat149(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat150':
                        caseid.caseid_giamsat150(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat151':
                        caseid.caseid_giamsat151(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat153':
                        caseid.caseid_giamsat153(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat154':
                        caseid.caseid_giamsat154(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat155':
                        caseid.caseid_giamsat155(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat156':
                        caseid.caseid_giamsat156(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat157':
                        caseid.caseid_giamsat157(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat158':
                        caseid.caseid_giamsat158(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat159':
                        caseid.caseid_giamsat159(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat160':
                        caseid.caseid_giamsat160(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat161':
                        caseid.caseid_giamsat161(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat162':
                        caseid.caseid_giamsat162(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat163':
                        caseid.caseid_giamsat163(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat164':
                        caseid.caseid_giamsat164(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat165':
                        caseid.caseid_giamsat165(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat165_1':
                        caseid.caseid_giamsat165_1(self)
                        caseid.caseid_giamsat165_2(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat165_2':
                        caseid.caseid_giamsat165_1(self)
                        caseid.caseid_giamsat165_2(self)
                except:
                    chucnangkhac.swich_tab_0()


                try:
                    if case == 'GiamSat166':
                        caseid.caseid_giamsat166(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat167':
                        caseid.caseid_giamsat167(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat168':
                        caseid.caseid_giamsat168(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat169':
                        caseid.caseid_giamsat169(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat170':
                        caseid.caseid_giamsat170(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat171':
                        caseid.caseid_giamsat171(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat172':
                        caseid.caseid_giamsat172(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat173':
                        caseid.caseid_giamsat173(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat174':
                        caseid.caseid_giamsat174(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat175':
                        caseid.caseid_giamsat175(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat176':
                        caseid.caseid_giamsat176(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat177':
                        caseid.caseid_giamsat177(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat178':
                        caseid.caseid_giamsat178(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat179':
                        caseid.caseid_giamsat179(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat180':
                        caseid.caseid_giamsat180(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat181':
                        caseid.caseid_giamsat181(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat182':
                        caseid.caseid_giamsat182(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat183':
                        caseid.caseid_giamsat183(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat184':
                        caseid.caseid_giamsat184(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat185':
                        caseid.caseid_giamsat185(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat186':
                        caseid.caseid_giamsat186(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat187':
                        caseid.caseid_giamsat187(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat188':
                        caseid.caseid_giamsat188(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat189':
                        caseid.caseid_giamsat189(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat190':
                        caseid.caseid_giamsat190(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat191':
                        caseid.caseid_giamsat191(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat192':
                        caseid.caseid_giamsat192(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat193':
                        caseid.caseid_giamsat193(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat194':
                        caseid.caseid_giamsat194(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat195':
                        caseid.caseid_giamsat195(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat196':
                        caseid.caseid_giamsat196(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat197':
                        caseid.caseid_giamsat197(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat198':
                        caseid.caseid_giamsat198(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat199':
                        caseid.caseid_giamsat199(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat200':
                        caseid.caseid_giamsat200(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat201':
                        caseid.caseid_giamsat201(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat202':
                        caseid.caseid_giamsat202(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat203':
                        caseid.caseid_giamsat203(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat204':
                        caseid.caseid_giamsat204(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat205':
                        caseid.caseid_giamsat205(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat206':
                        caseid.caseid_giamsat206(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat207':
                        caseid.caseid_giamsat207(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat209':
                        caseid.caseid_giamsat209(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat210':
                        caseid.caseid_giamsat210(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat211':
                        caseid.caseid_giamsat211(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat212':
                        caseid.caseid_giamsat212(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat213':
                        caseid.caseid_giamsat213(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat214':
                        caseid.caseid_giamsat214(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat215':
                        caseid.caseid_giamsat215(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat216':
                        caseid.caseid_giamsat216(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat217':
                        caseid.caseid_giamsat217(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat218':
                        caseid.caseid_giamsat218(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat219':
                        caseid.caseid_giamsat219(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat220':
                        caseid.caseid_giamsat220(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat221':
                        caseid.caseid_giamsat221(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat222':
                        caseid.caseid_giamsat222(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat223':
                        caseid.caseid_giamsat223(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat224':
                        caseid.caseid_giamsat224(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat225':
                        caseid.caseid_giamsat225(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat226':
                        caseid.caseid_giamsat226(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat227':
                        caseid.caseid_giamsat227(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat228':
                        caseid.caseid_giamsat228(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat229':
                        caseid.caseid_giamsat229(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat230':
                        caseid.caseid_giamsat230(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat231':
                        caseid.caseid_giamsat231(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat232':
                        caseid.caseid_giamsat232(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat233':
                        caseid.caseid_giamsat233(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat234':
                        caseid.caseid_giamsat234(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat235':
                        caseid.caseid_giamsat235(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat236':
                        caseid.caseid_giamsat236(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat237':
                        caseid.caseid_giamsat237(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat238':
                        caseid.caseid_giamsat238(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat239':
                        caseid.caseid_giamsat239(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat241':
                        caseid.caseid_giamsat241(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat242':
                        caseid.caseid_giamsat242(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat243':
                        caseid.caseid_giamsat243(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat244':
                        caseid.caseid_giamsat244(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat245':
                        caseid.caseid_giamsat245(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat246':
                        caseid.caseid_giamsat246(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat247':
                        caseid.caseid_giamsat247(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat248':
                        caseid.caseid_giamsat248(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat249':
                        caseid.caseid_giamsat249(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat250':
                        caseid.caseid_giamsat250(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat251':
                        caseid.caseid_giamsat251(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat252':
                        caseid.caseid_giamsat252(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat253':
                        caseid.caseid_giamsat253(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat254':
                        caseid.caseid_giamsat254(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat255':
                        caseid.caseid_giamsat255(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat256':
                        caseid.caseid_giamsat256(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat258':
                        caseid.caseid_giamsat258(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat259':
                        caseid.caseid_giamsat259(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat260':
                        caseid.caseid_giamsat260(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "4":
            print(list_mucdo4)
            for case in list_mucdo4:
                try:
                    if case == 'GiamSat01':
                        caseid.caseid_giamsat01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat02':
                        caseid.caseid_giamsat02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat03':
                        caseid.caseid_giamsat03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat04':
                        caseid.caseid_giamsat04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat05':
                        caseid.caseid_giamsat05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat06':
                        caseid.caseid_giamsat06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat07':
                        caseid.caseid_giamsat07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat08':
                        caseid.caseid_giamsat08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat09':
                        caseid.caseid_giamsat09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat10':
                        caseid.caseid_giamsat10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat11':
                        caseid.caseid_giamsat11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat12':
                        caseid.caseid_giamsat12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat13':
                        caseid.caseid_giamsat13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat14':
                        caseid.caseid_giamsat14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat15':
                        caseid.caseid_giamsat15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat16':
                        caseid.caseid_giamsat16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat17':
                        caseid.caseid_giamsat17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat18':
                        caseid.caseid_giamsat18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat19':
                        caseid.caseid_giamsat19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat20':
                        caseid.caseid_giamsat20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat21':
                        caseid.caseid_giamsat21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat22':
                        caseid.caseid_giamsat22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat23':
                        caseid.caseid_giamsat23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat24':
                        caseid.caseid_giamsat24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat25':
                        caseid.caseid_giamsat25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat26':
                        caseid.caseid_giamsat26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat27':
                        caseid.caseid_giamsat27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat28':
                        caseid.caseid_giamsat28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat29':
                        caseid.caseid_giamsat29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat30':
                        caseid.caseid_giamsat30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat31':
                        caseid.caseid_giamsat31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat32':
                        caseid.caseid_giamsat32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat33':
                        caseid.caseid_giamsat33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat34':
                        caseid.caseid_giamsat34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat35':
                        caseid.caseid_giamsat35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat36':
                        caseid.caseid_giamsat36(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat37':
                        caseid.caseid_giamsat37(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat38':
                        caseid.caseid_giamsat38(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat39':
                        caseid.caseid_giamsat39(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat40':
                        caseid.caseid_giamsat40(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat41':
                        caseid.caseid_giamsat41(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat42':
                        caseid.caseid_giamsat42(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat43':
                        caseid.caseid_giamsat43(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat44':
                        caseid.caseid_giamsat44(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat45':
                        caseid.caseid_giamsat45(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat46':
                        caseid.caseid_giamsat46(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat47':
                        caseid.caseid_giamsat47(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat48':
                        caseid.caseid_giamsat48(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat49':
                        caseid.caseid_giamsat49(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat50':
                        caseid.caseid_giamsat50(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat51':
                        caseid.caseid_giamsat51(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat52':
                        caseid.caseid_giamsat52(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat53':
                        caseid.caseid_giamsat53(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat54':
                        caseid.caseid_giamsat54(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat55':
                        caseid.caseid_giamsat55(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat56':
                        caseid.caseid_giamsat56(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat57':
                        caseid.caseid_giamsat57(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat58':
                        caseid.caseid_giamsat58(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat59':
                        caseid.caseid_giamsat59(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat60':
                        caseid.caseid_giamsat60(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat61':
                        caseid.caseid_giamsat61(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat62':
                        caseid.caseid_giamsat62(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat63':
                        caseid.caseid_giamsat63(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat64':
                        caseid.caseid_giamsat64(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat65':
                        caseid.caseid_giamsat65(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat66':
                        caseid.caseid_giamsat66(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat67':
                        caseid.caseid_giamsat67(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat68':
                        caseid.caseid_giamsat68(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat69':
                        caseid.caseid_giamsat69(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat70':
                        caseid.caseid_giamsat70(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat71':
                        caseid.caseid_giamsat71(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat72':
                        caseid.caseid_giamsat72(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat73':
                        caseid.caseid_giamsat73(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat74':
                        caseid.caseid_giamsat74(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat75':
                        caseid.caseid_giamsat75(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat76':
                        caseid.caseid_giamsat76(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat77':
                        caseid.caseid_giamsat77(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat78':
                        caseid.caseid_giamsat78(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat79':
                        caseid.caseid_giamsat79(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat80':
                        caseid.caseid_giamsat80(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat80_1':
                        caseid.caseid_giamsat80_1(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat81':
                        caseid.caseid_giamsat81(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat82':
                        caseid.caseid_giamsat82(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat83':
                        caseid.caseid_giamsat83(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat84':
                        caseid.caseid_giamsat84(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat85':
                        caseid.caseid_giamsat85(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat86':
                        caseid.caseid_giamsat86(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat87':
                        caseid.caseid_giamsat87(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat88':
                        caseid.caseid_giamsat88(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat89':
                        caseid.caseid_giamsat89(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat90':
                        caseid.caseid_giamsat90(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat91':
                        caseid.caseid_giamsat91(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat92':
                        caseid.caseid_giamsat92(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat93':
                        caseid.caseid_giamsat93(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat94':
                        caseid.caseid_giamsat94(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat95':
                        caseid.caseid_giamsat95(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat96':
                        caseid.caseid_giamsat96(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat97':
                        caseid.caseid_giamsat97(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat98':
                        caseid.caseid_giamsat98(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat99':
                        caseid.caseid_giamsat99(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat100':
                        caseid.caseid_giamsat100(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat101':
                        caseid.caseid_giamsat101(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat102':
                        caseid.caseid_giamsat102(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat103':
                        caseid.caseid_giamsat103(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat104':
                        caseid.caseid_giamsat104(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat105':
                        caseid.caseid_giamsat105(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat106':
                        caseid.caseid_giamsat106(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat107':
                        caseid.caseid_giamsat107(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat108':
                        caseid.caseid_giamsat108(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat109':
                        caseid.caseid_giamsat109(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat110':
                        caseid.caseid_giamsat110(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat111':
                        caseid.caseid_giamsat111(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat112':
                        caseid.caseid_giamsat112(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat113':
                        caseid.caseid_giamsat113(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat114':
                        caseid.caseid_giamsat114(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat115':
                        caseid.caseid_giamsat115(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat116':
                        caseid.caseid_giamsat116(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat117':
                        caseid.caseid_giamsat117(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat118':
                        caseid.caseid_giamsat118(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat119':
                        caseid.caseid_giamsat119(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat120':
                        caseid.caseid_giamsat120(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat121':
                        caseid.caseid_giamsat121(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat122':
                        caseid.caseid_giamsat122(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat123':
                        caseid.caseid_giamsat123(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat124':
                        caseid.caseid_giamsat124(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat125':
                        caseid.caseid_giamsat125(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat126':
                        caseid.caseid_giamsat126(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat127':
                        caseid.caseid_giamsat127(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat128':
                        caseid.caseid_giamsat128(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat129':
                        caseid.caseid_giamsat129(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat130':
                        caseid.caseid_giamsat130(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat131':
                        caseid.caseid_giamsat131(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat132':
                        caseid.caseid_giamsat132(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat133':
                        caseid.caseid_giamsat133(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat134':
                        caseid.caseid_giamsat134(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat135':
                        caseid.caseid_giamsat135(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat136':
                        caseid.caseid_giamsat136(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat137':
                        caseid.caseid_giamsat137(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat138':
                        caseid.caseid_giamsat138(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat139':
                        caseid.caseid_giamsat139(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat140':
                        caseid.caseid_giamsat140(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat143':
                        caseid.caseid_giamsat143(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat144':
                        caseid.caseid_giamsat144(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat145':
                        caseid.caseid_giamsat145(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat146':
                        caseid.caseid_giamsat146(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat147':
                        caseid.caseid_giamsat147(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat148':
                        caseid.caseid_giamsat148(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat149':
                        caseid.caseid_giamsat149(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat150':
                        caseid.caseid_giamsat150(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat151':
                        caseid.caseid_giamsat151(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat153':
                        caseid.caseid_giamsat153(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat154':
                        caseid.caseid_giamsat154(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat155':
                        caseid.caseid_giamsat155(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat156':
                        caseid.caseid_giamsat156(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat157':
                        caseid.caseid_giamsat157(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat158':
                        caseid.caseid_giamsat158(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat159':
                        caseid.caseid_giamsat159(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat160':
                        caseid.caseid_giamsat160(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat161':
                        caseid.caseid_giamsat161(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat162':
                        caseid.caseid_giamsat162(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat163':
                        caseid.caseid_giamsat163(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat164':
                        caseid.caseid_giamsat164(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat165':
                        caseid.caseid_giamsat165(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat165_1':
                        caseid.caseid_giamsat165_1(self)
                        caseid.caseid_giamsat165_2(self)
                except:
                    chucnangkhac.swich_tab_0()

                try:
                    if case == 'GiamSat165_2':
                        caseid.caseid_giamsat165_1(self)
                        caseid.caseid_giamsat165_2(self)
                except:
                    chucnangkhac.swich_tab_0()



                try:
                    if case == 'GiamSat166':
                        caseid.caseid_giamsat166(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat167':
                        caseid.caseid_giamsat167(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat168':
                        caseid.caseid_giamsat168(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat169':
                        caseid.caseid_giamsat169(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat170':
                        caseid.caseid_giamsat170(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat171':
                        caseid.caseid_giamsat171(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat172':
                        caseid.caseid_giamsat172(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat173':
                        caseid.caseid_giamsat173(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat174':
                        caseid.caseid_giamsat174(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat175':
                        caseid.caseid_giamsat175(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat176':
                        caseid.caseid_giamsat176(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat177':
                        caseid.caseid_giamsat177(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat178':
                        caseid.caseid_giamsat178(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat179':
                        caseid.caseid_giamsat179(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat180':
                        caseid.caseid_giamsat180(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat181':
                        caseid.caseid_giamsat181(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat182':
                        caseid.caseid_giamsat182(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat183':
                        caseid.caseid_giamsat183(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat184':
                        caseid.caseid_giamsat184(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat185':
                        caseid.caseid_giamsat185(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat186':
                        caseid.caseid_giamsat186(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat187':
                        caseid.caseid_giamsat187(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat188':
                        caseid.caseid_giamsat188(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat189':
                        caseid.caseid_giamsat189(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat190':
                        caseid.caseid_giamsat190(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat191':
                        caseid.caseid_giamsat191(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat192':
                        caseid.caseid_giamsat192(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat193':
                        caseid.caseid_giamsat193(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat194':
                        caseid.caseid_giamsat194(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat195':
                        caseid.caseid_giamsat195(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat196':
                        caseid.caseid_giamsat196(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat197':
                        caseid.caseid_giamsat197(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat198':
                        caseid.caseid_giamsat198(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat199':
                        caseid.caseid_giamsat199(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat200':
                        caseid.caseid_giamsat200(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat201':
                        caseid.caseid_giamsat201(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat202':
                        caseid.caseid_giamsat202(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat203':
                        caseid.caseid_giamsat203(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat204':
                        caseid.caseid_giamsat204(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat205':
                        caseid.caseid_giamsat205(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat206':
                        caseid.caseid_giamsat206(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat207':
                        caseid.caseid_giamsat207(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat209':
                        caseid.caseid_giamsat209(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat210':
                        caseid.caseid_giamsat210(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat211':
                        caseid.caseid_giamsat211(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat212':
                        caseid.caseid_giamsat212(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat213':
                        caseid.caseid_giamsat213(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat214':
                        caseid.caseid_giamsat214(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat215':
                        caseid.caseid_giamsat215(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat216':
                        caseid.caseid_giamsat216(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat217':
                        caseid.caseid_giamsat217(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat218':
                        caseid.caseid_giamsat218(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat219':
                        caseid.caseid_giamsat219(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat220':
                        caseid.caseid_giamsat220(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat221':
                        caseid.caseid_giamsat221(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat222':
                        caseid.caseid_giamsat222(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat223':
                        caseid.caseid_giamsat223(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat224':
                        caseid.caseid_giamsat224(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat225':
                        caseid.caseid_giamsat225(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat226':
                        caseid.caseid_giamsat226(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat227':
                        caseid.caseid_giamsat227(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat228':
                        caseid.caseid_giamsat228(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat229':
                        caseid.caseid_giamsat229(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat230':
                        caseid.caseid_giamsat230(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat231':
                        caseid.caseid_giamsat231(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat232':
                        caseid.caseid_giamsat232(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat233':
                        caseid.caseid_giamsat233(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat234':
                        caseid.caseid_giamsat234(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat235':
                        caseid.caseid_giamsat235(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat236':
                        caseid.caseid_giamsat236(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat237':
                        caseid.caseid_giamsat237(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat238':
                        caseid.caseid_giamsat238(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat239':
                        caseid.caseid_giamsat239(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat241':
                        caseid.caseid_giamsat241(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat242':
                        caseid.caseid_giamsat242(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat243':
                        caseid.caseid_giamsat243(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat244':
                        caseid.caseid_giamsat244(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat245':
                        caseid.caseid_giamsat245(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat246':
                        caseid.caseid_giamsat246(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat247':
                        caseid.caseid_giamsat247(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat248':
                        caseid.caseid_giamsat248(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat249':
                        caseid.caseid_giamsat249(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat250':
                        caseid.caseid_giamsat250(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat251':
                        caseid.caseid_giamsat251(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat252':
                        caseid.caseid_giamsat252(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat253':
                        caseid.caseid_giamsat253(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat254':
                        caseid.caseid_giamsat254(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat255':
                        caseid.caseid_giamsat255(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat256':
                        caseid.caseid_giamsat256(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat258':
                        caseid.caseid_giamsat258(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat259':
                        caseid.caseid_giamsat259(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'GiamSat260':
                        caseid.caseid_giamsat260(self)
                except:
                    chucnangkhac.swich_tab_0()



#3 lộ trình
def route(self):
    var.driver.set_page_load_timeout(20)
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 430
    while (rownum < 458):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo1)

    modetest = ''.join(re.findall(r'\d+', var.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Route01':
                        caseid.caseid_route01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route02':
                        caseid.caseid_route02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route03':
                        caseid.caseid_route03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route04':
                        caseid.caseid_route04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route05':
                        caseid.caseid_route05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route06':
                        caseid.caseid_route06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route07':
                        caseid.caseid_route07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route08':
                        caseid.caseid_route08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route09':
                        caseid.caseid_route09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route10':
                        caseid.caseid_route10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route11':
                        caseid.caseid_route11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route12':
                        caseid.caseid_route12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route13':
                        caseid.caseid_route13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route14':
                        caseid.caseid_route14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route15':
                        caseid.caseid_route15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route16':
                        caseid.caseid_route16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route17':
                        caseid.caseid_route17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route18':
                        caseid.caseid_route18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route19':
                        caseid.caseid_route19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route20':
                        caseid.caseid_route20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route21':
                        caseid.caseid_route21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route22':
                        caseid.caseid_route22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route23':
                        caseid.caseid_route23(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Route01':
                        caseid.caseid_route01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route02':
                        caseid.caseid_route02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route03':
                        caseid.caseid_route03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route04':
                        caseid.caseid_route04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route05':
                        caseid.caseid_route05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route06':
                        caseid.caseid_route06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route07':
                        caseid.caseid_route07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route08':
                        caseid.caseid_route08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route09':
                        caseid.caseid_route09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route10':
                        caseid.caseid_route10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route11':
                        caseid.caseid_route11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route12':
                        caseid.caseid_route12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route13':
                        caseid.caseid_route13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route14':
                        caseid.caseid_route14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route15':
                        caseid.caseid_route15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route16':
                        caseid.caseid_route16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route17':
                        caseid.caseid_route17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route18':
                        caseid.caseid_route18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route19':
                        caseid.caseid_route19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route20':
                        caseid.caseid_route20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route21':
                        caseid.caseid_route21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route22':
                        caseid.caseid_route22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route23':
                        caseid.caseid_route23(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Route01':
                        caseid.caseid_route01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route02':
                        caseid.caseid_route02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route03':
                        caseid.caseid_route03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route04':
                        caseid.caseid_route04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route05':
                        caseid.caseid_route05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route06':
                        caseid.caseid_route06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route07':
                        caseid.caseid_route07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route08':
                        caseid.caseid_route08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route09':
                        caseid.caseid_route09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route10':
                        caseid.caseid_route10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route11':
                        caseid.caseid_route11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route12':
                        caseid.caseid_route12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route13':
                        caseid.caseid_route13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route14':
                        caseid.caseid_route14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route15':
                        caseid.caseid_route15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route16':
                        caseid.caseid_route16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route17':
                        caseid.caseid_route17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route18':
                        caseid.caseid_route18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route19':
                        caseid.caseid_route19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route20':
                        caseid.caseid_route20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route21':
                        caseid.caseid_route21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route22':
                        caseid.caseid_route22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route23':
                        caseid.caseid_route23(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Route01':
                        caseid.caseid_route01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route02':
                        caseid.caseid_route02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route03':
                        caseid.caseid_route03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route04':
                        caseid.caseid_route04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route05':
                        caseid.caseid_route05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route06':
                        caseid.caseid_route06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route07':
                        caseid.caseid_route07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route08':
                        caseid.caseid_route08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route09':
                        caseid.caseid_route09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route10':
                        caseid.caseid_route10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route11':
                        caseid.caseid_route11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route12':
                        caseid.caseid_route12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route13':
                        caseid.caseid_route13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route14':
                        caseid.caseid_route14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route15':
                        caseid.caseid_route15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route16':
                        caseid.caseid_route16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route17':
                        caseid.caseid_route17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route18':
                        caseid.caseid_route18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route19':
                        caseid.caseid_route19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route20':
                        caseid.caseid_route20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route21':
                        caseid.caseid_route21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route22':
                        caseid.caseid_route22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Route23':
                        caseid.caseid_route23(self)
                except:
                    chucnangkhac.swich_tab_0()


#4 quản trị
def administration(self):
    var.driver.set_page_load_timeout(20)
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 459
    while (rownum < 527):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo1)

    modetest = ''.join(re.findall(r'\d+', var.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Admin01':
                        caseid.caseid_admin01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid.caseid_admin02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid.caseid_admin03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid.caseid_admin04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid.caseid_admin05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid.caseid_admin06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid.caseid_admin07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid.caseid_admin08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid.caseid_admin09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid.caseid_admin10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid.caseid_admin11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid.caseid_admin12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid.caseid_admin13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid.caseid_admin14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid.caseid_admin15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid.caseid_admin16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid.caseid_admin17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid.caseid_admin18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid.caseid_admin19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid.caseid_admin20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid.caseid_admin21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid.caseid_admin22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid.caseid_admin23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid.caseid_admin24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid.caseid_admin25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid.caseid_admin26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid.caseid_admin27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid.caseid_admin28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid.caseid_admin29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid.caseid_admin30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid.caseid_admin31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid.caseid_admin32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid.caseid_admin33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid.caseid_admin34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid.caseid_admin35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User01':
                        caseid.caseid_user01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User02':
                        caseid.caseid_user02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User03':
                        caseid.caseid_user03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User04':
                        caseid.caseid_user04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User05':
                        caseid.caseid_user05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User06':
                        caseid.caseid_user06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User07':
                        caseid.caseid_user07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User08':
                        caseid.caseid_user08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User09':
                        caseid.caseid_user09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User10':
                        caseid.caseid_user10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User11':
                        caseid.caseid_user11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User12':
                        caseid.caseid_user12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User12':
                        caseid.caseid_user12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User13':
                        caseid.caseid_user13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User14':
                        caseid.caseid_user14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User15':
                        caseid.caseid_user15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User16':
                        caseid.caseid_user16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User17':
                        caseid.caseid_user17(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Admin01':
                        caseid.caseid_admin01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid.caseid_admin02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid.caseid_admin03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid.caseid_admin04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid.caseid_admin05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid.caseid_admin06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid.caseid_admin07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid.caseid_admin08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid.caseid_admin09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid.caseid_admin10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid.caseid_admin11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid.caseid_admin12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid.caseid_admin13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid.caseid_admin14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid.caseid_admin15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid.caseid_admin16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid.caseid_admin17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid.caseid_admin18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid.caseid_admin19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid.caseid_admin20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid.caseid_admin21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid.caseid_admin22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid.caseid_admin23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid.caseid_admin24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid.caseid_admin25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid.caseid_admin26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid.caseid_admin27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid.caseid_admin28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid.caseid_admin29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid.caseid_admin30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid.caseid_admin31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid.caseid_admin32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid.caseid_admin33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid.caseid_admin34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid.caseid_admin35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User01':
                        caseid.caseid_user01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User02':
                        caseid.caseid_user02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User03':
                        caseid.caseid_user03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User04':
                        caseid.caseid_user04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User05':
                        caseid.caseid_user05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User06':
                        caseid.caseid_user06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User07':
                        caseid.caseid_user07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User08':
                        caseid.caseid_user08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User09':
                        caseid.caseid_user09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User10':
                        caseid.caseid_user10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User11':
                        caseid.caseid_user11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User12':
                        caseid.caseid_user12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User13':
                        caseid.caseid_user13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User14':
                        caseid.caseid_user14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User15':
                        caseid.caseid_user15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User16':
                        caseid.caseid_user16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User17':
                        caseid.caseid_user17(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Admin01':
                        caseid.caseid_admin01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid.caseid_admin02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid.caseid_admin03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid.caseid_admin04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid.caseid_admin05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid.caseid_admin06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid.caseid_admin07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid.caseid_admin08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid.caseid_admin09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid.caseid_admin10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid.caseid_admin11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid.caseid_admin12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid.caseid_admin13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid.caseid_admin14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid.caseid_admin15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid.caseid_admin16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid.caseid_admin17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid.caseid_admin18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid.caseid_admin19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid.caseid_admin20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid.caseid_admin21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid.caseid_admin22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid.caseid_admin23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid.caseid_admin24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid.caseid_admin25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid.caseid_admin26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid.caseid_admin27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid.caseid_admin28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid.caseid_admin29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid.caseid_admin30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid.caseid_admin31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid.caseid_admin32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid.caseid_admin33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid.caseid_admin34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid.caseid_admin35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User01':
                        caseid.caseid_user01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User02':
                        caseid.caseid_user02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User03':
                        caseid.caseid_user03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User04':
                        caseid.caseid_user04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User05':
                        caseid.caseid_user05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User06':
                        caseid.caseid_user06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User07':
                        caseid.caseid_user07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User08':
                        caseid.caseid_user08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User09':
                        caseid.caseid_user09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User10':
                        caseid.caseid_user10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User11':
                        caseid.caseid_user11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User12':
                        caseid.caseid_user12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User13':
                        caseid.caseid_user13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User14':
                        caseid.caseid_user14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User15':
                        caseid.caseid_user15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User16':
                        caseid.caseid_user16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User17':
                        caseid.caseid_user17(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Admin01':
                        caseid.caseid_admin01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid.caseid_admin02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid.caseid_admin03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid.caseid_admin04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid.caseid_admin05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid.caseid_admin06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid.caseid_admin07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid.caseid_admin08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid.caseid_admin09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid.caseid_admin10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid.caseid_admin11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid.caseid_admin12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid.caseid_admin13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid.caseid_admin14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid.caseid_admin15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid.caseid_admin16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid.caseid_admin17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid.caseid_admin18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid.caseid_admin19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid.caseid_admin20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid.caseid_admin21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid.caseid_admin22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid.caseid_admin23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid.caseid_admin24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid.caseid_admin25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid.caseid_admin26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid.caseid_admin27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid.caseid_admin28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid.caseid_admin29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid.caseid_admin30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid.caseid_admin31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid.caseid_admin32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid.caseid_admin33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid.caseid_admin34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid.caseid_admin35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User01':
                        caseid.caseid_user01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User02':
                        caseid.caseid_user02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User03':
                        caseid.caseid_user03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User04':
                        caseid.caseid_user04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User05':
                        caseid.caseid_user05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User06':
                        caseid.caseid_user06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User07':
                        caseid.caseid_user07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User08':
                        caseid.caseid_user08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User09':
                        caseid.caseid_user09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User10':
                        caseid.caseid_user10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User11':
                        caseid.caseid_user11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User12':
                        caseid.caseid_user12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User13':
                        caseid.caseid_user13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User14':
                        caseid.caseid_user14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User15':
                        caseid.caseid_user15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User16':
                        caseid.caseid_user16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'User17':
                        caseid.caseid_user17(self)
                except:
                    chucnangkhac.swich_tab_0()



#5 báo cáo
def report(self):
    var.driver.set_page_load_timeout(20)
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 528
    while (rownum < 630):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo1)

    modetest = ''.join(re.findall(r'\d+', var.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Report01':
                        caseid.caseid_report01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid.caseid_report02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid.caseid_report03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid.caseid_report04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid.caseid_report05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid.caseid_report06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid.caseid_report07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report08':
                        caseid.caseid_report08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid.caseid_report09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid.caseid_report10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid.caseid_report11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid.caseid_report12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid.caseid_report13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid.caseid_report14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid.caseid_report15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid.caseid_report16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid.caseid_report17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid.caseid_report18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid.caseid_report19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid.caseid_report20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid.caseid_report21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid.caseid_report22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid.caseid_report23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid.caseid_report24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid.caseid_report25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid.caseid_report26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid.caseid_report27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid.caseid_report28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid.caseid_report29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid.caseid_report30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid.caseid_report31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid.caseid_report32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid.caseid_report33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid.caseid_report34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid.caseid_report35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid.caseid_report36(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid.caseid_report37(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid.caseid_report38(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid.caseid_report39(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid.caseid_report40(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid.caseid_report41(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid.caseid_report42(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid.caseid_report43(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid.caseid_report44(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid.caseid_report45(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid.caseid_report46(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid.caseid_report47(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid.caseid_report48(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid.caseid_report49(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid.caseid_report50(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid.caseid_report51(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid.caseid_report52(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid.caseid_report53(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid.caseid_report54(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid.caseid_report55(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid.caseid_report56(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid.caseid_report57(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid.caseid_report58(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid.caseid_report59(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid.caseid_report60(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid.caseid_report61(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid.caseid_report62(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid.caseid_report63(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid.caseid_report64(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid.caseid_report65(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid.caseid_report66(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid.caseid_report67(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid.caseid_report68(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid.caseid_report69(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid.caseid_report70(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Report01':
                        caseid.caseid_report01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid.caseid_report02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid.caseid_report03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid.caseid_report04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid.caseid_report05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid.caseid_report06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid.caseid_report07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report08':
                        caseid.caseid_report08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid.caseid_report09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid.caseid_report10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid.caseid_report11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid.caseid_report12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid.caseid_report13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid.caseid_report14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid.caseid_report15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid.caseid_report16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid.caseid_report17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid.caseid_report18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid.caseid_report19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid.caseid_report20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid.caseid_report21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid.caseid_report22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid.caseid_report23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid.caseid_report24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid.caseid_report25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid.caseid_report26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid.caseid_report27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid.caseid_report28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid.caseid_report29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid.caseid_report30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid.caseid_report31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid.caseid_report32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid.caseid_report33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid.caseid_report34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid.caseid_report35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid.caseid_report36(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid.caseid_report37(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid.caseid_report38(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid.caseid_report39(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid.caseid_report40(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid.caseid_report41(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid.caseid_report42(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid.caseid_report43(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid.caseid_report44(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid.caseid_report45(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid.caseid_report46(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid.caseid_report47(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid.caseid_report48(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid.caseid_report49(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid.caseid_report50(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid.caseid_report51(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid.caseid_report52(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid.caseid_report53(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid.caseid_report54(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid.caseid_report55(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid.caseid_report56(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid.caseid_report57(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid.caseid_report58(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid.caseid_report59(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid.caseid_report60(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid.caseid_report61(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid.caseid_report62(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid.caseid_report63(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid.caseid_report64(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid.caseid_report65(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid.caseid_report66(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid.caseid_report67(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid.caseid_report68(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid.caseid_report69(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid.caseid_report70(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Report01':
                        caseid.caseid_report01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid.caseid_report02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid.caseid_report03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid.caseid_report04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid.caseid_report05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid.caseid_report06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid.caseid_report07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report08':
                        caseid.caseid_report08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid.caseid_report09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid.caseid_report10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid.caseid_report11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid.caseid_report12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid.caseid_report13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid.caseid_report14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid.caseid_report15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid.caseid_report16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid.caseid_report17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid.caseid_report18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid.caseid_report19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid.caseid_report20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid.caseid_report21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid.caseid_report22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid.caseid_report23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid.caseid_report24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid.caseid_report25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid.caseid_report26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid.caseid_report27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid.caseid_report28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid.caseid_report29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid.caseid_report30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid.caseid_report31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid.caseid_report32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid.caseid_report33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid.caseid_report34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid.caseid_report35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid.caseid_report36(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid.caseid_report37(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid.caseid_report38(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid.caseid_report39(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid.caseid_report40(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid.caseid_report41(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid.caseid_report42(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid.caseid_report43(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid.caseid_report44(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid.caseid_report45(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid.caseid_report46(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid.caseid_report47(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid.caseid_report48(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid.caseid_report49(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid.caseid_report50(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid.caseid_report51(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid.caseid_report52(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid.caseid_report53(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid.caseid_report54(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid.caseid_report55(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid.caseid_report56(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid.caseid_report57(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid.caseid_report58(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid.caseid_report59(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid.caseid_report60(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid.caseid_report61(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid.caseid_report62(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid.caseid_report63(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid.caseid_report64(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid.caseid_report65(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid.caseid_report66(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid.caseid_report67(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid.caseid_report68(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid.caseid_report69(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid.caseid_report70(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Report01':
                        caseid.caseid_report01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid.caseid_report02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid.caseid_report03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid.caseid_report04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid.caseid_report05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid.caseid_report06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid.caseid_report07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report08':
                        caseid.caseid_report08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid.caseid_report09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid.caseid_report10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid.caseid_report11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid.caseid_report12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid.caseid_report13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid.caseid_report14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid.caseid_report15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid.caseid_report16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid.caseid_report17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid.caseid_report18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid.caseid_report19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid.caseid_report20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid.caseid_report21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid.caseid_report22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid.caseid_report23(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid.caseid_report24(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid.caseid_report25(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid.caseid_report26(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid.caseid_report27(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid.caseid_report28(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid.caseid_report29(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid.caseid_report30(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid.caseid_report31(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid.caseid_report32(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid.caseid_report33(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid.caseid_report34(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid.caseid_report35(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid.caseid_report36(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid.caseid_report37(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid.caseid_report38(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid.caseid_report39(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid.caseid_report40(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid.caseid_report41(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid.caseid_report42(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid.caseid_report43(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid.caseid_report44(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid.caseid_report45(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid.caseid_report46(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid.caseid_report47(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid.caseid_report48(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid.caseid_report49(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid.caseid_report50(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid.caseid_report51(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid.caseid_report52(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid.caseid_report53(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid.caseid_report54(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid.caseid_report55(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid.caseid_report56(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid.caseid_report57(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid.caseid_report58(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid.caseid_report59(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid.caseid_report60(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid.caseid_report61(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid.caseid_report62(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid.caseid_report63(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid.caseid_report64(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid.caseid_report65(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid.caseid_report66(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid.caseid_report67(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid.caseid_report68(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid.caseid_report69(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid.caseid_report70(self)
                except:
                    chucnangkhac.swich_tab_0()




#6 video clip
def videoclip(self):
    var.driver.set_page_load_timeout(20)
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 631
    while (rownum < 649):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo4)

    modetest = ''.join(re.findall(r'\d+', var.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Video01':
                        caseid.caseid_video01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video02':
                        caseid.caseid_video02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video03':
                        caseid.caseid_video03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video04':
                        caseid.caseid_video04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video05':
                        caseid.caseid_video05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video06':
                        caseid.caseid_video06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video07':
                        caseid.caseid_video07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video08':
                        caseid.caseid_video08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video09':
                        caseid.caseid_video09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video10':
                        caseid.caseid_video10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video11':
                        caseid.caseid_video11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video12':
                        caseid.caseid_video12(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Video01':
                        caseid.caseid_video01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video02':
                        caseid.caseid_video02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video03':
                        caseid.caseid_video03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video04':
                        caseid.caseid_video04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video05':
                        caseid.caseid_video05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video06':
                        caseid.caseid_video06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video07':
                        caseid.caseid_video07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video08':
                        caseid.caseid_video08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video09':
                        caseid.caseid_video09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video10':
                        caseid.caseid_video10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video11':
                        caseid.caseid_video11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video12':
                        caseid.caseid_video12(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Video01':
                        caseid.caseid_video01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video02':
                        caseid.caseid_video02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video03':
                        caseid.caseid_video03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video04':
                        caseid.caseid_video04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video05':
                        caseid.caseid_video05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video06':
                        caseid.caseid_video06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video07':
                        caseid.caseid_video07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video08':
                        caseid.caseid_video08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video09':
                        caseid.caseid_video09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video10':
                        caseid.caseid_video10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video11':
                        caseid.caseid_video11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video12':
                        caseid.caseid_video12(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Video01':
                        caseid.caseid_video01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video02':
                        caseid.caseid_video02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video03':
                        caseid.caseid_video03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video04':
                        caseid.caseid_video04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video05':
                        caseid.caseid_video05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video06':
                        caseid.caseid_video06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video07':
                        caseid.caseid_video07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video08':
                        caseid.caseid_video08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video09':
                        caseid.caseid_video09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video10':
                        caseid.caseid_video10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video11':
                        caseid.caseid_video11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Video12':
                        caseid.caseid_video12(self)
                except:
                    chucnangkhac.swich_tab_0()




#7 hình ảnh
def image(self):
    var.driver.set_page_load_timeout(20)
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 650
    while (rownum < 680):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)


    modetest = ''.join(re.findall(r'\d+', var.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Image01':
                        caseid.caseid_image01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image02':
                        caseid.caseid_image02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image03':
                        caseid.caseid_image03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image04':
                        caseid.caseid_image04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image05':
                        caseid.caseid_image05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image06':
                        caseid.caseid_image06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image07':
                        caseid.caseid_image07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image08':
                        caseid.caseid_image08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image09':
                        caseid.caseid_image09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image10':
                        caseid.caseid_image10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image11':
                        caseid.caseid_image11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image12':
                        caseid.caseid_image12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image13':
                        caseid.caseid_image13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image14':
                        caseid.caseid_image14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image15':
                        caseid.caseid_image15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image16':
                        caseid.caseid_image16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image17':
                        caseid.caseid_image17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image18':
                        caseid.caseid_image18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image19':
                        caseid.caseid_image19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image20':
                        caseid.caseid_image20(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Image01':
                        caseid.caseid_image01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image02':
                        caseid.caseid_image02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image03':
                        caseid.caseid_image03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image04':
                        caseid.caseid_image04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image05':
                        caseid.caseid_image05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image06':
                        caseid.caseid_image06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image07':
                        caseid.caseid_image07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image08':
                        caseid.caseid_image08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image09':
                        caseid.caseid_image09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image10':
                        caseid.caseid_image10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image11':
                        caseid.caseid_image11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image12':
                        caseid.caseid_image12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image13':
                        caseid.caseid_image13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image14':
                        caseid.caseid_image14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image15':
                        caseid.caseid_image15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image16':
                        caseid.caseid_image16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image17':
                        caseid.caseid_image17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image18':
                        caseid.caseid_image18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image19':
                        caseid.caseid_image19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image20':
                        caseid.caseid_image20(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Image01':
                        caseid.caseid_image01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image02':
                        caseid.caseid_image02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image03':
                        caseid.caseid_image03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image04':
                        caseid.caseid_image04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image05':
                        caseid.caseid_image05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image06':
                        caseid.caseid_image06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image07':
                        caseid.caseid_image07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image08':
                        caseid.caseid_image08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image09':
                        caseid.caseid_image09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image10':
                        caseid.caseid_image10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image11':
                        caseid.caseid_image11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image12':
                        caseid.caseid_image12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image13':
                        caseid.caseid_image13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image14':
                        caseid.caseid_image14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image15':
                        caseid.caseid_image15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image16':
                        caseid.caseid_image16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image17':
                        caseid.caseid_image17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image18':
                        caseid.caseid_image18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image19':
                        caseid.caseid_image19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image20':
                        caseid.caseid_image20(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Image01':
                        caseid.caseid_image01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image02':
                        caseid.caseid_image02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image03':
                        caseid.caseid_image03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image04':
                        caseid.caseid_image04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image05':
                        caseid.caseid_image05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image06':
                        caseid.caseid_image06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image07':
                        caseid.caseid_image07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image08':
                        caseid.caseid_image08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image09':
                        caseid.caseid_image09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image10':
                        caseid.caseid_image10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image11':
                        caseid.caseid_image11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image12':
                        caseid.caseid_image12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image13':
                        caseid.caseid_image13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image14':
                        caseid.caseid_image14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image15':
                        caseid.caseid_image15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image16':
                        caseid.caseid_image16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image17':
                        caseid.caseid_image17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image18':
                        caseid.caseid_image18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image19':
                        caseid.caseid_image19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Image20':
                        caseid.caseid_image20(self)
                except:
                    chucnangkhac.swich_tab_0()




#8 tien ích
def utility(self):
    var.driver.set_page_load_timeout(20)
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 681
    while (rownum < 714):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)


    modetest = ''.join(re.findall(r'\d+', var.modetest))
    for i in modetest:

        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Utility01':
                        caseid.caseid_utility01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility02':
                        caseid.caseid_utility02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility03':
                        caseid.caseid_utility03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility04':
                        caseid.caseid_utility04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility05':
                        caseid.caseid_utility05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility06':
                        caseid.caseid_utility06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility07':
                        caseid.caseid_utility07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility08':
                        caseid.caseid_utility08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility09':
                        caseid.caseid_utility09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility10':
                        caseid.caseid_utility10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility11':
                        caseid.caseid_utility11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility12':
                        caseid.caseid_utility12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility13':
                        caseid.caseid_utility13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility14':
                        caseid.caseid_utility14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility15':
                        caseid.caseid_utility15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility16':
                        caseid.caseid_utility16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility17':
                        caseid.caseid_utility17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility18':
                        caseid.caseid_utility18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility19':
                        caseid.caseid_utility19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility20':
                        caseid.caseid_utility20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility21':
                        caseid.caseid_utility21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility22':
                        caseid.caseid_utility22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility23':
                        caseid.caseid_utility23(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Utility01':
                        caseid.caseid_utility01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility02':
                        caseid.caseid_utility02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility03':
                        caseid.caseid_utility03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility04':
                        caseid.caseid_utility04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility05':
                        caseid.caseid_utility05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility06':
                        caseid.caseid_utility06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility07':
                        caseid.caseid_utility07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility08':
                        caseid.caseid_utility08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility09':
                        caseid.caseid_utility09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility10':
                        caseid.caseid_utility10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility11':
                        caseid.caseid_utility11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility12':
                        caseid.caseid_utility12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility13':
                        caseid.caseid_utility13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility14':
                        caseid.caseid_utility14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility15':
                        caseid.caseid_utility15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility16':
                        caseid.caseid_utility16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility17':
                        caseid.caseid_utility17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility18':
                        caseid.caseid_utility18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility19':
                        caseid.caseid_utility19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility20':
                        caseid.caseid_utility20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility21':
                        caseid.caseid_utility21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility22':
                        caseid.caseid_utility22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility23':
                        caseid.caseid_utility23(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Utility01':
                        caseid.caseid_utility01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility02':
                        caseid.caseid_utility02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility03':
                        caseid.caseid_utility03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility04':
                        caseid.caseid_utility04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility05':
                        caseid.caseid_utility05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility06':
                        caseid.caseid_utility06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility07':
                        caseid.caseid_utility07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility08':
                        caseid.caseid_utility08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility09':
                        caseid.caseid_utility09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility10':
                        caseid.caseid_utility10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility11':
                        caseid.caseid_utility11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility12':
                        caseid.caseid_utility12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility13':
                        caseid.caseid_utility13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility14':
                        caseid.caseid_utility14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility15':
                        caseid.caseid_utility15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility16':
                        caseid.caseid_utility16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility17':
                        caseid.caseid_utility17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility18':
                        caseid.caseid_utility18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility19':
                        caseid.caseid_utility19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility20':
                        caseid.caseid_utility20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility21':
                        caseid.caseid_utility21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility22':
                        caseid.caseid_utility22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility23':
                        caseid.caseid_utility23(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Utility01':
                        caseid.caseid_utility01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility02':
                        caseid.caseid_utility02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility03':
                        caseid.caseid_utility03(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility04':
                        caseid.caseid_utility04(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility05':
                        caseid.caseid_utility05(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility06':
                        caseid.caseid_utility06(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility07':
                        caseid.caseid_utility07(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility08':
                        caseid.caseid_utility08(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility09':
                        caseid.caseid_utility09(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility10':
                        caseid.caseid_utility10(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility11':
                        caseid.caseid_utility11(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility12':
                        caseid.caseid_utility12(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility13':
                        caseid.caseid_utility13(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility14':
                        caseid.caseid_utility14(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility15':
                        caseid.caseid_utility15(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility16':
                        caseid.caseid_utility16(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility17':
                        caseid.caseid_utility17(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility18':
                        caseid.caseid_utility18(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility19':
                        caseid.caseid_utility19(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility20':
                        caseid.caseid_utility20(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility21':
                        caseid.caseid_utility21(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility22':
                        caseid.caseid_utility22(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Utility23':
                        caseid.caseid_utility23(self)
                except:
                    chucnangkhac.swich_tab_0()





#9 AI
def ai(self):
    var.driver.set_page_load_timeout(20)
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 715
    while (rownum < 725):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)

    modetest = ''.join(re.findall(r'\d+', var.modetest))
    for i in modetest:

        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Ai01':
                        caseid.caseid_ai01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Ai02':
                        caseid.caseid_ai02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Ai03':
                        caseid.caseid_ai03(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Ai01':
                        caseid.caseid_ai01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Ai02':
                        caseid.caseid_ai02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Ai03':
                        caseid.caseid_ai03(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Ai01':
                        caseid.caseid_ai01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Ai02':
                        caseid.caseid_ai02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Ai03':
                        caseid.caseid_ai03(self)
                except:
                    chucnangkhac.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Ai01':
                        caseid.caseid_ai01(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Ai02':
                        caseid.caseid_ai02(self)
                except:
                    chucnangkhac.swich_tab_0()
                try:
                    if case == 'Ai03':
                        caseid.caseid_ai03(self)
                except:
                    chucnangkhac.swich_tab_0()






















