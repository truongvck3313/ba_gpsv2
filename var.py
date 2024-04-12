from selenium.webdriver.common.by import By
#trình duyệt 2
import openpyxl

from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
capa = DesiredCapabilities.CHROME
capa["pageLoadStrategy"] = "none"
from seleniumwire import webdriver
driver = webdriver.Chrome(desired_capabilities=capa)
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')
chrome_options.add_argument('window-size=1920x1480')




PATH ="C:/Users/truongtq.BA/PycharmProjects/pythonProject/ba_v2/chromedriver.exe"
path_baocao = "C:/Users/truongtq.BA/PycharmProjects/pythonProject/ba_v2/GPS_ChecklistForAutoTest.xlsx"
path_luutamthoi = "C:/Users/truongtq.BA/PycharmProjects/pythonProject/ba_v2/bangluudulieu_tamthoi.xlsx"


def readData(file,sheetName,rownum,columnno):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,rowum,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowum,column=columnno).value = data
    wordbook.save(file)


#đọc file config
f = open("file_config.txt", 'r')
for x in f:
     if x[0:11] == "- ModeTest:":   #1,2,3,4
         modetest = x[13:-2]
     if x[0:11] == "- LinkTest:":   #https://gps.binhanh.vn
         linktest = x[13:-2]
     if x[0:10] == "- TimeRun:":    #09:41
         timerun = x[12:-2]
     if x[0:10] == "- LogPath:":    #C:/Users/truongtq.BA/PycharmProjects/pythonProject/ba_v2/ba.log
         logpath = x[12:-2]
     if x[0:16] == "- CheckListPath:":  #C:/Users/truongtq.BA/PycharmProjects/pythonProject/ba_v2/GPS_ChecklistForAutoTest_1.xlsx
         checklistpath = x[18:-2]
     if x[0:12] == "- ImagePath:":  #C:/Users/truongtq.BA/PycharmProjects/pythonProject/anhchup
         imagepath = x[14:-2]
     if x[0:15] == "- DataTestPath:":   #C:/Users/truongtq.BA/PycharmProjects/pythonProject/ba_v2/data_ba.json
         datatestpath = x[17:-2]
     if x[0:13] == "- UploadPath:":     #C:/Users/truongtq.BA/PycharmProjects/pythonProject/file
         uploadpath = x[15:-2]
     if x[0:23] == "- LuuDuLieuTamThoiPath:":
         luudulieutamthoipath = x[25:-2]






login_user = "//*[@placeholder='Tên đăng nhập']"
login_password = "//*[@placeholder='Mật khẩu']"
login_ghinhodangnhap = "//*[@checked='checked']"
dangnhap = "//*[@id='UserLogin1_btnLogin']"

icon_bagps = "//*[@id='imgLogo']"
check_loginsai = "//*[text()='Tên đăng nhập hoặc mật khẩu không hợp lệ.']"
login_iconappstore = "//*[@src='/Images/Login/iconLogin/SVG/photo_logo-IOS.svg']"
check_login_appstore = "//*[text()='BA GPS']"

hopthoai = By.XPATH, "//*[@class='chat-list custom-scroll']//*[text()='Cảnh báo Autotest BA_GPS V2']"
hopthoai_input = By.XPATH, "//*[@id='editable-message-text']"
hopthoai_iconlink = By.XPATH, "//*[@class='icon icon-attach']"
hopthoai_iconlink_file = By.XPATH, "//*[@id='attach-menu-controls']/div/div[2]//*[@class='icon icon-document']"
hopthoai_send = By.XPATH, "//*[@class='modal-dialog']/div[2]//*[text()='Send']"
login_iconchplay = "//*[@src='/Images/Login/iconLogin/SVG/photo_logo-android.svg']"
login_icontrangchu = "//*[@src='../../Images/Login/iconLogin/SVG/ic_home.svg']"
login_iconlienhezalo = "//*[@id='lnkCallZalo']"
login_iconzalo_x = "//*[@src='../../Images/X.png']"
login_iconsodienthoai = "//*[@src='../../Images/Login/iconLogin/png-phone-icon-13.jpg']"
login_bagps = "//*[@id='lnkGlobal']"
login_hotlinemuahang = "//*[text()='HOTLINE MUA HÀNG']"
login_muasamsanpham = "//*[text()='MUA SẮM SẢN PHẨM']"
login_thongtingiaiphap = "//*[text()='THÔNG TIN GIẢI PHÁP']"
check_login_lienhezalo = "//*[@class='contactzalo-header']"
check_login_trangchu = "//*[@class='row align-items-center col-mar-5']/div[1]/a/img"
check_login_chplay = "//*[text()='BA GPS']"
check_login_muasamsanpham = "//*[text()='Sản phẩm ']"
login_vechungtoi = "//*[text()='VỀ CHÚNG TÔI']"
check_login_vechungtoi = "//*[@class='about-1 bg']/div/div[1]/div[1]/h1"
login_mangluoi = "//*[text()='MẠNG LƯỚI']"
check_login_mangluoi = "//*[@class='q-main-page']/div/h1"
login_huongdansudung = "//*[text()='HƯỚNG DẪN SỬ DỤNG']"
check_login_huongdansudung = "//*[@id='channel-container']//*[text()='BA GPS']"
login_huongdandongphi = "//*[text()='HƯỚNG DẪN ĐÓNG PHÍ']"
check_login_huongdandongphi = "//*[@class='q-main-page v2 q-bg-gray']/div/div/div[1]/div/div[1]//*[text()='Hướng dẫn đóng phí dịch vụ BA GPS']"
linklienket_hopthoaizalo = "//*[@class='zalo-chat-widget']/iframe"
tiengviet = "/html/body/div/div[1]/div[1]/div/div/div[3]/div/div/div/div[1]"
english = "//*[@class='box-wrapper']//*[@class='zbtn zbtn-clear']"
chatnhanh = "//*[text()='Chat nhanh']"
hopthoaizalo_dau3cham = "//*[@class='far fa-ellipsis-h']"
hopthoaizalo_ketthuccuoctrochuyen = "//*[text()='Kết thúc cuộc trò chuyện']"
hopthoaizalo_icondong = "//*[@class='fal fa-chevron-down']"
check_login_iconzalo_doingonngu = "//*[@class='box__welcome']/h1"
check_login_iconzalo_chatnhanh = "//*[text()='Xin chào! BA GPS rất vui được hỗ trợ bạn.']"
iconngonngu_tienganh = "//*[@id='header-authentication']/div[3]/a//*[@src='/icons/flags/110-united%20kingdom.png']"
taikhoan = "//*[@id='header-authentication']//*[@id='lblAccount']"
dangxuat = "//*[text()='Đăng xuất']"
check_login_giamsat = "//*[@id='master-menu']//*[text()='Giám sát']"
check_login_khac = "//*[@id='master-menu']//*[text()='KHÁC']"
check_login_quantri = "//*[@id='master-menu']//*[text()='Quản trị']"


##GIÁM SÁT
menu_giamsat = "//*[@id='master-menu']//*[text()='Giám sát']"
timkiem_icon = "//*[@class='SearchSelect']"
timkiem_icon_timxe = "//*[@id='SelectSearchType']//*[@id='liVehicle']"
timkiem_icon_timdiachi = "//*[@id='SelectSearchType']//*[@id='liAddress']"
timkiem_icon_tendiem = "//*[@id='SelectSearchType']//*[@id='liLandmarkName']"
timkiem_icon_timtoado = "//*[@id='SelectSearchType']//*[@id='liLatLng']"
timkiem_timxe_input = "//*[@placeholder='Biển kiểm soát']"
timkiem_iconsearch = "//*[@id='SearchBtn']"
danhsachxe_tenphuongtien = "//*[@id='idClearOnline']/table/tbody/tr[1]/td[2]/div[2]"
saved = "//*[text()=' Saved ']"
timkiem_tendiem_input = "//*[@placeholder='Tên điểm']"
danhsachxe_tendiem = "//*[@class='ui-autocomplete ui-menu ui-widget ui-widget-content ui-corner-all']/li[1]/a"
timkiem_icon1 = "//*[@class='SearchSelect Landmark LandmarkName']"
timkiem_timtoado_input = "//*[@placeholder='Tọa độ']"
nhomxe = "//*[@id='ddlVehicleGroup']"
chonnhomxe1 = "//*[@id='ddlVehicleGroup']/option[2]"
bienso1 = "//*[@id='tblVehicleList']/tbody/tr[1]/td[2]/div[2]"
tatcanhomxe = "//*[@id='ddlVehicleGroup']//*[text()='Tất cả']"
trangthai = "//*[@id='ddlVehicleState']"
trangthai_batmay1 = "//*[@id='ddlVehicleState']//*[@value='1']"
checkiconxe_batmay1 = "//*[@id='tblVehicleList']/tbody/tr[2]/td[2]/div[1]/img"
soluong_dungbat = "//*[@id='tr_ColorSumVehicle']/td/table/tbody/tr[1]/td[2]/div"
soluong_dichuyen = "//*[@id='tr_ColorSumVehicle']/td/table/tbody/tr[1]/td[3]/div"
soluong_quatocdo = "//*[@id='tr_ColorSumVehicle']/td/table/tbody/tr[1]/td[4]/div"
tongsoxe_tren = "//*[@id='tr_ColorSumVehicle']/td/table/tbody/tr[1]/td[6]/div/div[1]"
trangthai_tatmay1 = "//*[@id='ddlVehicleState']//*[@value='2']"
soluong_dungtat = "//*[@id='tr_ColorSumVehicle']/td/table/tbody/tr[1]/td[1]/div"
trangthai_quatocdo = "//*[@id='ddlVehicleState']//*[@value='10']"
trangthai_dungdo = "//*[@id='ddlVehicleState']//*[@value='4']"
trangthai_dichuyen = "//*[@id='ddlVehicleState']//*[@value='3']"
trangthai_tatmay2 = "//*[@id='ddlVehicleState']//*[@value='8']"
trangthai_batmay2 = "//*[@id='ddlVehicleState']//*[@value='9']"
trangthai_mattinhieu = "//*[@id='ddlVehicleState']//*[@value='5']"
soluong_mattinhieu = "//*[@id='tr_ColorSumVehicle']/td/table/tbody/tr[1]/td[5]/div"
icon_xuatexcel = "//*[@class='tab-content']//*[@src='/Images/s_icon_excelExport.png']"
icon_capnhatdulieu = "//*[@class='tab-content']//*[@title='Cập nhật mới dữ liệu']"
icon_hientranghethong = "//*[@class='tab-content']//*[@title='Hiện trạng hệ thống']"
hientranghethong_x = "//*[@id='panelStatus']//*[@src='/Images/s_icon_x.png']"
icon_ynghiabieutuongxe = "//*[@src='/Images/s_icon_question.png']"
ynghiabieutuongxe_x = "//*[@id='closeMeaningOfIcon']"
xemlailotrinh = "//*[text()='Xem lại lộ trình ']"
tamgioganday = "//*[text()='8h gần đây']"
xemnhanh = "/html/body/div[17]/div[2]/div[2]/div[1]"
# xemnhanh = "//*[text()='Xem nhanh']"
check_popup_xemnhanhlotrinh = "//*[@id='divViewRouteQuick']"
xemnhanhlotrinh_xoa = "//*[@id='divViewRouteQuick']//*[@title='Xóa lộ trình']"
xemnhanhlotrinh_x = "//*[@id='divViewRouteQuick']//*[@title='Đóng']"
xemchitiettrencuasomoi = "/html/body/div[17]/div[2]/div[2]/div[2]"
# trongngay = "//*[text()='Trong ngày']"
trongngay = "/html/body/div[18]/div[2]/div[2]/div[9]"
tuychon = "/html/body/div[18]/div[2]/div[2]/div[10]"
nhapthongtinxe = "//*[text()='Nhập thông tin xe']"
popupthongtinxe_trongtaiinput = "//*[@id='vehicleInfomation']/div[2]/table/tbody/tr[6]/td[2]/input"
popupthongtinxe_capnhat = "//*[@id='vehicleInfomation']//*[@class='close']"
capnhatthanhcong = "//*[text()='Cập nhật thành công']"
ok = "//*[text()='OK']"
huy = "//*[@id='vehicleInfomation']//*[text()='Hủy']"
xemlotrinhnhieuxe = "/html/body/div[20]/div[2]/div[2]/div[3]"
hientrang = "/html/body/div[20]/div[2]/div[2]/div[4]"
check_danhsachxe_hientrang = "//*[@class='leaflet-popup-content-wrapper']"
gannhomdacbiet_hover = "//*[text()='Gán nhóm đặc biệt']"
themnhomdacbiet = "/html/body/div[19]/div[2]/div[2]/div"
check_themnhomdacbiet = "//*[@id='masterTwoColumnRight']/div/div/div/table/tbody/tr[1]/td"
anxe = "/html/body/div[20]/div[2]/div[2]/div[6]"
anxe_antoanbotrang = "//*[@id='panelHideVehicle']/div/div[2]/div[3]/div[2]/div/div[1]/input"
anxe_truyen = "//*[@id='panelHideVehicle']/div/div[2]/div[4]/div[2]/div/div[1]/input"
anxe_nguyennhan = "//*[@id='panelHideVehicle']/div/div[2]/div[5]/div[2]/select"
anxe_nguyennhan_xetainan = "//*[@id='panelHideVehicle']/div/div[2]/div[5]/div[2]/select//*[@value='3']"
anxe_ghichu = "//*[@id='panelHideVehicle']/div/div[2]/div[6]/div[2]/textarea"
luu = "//*[@id='panelHideVehicle']//*[text()='Lưu']"
icon_danhsachxedangan = "//*[@title='Danh sách xe đang ẩn']"
check_danhsachxedangan_tenphuongtien     = "//*[@id='panelFavourite']/div/div[8]/table/tbody/tr[1]/td[2]"
icon_danhsachxedangan_x = "//*[@id='panelFavourite']/div/div[8]/table/tbody/tr[1]//*[@src='/Images/delete_item.png']"
thongtinthietbi = "/html/body/div[20]/div[2]/div[2]/div[7]"
check_danhsachxe_thongtinthietbi = "//*[@id='divDeviceInfo']"
thongtinthietbi_x = "//*[@class='ui-icon ui-icon-closethick']"
xemhinhanhnhanh = "/html/body/div[20]/div[2]/div[2]/div[8]"
xemanhcamera = "/html/body/div[20]/div[2]/div[2]/div[9]"
check_xemanhcamera_tuchoitruycap = "//*[text()='Từ chối truy cập']"
xemanhcamera_nd10 = "/html/body/div[20]/div[2]/div[2]/div[10]"
check_xemanhcamera_nd10 = "//*[@id='cameraLeftPanel']/div/table/tbody/tr[1]/td"
giamsatcamera_nd10 = "/html/body/div[20]/div[2]/div[2]/div[11]"
check_giamsatcamera_nd10 = "//*[text()='Giám sát video']"
bieudonhienlieu = "/html/body/div[20]/div[2]/div[2]/div[12]"
check_popupbieudonhienlieu = "//*[@id='popUpBox']/div/div/div[2]"
bieudonhienlieu_x = "//*[@id='popUpBox']//*[@title='Đóng']"
bieudonhienlieumoi = "/html/body/div[20]/div[2]/div[2]/div[13]"
check_popupbieudonhienlieumoi = "//*[@class='popUpBox popup-online ui-draggable']/div/div/div[1]"
bieudonhienlieumoi_x = "//*[@class='popUpBox popup-online ui-draggable']//*[@title='Đóng']"
bieudonhietdo = "/html/body/div[20]/div[2]/div[2]/div[14]"
check_popupbieudonhietdo = "//*[text()='Xe chưa cấu hình hiển thị nhiệt độ.']"
khoangcachdencacxe = "/html/body/div[20]/div[2]/div[2]/div[15]"
khoancach_ngan = "//*[@class='float-left;']/table/tbody/tr[1]/td[1]"
khoancach_dai = "//*[@class='float-left;']/table/tbody/tr[1]/td[2]/div[3]"
button_keo = "//*[@class='float-left;']/table/tbody/tr[1]/td[2]/div[1]/div/a"
khoancachdencacxe_x = "//*[@id='panelMeasureDistance']/div/div[1]//*[@src='/Images/s_icon_x.png']"
khoangcachdencacdiem = "/html/body/div[20]/div[2]/div[2]/div[16]"
goicuoccamera = "//*[@pid='panelTabBacam1']"
popupthongtinxe_masorieng = "//*[@id='panelTabNormal1']/div[1]/div[1]"
popupthongtinxe_bienso = "//*[@id='panelTabNormal1']/div[1]/div[2]"
popupthongtinxe_giocapnhat = "//*[@id='panelTabNormal1']/div[1]/div[5]"
popupthongtinxe_vantocgps = "//*[@id='panelTabNormal1']/div[1]/div[8]"
popupthongtinxe_vantocco = "//*[@id='panelTabNormal1']/div[1]/div[11]"
popupthongtinxe_dungdo = "//*[@id='panelTabNormal1']/div[1]/div[14]"
popupthongtinxe_dangdo = "//*[@id='panelTabNormal1']/div[1]/div[16]"
popupthongtinxe_kmtrongngay = "//*[@id='panelTabNormal1']/div[1]/div[17]"
popupthongtinxe_may = "//*[@id='panelTabNormal1']/div[1]/div[20]"
popupthongtinxe_dieuhoa = "//*[@id='panelTabNormal1']/div[1]/div[23]"
popupthongtinxe_diachi = "//*[@id='panelTabNormal1']/div[1]/div[26]"
popupthongtinxe_nhienlieu = "//*[@id='panelTabNormal1']/div[1]/div[29]/div[2]"
popupthongtinxe_laixe = "//*[@id='panelTabNormal1']/div[2]/div[2]"
popupthongtinxe_dienthoai = "//*[@id='panelTabNormal1']/div[2]/div[4]"
popupthongtinxe_giaypheplaixe = "//*[@id='panelTabNormal1']/div[2]/div[8]"
popupthongtinxe_quatocdo = "//*[@id='panelTabNormal1']/div[2]/div[11]"
popupthongtinxe_thoigianlaixelientuc = "//*[@id='panelTabNormal1']/div[2]/div[14]"
popupthongtinxe_thoigianlaixetrongngay = "//*[@id='panelTabNormal1']/div[2]/div[17]"
popupthongtinxe_thongtinthenho = "//*[@id='panelTabNormal1']/div[2]/div[20]"
popupthongtinxe_soquanly = "//*[@id='panelTabNormal1']/div[2]/div[23]"
popupthongtinxe_thongtinphi = "//*[@id='panelTabNormal1']/div[3]/div[2]"
popupthongtinxe_goicuocdichvuvienthong = "//*[@id='panelTabBacam1']/div[1]/div[5]"
popupthongtinxe_nhamang = "//*[@id='panelTabBacam1']/div[1]/div[8]"
popupthongtinxe_dungluonggoicuoc = "//*[@id='panelTabBacam1']/div[1]/div[11]"
popupthongtinxe_songayluutru = "//*[@id='panelTabBacam1']/div[1]/div[14]"
popupthongtinxe_sokenhluutru = "//*[@id='panelTabBacam1']/div[1]/div[17]"
popupthongtinxe_tinhnangdinhvi = "//*[@id='panelTabBacam1']/div[1]/div[20]"
popupthongtinxe_tinhnanganh = "//*[@id='panelTabBacam1']/div[1]/div[23]"
popupthongtinxe_tinhnangvideo = "//*[@id='panelTabBacam1']/div[1]/div[26]"
popupthongtinxe_kenhlapcamera = "//*[@id='panelTabBacam1']/div[1]/div[32]"
popupthongtinxe_kenhhoatdong = "//*[@id='panelTabBacam1']/div[1]/div[35]"
popupthongtinxe_dungluongocung = "//*[@id='panelTabBacam1']/div[1]/div[38]"
popupthongtinxe_mangketnoi = "//*[@id='panelTabBacam1']/div[1]/div[41]"
icon_hientranghethong1 = "//*[@class='toolboxRightButton']//*[@src='/Images/btn_rightPanel_hienTrang.png']"
hientranghethong_biensoinput = "//*[@id='panelStatus']/div/div[4]/table/tbody/tr/td[4]/input"
timkiem_icon2 = "//*[@class='SearchSelect']/span"
danhsachxe2g_x = "//*[@id='divListVehicle2G']//*[@src='/Images/s_icon_x.png']"
chuotphaimap_phongto = "/html/body/div[21]/div[2]/div[2]/div[1]"
chuotphaimap_thunho = "/html/body/div[21]/div[2]/div[2]/div[2]"
canhbaoquatocdo_x = "//*[@id='warningSpeedOverTimeLine']//*[@src='/icons/IconX.png']"
chuotphaimap_trungtamoday = "/html/body/div[21]/div[2]/div[2]/div[3]"









