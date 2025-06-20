import logging
from selenium.common.exceptions import NoSuchElementException
import var
import time
import openpyxl
import subprocess
from selenium.webdriver.common.keys import Keys
import os
from selenium.webdriver.common.by import By
from retry import retry
import module_gpsv2
from playsound import playsound
from gtts import gTTS
import json
import requests
from requests.auth import HTTPBasicAuth

def timerun():
    while True:
        time.sleep(3)
        timerun = time.strftime("%H:%M:%S", time.localtime())
        print("Thời gian hiện tại:", timerun)
        print("Thời gian chạy tool:", var.timerun)
        var.writeData(var.path_luutamthoi, "Sheet1", 47, 2, timerun)
        if var.timerun == "":
            var.writeData(var.path_luutamthoi, "Sheet1", 47, 2, timerun)
        else:
            var.writeData(var.path_luutamthoi, "Sheet1", 47, 2, var.timerun)


        if var.timerun == time.strftime("%H:%M", time.localtime()):
            break
        if var.timerun == "":
            break



def clear_log():
    logging.basicConfig(handlers=[logging.FileHandler(filename=var.logpath,
                                                      encoding='utf-8', mode='w')],  # mode='a+', w
                        format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                        datefmt="%F %A %T",
                        level=logging.INFO)




def clearData(file, sheetName, ketqua, trangthai, tenanh):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 9
    while (i < 1000):
        i += 1
        i = str(i)
        sheet["F"+i] = ketqua
        sheet["G"+i] = trangthai
        sheet["M"+i] = tenanh
        i = int(i)
    wordbook.save(file)




def clearData_luutamthoi(file,sheetName, api, web, popup):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 6
    while (i < 37):
        i += 1
        i = str(i)
        sheet["B"+i] = api
        sheet["C"+i] = web
        sheet["D"+i] = popup
        i = int(i)
    wordbook.save(file)




def getRowCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)



def getColumnCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)



def readData(file,sheetName,rownum,columnno):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value


def writeData(file,sheetName,caseid,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 0
    while (i < 5000):
        i += 1
        i = str(i)
        if sheet["A"+i].value == caseid:
            i = int(i)
            sheet.cell(row=i, column=columnno).value = data
            break
        i = int(i)
    wordbook.save(file)








def generate_audio(text, filename):
    tts = gTTS(text=text, lang='vi')
    tts.save(filename)
    print(f"Đã tạo file âm thanh: {filename}")


def allow_mic(driver2):
    try:
        allow_button = driver2.ele('xpath://button[contains(text(), "Allow while visiting the site")]', timeout=5)
        if allow_button:
            allow_button.click()
            print("Đã bấm Cho phép dùng Microphone.")
    except Exception as e:
        print(f"Không thấy popup xin quyền micro: {e}")


def play_mp3_hidden(filepath):
    playsound(filepath)


def tele_search(tag, phone, data):
    global driver2
    driver2.ele(var.tele_search_input).clear()
    time.sleep(0.5)
    driver2.ele(var.tele_search_input).input(tag)
    time.sleep(1)
    driver2.ele(var.tele_search_name1, timeout=10).click()
    time.sleep(2)
    driver2.ele(var.tele_profile_name).click()
    time.sleep(2.5)
    check_phone = driver2.ele(var.tele_profile_phone).text
    check_tag = driver2.ele(var.tele_profile_tag).text
    print(check_phone)
    print(phone)
    print(check_tag)
    print(tag)
    if (check_phone == phone) and (check_tag == tag):
        driver2.ele(var.hopthoai_input).input(data)
        driver2.ele(var.hopthoai_input).input(Keys.ENTER)
        time.sleep(1)



def tele_call(tag, phone, count, cases):
    global driver2
    check_phone = driver2.ele(var.tele_profile_phone).text
    check_tag = driver2.ele(var.tele_profile_tag).text
    print(check_phone)
    print(phone)
    print(check_tag)
    print(tag)
    if (check_phone == phone) and (check_tag == tag):
        driver2.ele(var.tele_profile_call).click()
        time.sleep(1)
        print("n10")
        allow_mic(driver2)
        n = 0
        while n < 15:
            n += 1
            time.sleep(1)
            try:
                check_call = driver2.ele(var.calling).text
                print("n11")
                if check_call not in ["waiting...", "exchanging encryption keys...", "Connected"]:
                    # if check_call not in ["Ringing...", "Connecting..."]:
                    generate_audio(f"Bạn đang có {count} lỗi cần xử lý, mã lỗi: {cases}", var.uploadpath + "output.mp3")
                    print("n12")
                    play_mp3_hidden(var.uploadpath + "output.mp3")
                    print("n13")
                    time.sleep(2)
                    var.writeData(var.path_luutamthoi, "Sheet1", 93, 2, "Đã gọi")
                    break
            except Exception as e:
                print(f"Lỗi kiểm tra cuộc gọi: {e}")
                pass

    driver2.ele(var.tele_profile_end_call).click()
    time.sleep(2)



def open_tele():
    from DrissionPage import ChromiumPage, ChromiumOptions
    do1 = ChromiumOptions().set_paths(local_port=9201, user_data_path=r""+var.uploadpath+"Profile 30""")
    driver2 = ChromiumPage(addr_or_opts=do1)
    driver2.get("https://web.telegram.org/a/")
    time.sleep(2)
    var.driver.close()



def call_telegram():
    from DrissionPage import ChromiumPage, ChromiumOptions
    do1 = ChromiumOptions().set_paths(local_port=9201, user_data_path=r""+var.uploadpath+"Profile 30""")
    global driver2
    driver2 = ChromiumPage(addr_or_opts=do1)
    clearData_luutamthoi2(var.path_luutamthoi, "Sheet1", "", "", "", "", "")
    module_gpsv2.retest_serious()



    driver2.get("https://web.telegram.org/a/")
    time.sleep(2)
    try:
        driver2.ele("//*[text()='OK']").click()
    except:
        pass

    var.writeData(var.path_luutamthoi, "Sheet1", 93, 2, "Đang check cuộc gọi")
    count = int(var.readData(var.path_luutamthoi, 'Sheet1', 91, 2))
    data = str(var.readData(var.path_luutamthoi, 'Sheet1', 92, 2))
    cases = str(var.readData(var.path_luutamthoi, 'Sheet1', 90, 3))


    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 3
    while rownum < 6:
        rownum += 1
        row_str = str(rownum)
        phone_tag = sheet["N" + row_str].value
        print(f"Dòng {row_str}, Phone_tag: {phone_tag}")
        phone, tag = phone_tag.split('(')
        phone = phone.strip()
        tag = tag.replace(')', '').strip()
        print(phone)
        print(tag)
        try:
            tele_search(tag, phone, data)
            tele_call(tag, phone, count, cases)
            check_call = str(var.readData(var.path_luutamthoi, 'Sheet1', 93, 2))
            if check_call == "Đã gọi":
                break

        except Exception as e:
            print(f"Lỗi kiểm tra cuộc gọi: {e}")
            pass



def check_info_telegram():
    global driver2
    driver2.get("https://web.telegram.org/a/")










def clearData_luutamthoi2(file,sheetName, column1, column2, column3, column4, column5):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 89
    while (i < 100):
        i += 1
        i = str(i)
        sheet["B"+i] = column1
        sheet["C"+i] = column2
        sheet["D"+i] = column3
        sheet["E"+i] = column4
        sheet["F"+i] = column5
        i = int(i)
    wordbook.save(file)





# @retry(tries=2, delay=2, backoff=1, jitter=5, )
# def notification_telegram():
#     from DrissionPage import ChromiumPage, ChromiumOptions
#     do1 = ChromiumOptions().set_paths(local_port=9201, user_data_path=r""+var.uploadpath+"Profile 30""")
#     driver2 = ChromiumPage(addr_or_opts=do1)
#
#
#     wordbook = openpyxl.load_workbook(var.checklistpath)
#     sheet = wordbook.get_sheet_by_name("Checklist")
#     module_gpsv2.check_casenone()
#     module_gpsv2.change_casenone()
#     module_gpsv2.check_casefail()
#     module_gpsv2.check_casepass()
#
#     mucnghiemtrong = str(var.readData(var.path_luutamthoi, 'Sheet1', 65, 2))
#     tong_case_trong = str(var.readData(var.path_luutamthoi, 'Sheet1', 66, 2))
#
#     case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 77, 2))
#     case_pass = str(var.readData(var.path_luutamthoi, 'Sheet1', 87, 2))
#
#     thoigianbatdauchay = str(var.readData(var.path_luutamthoi , 'Sheet1', 47, 2))
#
#
#     # if case_fail >= 1:
#     time_string1 = time.strftime("%d/%m/%Y, ", time.localtime())
#     time_string1 = str(time_string1)
#     time_string2 = time.strftime("%H:%M", time.localtime())
#     time_string2 = str(time_string2)
#     print("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
#                                               "\n- LinkTest: " + var.linktest+
#                                               "\n- ModeTest: " + var.modetest+
#                                               "\n- Số case Pass: " + case_pass+
#                                               "\n- Số case False: "+ case_fail+
#                                               "\n- Số case False nghiêm trọng: "+ mucnghiemtrong)
#
#
#     if int(case_fail) >= 1 or int(tong_case_trong)>=1:
#         print("đã vào telegram")
#         driver2.get("https://web.telegram.org/a/")
#         time.sleep(2)
#         case_pass = str(case_pass)
#         case_fail = str(case_fail)
#         try:
#             driver2.ele("//*[text()='OK']").click()
#         except:
#             pass
#
#         # if var.linktest[0:27] == "https://testgps2.binhanh.vn":
#         if var.linktest[0:22] == "https://gps.binhanh.vn":
#             driver2.ele(var.hopthoai).click()
#         else:
#             driver2.ele(var.hopthoai1).click()
#         time.sleep(0.5)
#         time_string1 = time.strftime("%d/%m/%Y, ", time.localtime())
#         time_string1 = str(time_string1)
#         time_string2 = time.strftime("%H:%M", time.localtime())
#         time_string2 = str(time_string2)
#         driver2.ele(var.hopthoai_input).input("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
#                                                   "\n- LinkTest: " + var.linktest+
#                                                   "\n- ModeTest: " + var.modetest+
#                                                   "\n- Số case Pass: " + case_pass+
#                                                   "\n- Số case False: "+ case_fail+
#                                                   "\n- Số case False nghiêm trọng: "+ mucnghiemtrong)
#         driver2.ele(var.hopthoai_input).input(Keys.ENTER)
#         time.sleep(1)
#         driver2.ele(var.hopthoai_iconlink).click()
#         time.sleep(1)
#         driver2.ele(var.hopthoai_iconlink_file).click()
#         time.sleep(1)
#         subprocess.Popen(var.uploadpath+"checklist_bagps.exe")
#         time.sleep(1)
#         driver2.ele(var.hopthoai_send).click()
#         time.sleep(2)
#         driver2.ele(var.hopthoai_iconlink).click()
#         time.sleep(1)
#         driver2.ele(var.hopthoai_iconlink_file).click()
#         time.sleep(1)
#         subprocess.Popen(var.uploadpath+"ba_log.exe")
#         time.sleep(1)
#         driver2.ele(var.hopthoai_send).click()
#         time.sleep(1)
#
#         time.sleep(30)
#         driver2.close()

@retry(tries=2, delay=2, backoff=1, jitter=5, )
def notification_telegram():
    from DrissionPage import ChromiumPage, ChromiumOptions
    do1 = ChromiumOptions().set_paths(local_port=9201, user_data_path=r""+var.uploadpath+"Profile 30""")
    driver2 = ChromiumPage(addr_or_opts=do1)


    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    module_gpsv2.check_casenone()
    module_gpsv2.change_casenone()
    module_gpsv2.check_casefail()
    module_gpsv2.check_casepass()

    mucnghiemtrong = str(var.readData(var.path_luutamthoi, 'Sheet1', 65, 2))
    tong_case_trong = str(var.readData(var.path_luutamthoi, 'Sheet1', 66, 2))

    case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 77, 2))
    case_pass = str(var.readData(var.path_luutamthoi, 'Sheet1', 87, 2))

    thoigianbatdauchay = str(var.readData(var.path_luutamthoi , 'Sheet1', 47, 2))


    # if case_fail >= 1:
    time_string1 = time.strftime("%d/%m/%Y, ", time.localtime())
    time_string1 = str(time_string1)
    time_string2 = time.strftime("%H:%M", time.localtime())
    time_string2 = str(time_string2)
    print("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                              "\n- LinkTest: " + var.linktest+
                                              "\n- ModeTest: " + var.modetest+
                                              "\n- Số case Pass: " + case_pass+
                                              "\n- Số case False: "+ case_fail+
                                              "\n- Số case False nghiêm trọng: "+ mucnghiemtrong)

    driver2.get("https://web.telegram.org/a/")
    time.sleep(3)
    try:
        driver2.ele("//*[text()='OK']").click()
        time.sleep(2)
    except:
        pass

    if var.linktest[0:22] == "https://gps.binhanh.vn":
        if int(case_fail) >= 1:
            driver2.ele(var.hopthoai).click()
            time.sleep(1.5)
            time_string1 = time.strftime("%d/%m/%Y, ", time.localtime())
            time_string1 = str(time_string1)
            time_string2 = time.strftime("%H:%M", time.localtime())
            time_string2 = str(time_string2)
            driver2.ele(var.hopthoai_input).input("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                                      "\n- LinkTest: " + var.linktest+
                                                      "\n- ModeTest: " + var.modetest+
                                                      "\n- Số case Pass: " + case_pass+
                                                      "\n- Số case False: "+ case_fail+
                                                      "\n- Số case False nghiêm trọng: "+ mucnghiemtrong)
            driver2.ele(var.hopthoai_input).input(Keys.ENTER)
            time.sleep(1)
            driver2.ele(var.hopthoai_iconlink).click()
            time.sleep(1)
            driver2.ele(var.hopthoai_iconlink_file).click()
            time.sleep(1)
            subprocess.Popen(var.uploadpath+"checklist_bagps.exe")
            time.sleep(1)
            driver2.ele(var.hopthoai_send).click()
            time.sleep(2)
            driver2.ele(var.hopthoai_iconlink).click()
            time.sleep(1)
            driver2.ele(var.hopthoai_iconlink_file).click()
            time.sleep(1)
            subprocess.Popen(var.uploadpath+"ba_log.exe")
            time.sleep(1)
            driver2.ele(var.hopthoai_send).click()
            time.sleep(1)

            time.sleep(30)
            driver2.close()
    else:
        driver2.ele(var.hopthoai1).click()
        time.sleep(1.5)
        time_string1 = time.strftime("%d/%m/%Y, ", time.localtime())
        time_string1 = str(time_string1)
        time_string2 = time.strftime("%H:%M", time.localtime())
        time_string2 = str(time_string2)
        driver2.ele(var.hopthoai_input).input(
            "- DateTest : " + time_string1 + "" + thoigianbatdauchay + " - " + time_string2 +
            "\n- LinkTest: " + var.linktest +
            "\n- ModeTest: " + var.modetest +
            "\n- Số case Pass: " + case_pass +
            "\n- Số case False: " + case_fail +
            "\n- Số case False nghiêm trọng: " + mucnghiemtrong)
        driver2.ele(var.hopthoai_input).input(Keys.ENTER)
        time.sleep(1)
        driver2.ele(var.hopthoai_iconlink).click()
        time.sleep(1)
        driver2.ele(var.hopthoai_iconlink_file).click()
        time.sleep(1)
        subprocess.Popen(var.uploadpath + "checklist_bagps.exe")
        time.sleep(1)
        driver2.ele(var.hopthoai_send).click()
        time.sleep(2)
        driver2.ele(var.hopthoai_iconlink).click()
        time.sleep(1)
        driver2.ele(var.hopthoai_iconlink_file).click()
        time.sleep(1)
        subprocess.Popen(var.uploadpath + "ba_log.exe")
        time.sleep(1)
        driver2.ele(var.hopthoai_send).click()
        time.sleep(30)
        driver2.close()










def viber_send_text():
    wordbook = openpyxl.load_workbook(var.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    module_gpsv2.check_casenone()
    module_gpsv2.change_casenone()
    module_gpsv2.check_casefail()
    module_gpsv2.check_casepass()

    mucnghiemtrong = str(var.readData(var.path_luutamthoi, 'Sheet1', 65, 2))
    tong_case_trong = str(var.readData(var.path_luutamthoi, 'Sheet1', 66, 2))

    case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 77, 2))
    case_pass = str(var.readData(var.path_luutamthoi, 'Sheet1', 87, 2))

    thoigianbatdauchay = str(var.readData(var.path_luutamthoi , 'Sheet1', 47, 2))

    # if case_fail >= 1:
    time_string1 = time.strftime("%d/%m/%Y, ", time.localtime())
    time_string1 = str(time_string1)
    time_string2 = time.strftime("%H:%M", time.localtime())
    time_string2 = str(time_string2)
    print("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                              "\n- LinkTest: " + var.linktest+
                                              "\n- ModeTest: " + var.modetest+
                                              "\n- Số case Pass: " + case_pass+
                                              "\n- Số case False: "+ case_fail+
                                              "\n- Số case False nghiêm trọng: "+ mucnghiemtrong)

    AUTH_TOKEN = "54c3dd27a1b404ea-12a24f08fb9f6d31-2de1680f6bbb4010"  # id Group QA Test
    FROM_USER_ID = "q1VLwFu1K0sq/MQm+lsG0A=="

    if var.linktest[0:22] == "https://gps.binhanh.vn":
        if int(case_fail) >= 1:
            AUTH_TOKEN = "54c3ea38cef1116a-f3e8f9c4b7180a2c-af4f51098941df75"  # id Cảnh báo Autotest BA_GPS V2
            FROM_USER_ID = "JY+dUO0MPuuP0Zfqxme+eA=="


    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    print(AUTH_TOKEN)
    print(FROM_USER_ID)


    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return


    # 2. Gửi tin nhắn văn bản
    send_url = "https://chatapi.viber.com/pa/post"
    payload = {
        "auth_token": AUTH_TOKEN,
        "from": FROM_USER_ID,
        "type": "text",
        "text": ("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                                      "\n- LinkTest: " + var.linktest+
                                                      "\n- ModeTest: " + var.modetest+
                                                      "\n- Số case Pass: " + case_pass+
                                                      "\n- Số case False: "+ case_fail+
                                                      "\n- Số case False nghiêm trọng: "+ mucnghiemtrong)}

    headers = {
        "Content-Type": "application/json"}

    response = requests.post(send_url, json=payload, headers=headers)

    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())


def check_user_id():
    res = requests.post(
        "https://chatapi.viber.com/pa/get_account_info",
        headers={"X-Viber-Auth-Token": "54c3ea38cef1116a-f3e8f9c4b7180a2c-af4f51098941df75"}
    )
    print(res.json())


# def upload_to_catbox(file_path):
#     url = "https://catbox.moe/user/api.php"
#     files = {
#         'fileToUpload': open(file_path, 'rb')
#     }
#     data = {
#         'reqtype': 'fileupload'
#     }
#
#     try:
#         response = requests.post(url, files=files, data=data)
#         response.raise_for_status()
#     except requests.RequestException as e:
#         print("❌ Lỗi upload:", e)
#         return None
#
#     direct_link = response.text.strip()
#     # Catbox trả về URL trực tiếp của file, vd: https://files.catbox.moe/abc123.png
#     print(f"✅ Upload thành công! Link tải trực tiếp:\n{direct_link}")
#     return direct_link


def upload_pixeldrain_auth(filepath):
    API_KEY = "c567bb13-f4c0-4aac-b9bd-c8add1e467fc"  # Thay bằng key thật

    with open(filepath, "rb") as f:
        res = requests.post(
            "https://pixeldrain.com/api/file",
            files={"file": f},
            auth=HTTPBasicAuth('', API_KEY)
        )
        res_json = json.loads(res.text)
        file_id = res_json["id"]
        link_download = (f"https://pixeldrain.com/api/file/{file_id}")
        print(link_download)
        return link_download



def send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, file_path):
    file_url = upload_pixeldrain_auth(file_path)
    if not file_url:
        print("⚠️ Không thể upload file. Hủy gửi.")
        return

    file_name = os.path.basename(file_path)
    file_size = os.path.getsize(file_path)


    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return

    # 2. Gửi tin nhắn văn bản
    payload = {
        "auth_token": AUTH_TOKEN,
        "from": FROM_USER_ID,           # Viber user ID người nhận
        "min_api_version": 1,
        "tracking_data": "",               # Có thể để chuỗi rỗng nếu không dùng tracking
        "type": "file",
        "media": file_url,
        "size": file_size,
        "file_name": file_name}

    headers = {
        "X-Viber-Auth-Token": AUTH_TOKEN,
        "Content-Type": "application/json"}

    response = requests.post("https://chatapi.viber.com/pa/post", json=payload, headers=headers)
    print("📨 Phản hồi từ Viber:", response.status_code, response.json())
    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())


def viber_send_file():
    # ==== Ví dụ sử dụng ====
    AUTH_TOKEN = "54c3dd27a1b404ea-12a24f08fb9f6d31-2de1680f6bbb4010"  # id Group QA Test
    FROM_USER_ID = "q1VLwFu1K0sq/MQm+lsG0A=="
    case_fail = str(var.readData(var.path_luutamthoi, 'Sheet1', 77, 2))

    if var.linktest[0:22] == "https://gps.binhanh.vn":
        if int(case_fail) >= 1:
            AUTH_TOKEN = "54c3ea38cef1116a-f3e8f9c4b7180a2c-af4f51098941df75"  # id Cảnh báo Autotest BA_GPS V2
            FROM_USER_ID = "JY+dUO0MPuuP0Zfqxme+eA=="

    FILE_PATH_checklisst = var.checklistpath  # Thay bằng đường dẫn file thật
    FILE_PATH_log = var.logpath  # Thay bằng đường dẫn file thật

    print(var.checklistpath)
    print(var.logpath)


    send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, FILE_PATH_checklisst)
    send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, FILE_PATH_log)


def send_viber():
    viber_send_text()
    viber_send_file()













def delete_image():
    path = os.path.join(var.imagepath)
    l = os.listdir(path)
    print(l)
    for i in l:
        print(os.path.join(path, i))
        os.remove(os.path.join(path, i))




def delete_excel():
    path = os.path.join(var.excelpath)
    l = os.listdir(path)
    print(l)
    for i in l:
        print(os.path.join(path, i))
        os.remove(os.path.join(path, i))



def get_datachecklist(ma):
        wordbook = openpyxl.load_workbook(var.checklistpath)
        sheet = wordbook.get_sheet_by_name("Checklist")
        rownum = 9
        while (rownum < 3000):
            rownum += 1
            rownum = str(rownum)
            if sheet["A"+rownum].value == ma:
                tensukien = sheet["B"+rownum].value
                ketqua = sheet["E"+rownum].value
                print(ma)
                print(tensukien)
                print(ketqua)
                logging.info("đang chạy case: " + ma)
            rownum = int(rownum)



@retry(tries=3, delay=2, backoff=1, jitter=5, )
def swich_tab_0():
    var.driver.implicitly_wait(15)

    try:
        var.driver.implicitly_wait(1)
        var.driver.switch_to.alert.accept()
        time.sleep(1)
    except:
        pass


    try:
        var.driver.implicitly_wait(1)
        subprocess.Popen(var.uploadpath+"cancel.exe")
    except:
        pass

    # time.sleep(1)
    # try:
    #     var.driver.switch_to.window(var.driver.window_handles[0])
    #     time.sleep(1)
    #     var.driver.execute_script("window.open('');")
    #     time.sleep(2)
    #     var.driver.switch_to.window(var.driver.window_handles[1])
    #     var.driver.get("https://gps.binhanh.vn/Default.aspx")
    #     time.sleep(5)
    # except:
    #     var.driver.execute_script("window.open('');")
    #     time.sleep(2)
    #     var.driver.switch_to.window(var.driver.window_handles[1])
    #     var.driver.get("https://gps.binhanh.vn/Default.aspx")
    #     time.sleep(5)
    #     var.driver.switch_to.window(var.driver.window_handles[0])
    #
    #
    #
    # try:
    #     var.driver.switch_to.window(var.driver.window_handles[0])
    #     time.sleep(1)
    #     var.driver.execute_script("window.open('');")
    #     time.sleep(2)
    #     var.driver.switch_to.window(var.driver.window_handles[1])
    #     var.driver.get("https://gps.binhanh.vn/Default.aspx")
    #     time.sleep(5)
    # except:
    #     var.driver.execute_script("window.open('');")
    #     time.sleep(2)
    #     var.driver.switch_to.window(var.driver.window_handles[1])
    #     var.driver.get("https://gps.binhanh.vn/Default.aspx")
    #     time.sleep(5)
    #     var.driver.switch_to.window(var.driver.window_handles[0])



    # try:
    #     var.driver.switch_to.window(var.driver.window_handles[0])
    #     curr = var.driver.current_window_handle
    #     for handle in var.driver.window_handles:
    #         if handle != curr:
    #             var.driver.switch_to.window(handle)
    #             var.driver.close()
    #             time.sleep(1)
    #     var.driver.switch_to.window(curr)
    #     time.sleep(0.5)
    #
    # except:
    #     var.driver.execute_script("window.open('');")
    #     var.driver.switch_to.window(var.driver.window_handles[-1])  # Chuyển đến tab mới nhất
    #     var.driver.get("https://gps.binhanh.vn/Default.aspx")
    #     time.sleep(5)
    #     var.driver.switch_to.window(var.driver.window_handles[0])

    try:
        var.driver.switch_to.window(var.driver.window_handles[0])
        time.sleep(0.5)
    except:
        curr = var.driver.current_window_handle
        for handle in var.driver.window_handles:
            if handle != curr:
                var.driver.switch_to.window(handle)
                var.driver.close()
                time.sleep(1)
        var.driver.switch_to.window(curr)
        time.sleep(1)

    # Khởi tạo lại trình duyệt
    var.restart_driver()
    var.driver.get(var.linktest)
    time.sleep(10)


def write_caseif():
    n = 53
    while (n < 70):
        var.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   if case == 'Report"+n+"':\n       caseid.caseid_report"+n+"(self)\nexcept:\n    pass")
        n = int(n)




def write_caseif1():
    n = 53
    while (n < 70):
        var.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   caseid.caseid_report"+n+"(self)\nexcept:\n     pass")
        n = int(n)



def write_caseif2():
    n = 53
    while (n < 70):
        var.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("caseid.caseid_report"+n+"(self)")
        n = int(n)



def write_result_text_try_if(code, eventname, result, path_module, path_text, check_result, name_image):
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        writeData(var.checklistpath, "Checklist", code, 6, check_text)

        if check_text == check_result:
            logging.info("True")
            writeData(var.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + name_image)
            writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var.driver.save_screenshot(var.imagepath + code + name_image)
        writeData(var.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var.checklistpath, "Checklist", code, 13, code + name_image)
    # chucnangkhac.write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_or(code, eventname, result, path_module, path_text, check_result, check_result2, name_image):
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        logging.info(check_result)
        logging.info(check_result2)
        writeData(var.checklistpath, "Checklist", code, 6, check_text)

        if check_text == check_result or check_result2:
            logging.info("True")
            writeData(var.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + name_image)
            writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var.driver.save_screenshot(var.imagepath + code + name_image)
        writeData(var.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var.checklistpath, "Checklist", code, 13, code + name_image)



def write_result_text_try_if_url(code, eventname, result, path_module, desire, name_image):
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_url = var.driver.current_url
        logging.info(check_url)
        logging.info(desire)
        writeData(var.checklistpath, "Checklist", code, 6, check_url)

        if check_url == desire:
            logging.info("True")
            writeData(var.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + name_image)
            writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var.driver.save_screenshot(var.imagepath + code + name_image)
        writeData(var.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var.checklistpath, "Checklist", code, 13, code + name_image)




def write_result_text_try_if_title(code, eventname, result, path_module, desire, name_image):
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_title = var.driver.title
        print(check_title)
        logging.info(check_title)
        logging.info(desire)
        writeData(var.checklistpath, "Checklist", code, 6, check_title)

        if check_title == desire:
            logging.info("True")
            writeData(var.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + name_image)
            writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var.driver.save_screenshot(var.imagepath + code + name_image)
        writeData(var.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var.checklistpath, "Checklist", code, 13, code + name_image)
    # chucnangkhac.write_result_text_try_if_title(code, eventname, result, "Quản trị - Danh sách xe",
    #                                        "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_other(code, eventname, result, path_module, path_text, check_result, name_image):
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        writeData(var.checklistpath, "Checklist", code, 6, check_text)

        if check_text != check_result:
            logging.info("True")
            writeData(var.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + name_image)
            writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var.driver.save_screenshot(var.imagepath + code + name_image)
        writeData(var.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var.checklistpath, "Checklist", code, 13, code + name_image)
    # chucnangkhac.write_result_text_try_if_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_displayed_try(code, eventname, result, path_module, path_text, name_image):
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_displayed = var.driver.find_element(By.XPATH, path_text).is_displayed()
        logging.info("True")
        writeData(var.checklistpath, "Checklist", code, 7, "Pass")
    except NoSuchElementException:
        logging.info("False")
        var.driver.save_screenshot(var.imagepath + code + name_image)
        writeData(var.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var.checklistpath, "Checklist", code, 13, code + name_image)

    # logging.info("Tìm biển kiểm soát - " + data['quantri']['bienkiemsoat'])
    # chucnangkhac.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var.check_hide_car, "_QuanTri_DsXe_AnXe.png")




def write_result_displayed_try_close(code, eventname, result, path_module, button_close, name_image):
    var.driver.implicitly_wait(1)
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        var.driver.find_element(By.XPATH, button_close).click()
    except:
        button = var.driver.find_element(By.XPATH, button_close)
        var.driver.execute_script("arguments[0].click();", button)
    time.sleep(1.5)
    try:
        var.driver.find_element(By.XPATH, button_close).click()
        logging.info("False")
        var.driver.save_screenshot(var.imagepath + code + name_image)
        writeData(var.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("True")
        writeData(var.checklistpath, "Checklist", code, 7, "Pass")


    # chucnangkhac.write_result_displayed_try_close(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var.button_close, "_QuanTri_DsXe_AnXe.png")




def write_result_not_displayed_try(code, eventname, result, path_module, path_text, name_image):
    var.driver.implicitly_wait(2)
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_not_displayed = var.driver.find_element(By.XPATH, path_text).is_displayed()
        logging.info("False")
        var.driver.save_screenshot(var.imagepath + code + name_image)
        writeData(var.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var.checklistpath, "Checklist", code, 13, code + name_image)
    except NoSuchElementException:
        logging.info("True")
        writeData(var.checklistpath, "Checklist", code, 7, "Pass")
    # chucnangkhac.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var.check_hide_car, "_QuanTri_DsXe_AnXe.png")



def write_result_not_displayed_try1(code, eventname, result, path_module, path_text, data, name_image):
    var.driver.implicitly_wait(2)
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_not_displayed = var.driver.find_element(By.XPATH, path_text).is_displayed()
        logging.info("False")
        var.driver.save_screenshot(var.imagepath + code + name_image)
        writeData(var.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var.checklistpath, "Checklist", code, 13, code + name_image)
    except NoSuchElementException:
        logging.info("True")
        writeData(var.checklistpath, "Checklist", code, 7, "Pass")
        writeData(var.checklistpath, "Checklist", code, 6, data)
    # chucnangkhac.write_result_not_displayed_try1(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var.check_hide_car, "_QuanTri_DsXe_AnXe.png")



def write_result_excelreport(code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        logging.info("Check vị trí: "+number_row+":  "+output+"")
        if str(sheet[column + number_column].value) == output:
            logging.info("True")
            writeData(var.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var.checklistpath, "Checklist", code, 6, "File Báo cáo sai ô " + number_row)
    # chucnangkhac.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")



def write_result_excelreport1(code, sheet, column, name_sheet, number_column, number_row, output, number_row2, output2):
    data_excel = str(sheet[number_row2].value)
    output2 = str(output2)

    print("Check vị trí: " + number_row + ":  " + output + "")
    print("Dữ liệu web: " +output2)
    print("Dữ liệu excel: " +data_excel)
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        logging.info("Check vị trí: "+number_row+":  "+output+"")
        logging.info("Dữ liệu excel: "+ data_excel)
        logging.info("Dữ liệu web: "+ output2)
        if str(sheet[column + number_column].value) == output and data_excel == output2:
            logging.info("True")
            writeData(var.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var.checklistpath, "Checklist", code, 6, "File Báo cáo sai ô " + number_row2)
    if output2 != data_excel:
        writeData(var.checklistpath, "Checklist", code, 6, "File Báo cáo sai ô " + number_row)
    # chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "D5", "Ngày tháng", "D10", activity_synthesis_group_report_day_month)




def write_result_excelreport2(code, output_web, output_excel, name_output):
    logging.info(name_output + " web: " + output_web)
    logging.info(name_output + " excel: " + output_excel)
    if output_web == output_excel:
        logging.info("True")
        writeData(var.checklistpath, "Checklist", code, 7, "Pass")
    else:
        logging.info("False")
        writeData(var.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var.checklistpath, "Checklist", code, 6, "File Báo cáo sai dữ liệu: \nweb: " + output_web + "\nexcel: " + output_excel)




def write_result_excel_checkweb(code, data_web, desire):
    logging.info("Dữ liệu web: " + data_web)
    logging.info("Dữ liệu mong muốn: " + desire)
    if data_web == desire:
        logging.info("True")
        writeData(var.checklistpath, "Checklist", code, 7, "Pass")
    else:
        logging.info("False")
        writeData(var.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var.checklistpath, "Checklist", code, 6, "File Báo cáo mất sai trường" + desire)




def write_result_excelreport_clear_data(code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        logging.info("Check vị trí: "+number_row+": "+output+"")
        if str(sheet[column + number_column].value) == output:
            logging.info("True")
            writeData(var.checklistpath, "Checklist", code, 6, " ")
            writeData(var.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var.checklistpath, "Checklist", code, 6, "File Báo cáo sai ô " + number_row)
    # chucnangkhac.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")





