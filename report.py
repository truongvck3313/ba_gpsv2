import logging
from selenium.webdriver.common.keys import Keys
import var
import time
import json
from selenium.webdriver.common.by import By
import chucnangkhac
import login
import openpyxl
from xls2xlsx import XLS2XLSX
import os
import shutil
import giamsat
import re












file_name = var.datatestpath
with open(file_name, 'r', encoding='utf-8') as f:
    data = json.load(f, strict=False)

logging.basicConfig(handlers=[logging.FileHandler(filename=var.logpath,
                                                  encoding='utf-8', mode='a+')],  # mode='a+'
                    format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                    datefmt="%F %A %T",
                    level=logging.INFO)



#hàm nhập từ ngày -1
def write_from_date(fromdate_input):
    var.driver.implicitly_wait(5)
    from_date = var.driver.find_element(By.XPATH, fromdate_input).get_attribute('value')
    print(from_date)
    from_date_day = from_date[0:2]
    print(from_date_day)
    from_date_month_year = from_date[2::]
    from_date_day1 = int(from_date_day) - 1
    print(from_date_day1)
    if from_date_day1 == 10:
        from_date_day1 = "09"
    if from_date_day1 == 9:
        from_date_day1 = "08"
    if from_date_day1 == 8:
        from_date_day1 = "07"
    if from_date_day1 == 7:
        from_date_day1 = "06"
    if from_date_day1 == 6:
        from_date_day1 = "05"
    if from_date_day1 == 5:
        from_date_day1 = "04"
    if from_date_day1 == 4:
        from_date_day1 = "03"
    if from_date_day1 == 3:
        from_date_day1 = "02"
    if from_date_day1 == 2:
        from_date_day1 = "01"
    if from_date_day1 == 1:
        from_date_day1 = "01"
    if from_date_day1 == 0:
        from_date_day1 = "01"
    # from_date_day2 = str(from_date_day1)
    from_date1 = str(from_date_day1) + str(from_date_month_year)
    print(from_date1)
    xoa = var.driver.find_element(By.XPATH, fromdate_input)
    xoa.send_keys(Keys.CONTROL, "a")
    var.driver.find_element(By.XPATH, fromdate_input).send_keys(from_date1)


def write_from_date_month(fromdate_month_input):
    var.driver.implicitly_wait(5)
    from_date = var.driver.find_element(By.XPATH, fromdate_month_input).get_attribute('value')
    print(from_date)
    from_date_day = from_date[0:2]
    print(from_date_day)
    from_date_month = from_date[3:5]
    print(from_date_month)
    from_date_year = from_date[6::]
    print(from_date_year)
    from_date_month = int(from_date_month) - 1
    print(from_date_month)
    if from_date_month == 9:
        from_date_month = "09"
    if from_date_month == 8:
        from_date_month = "08"
    if from_date_month == 7:
        from_date_month = "07"
    if from_date_month == 6:
        from_date_month = "06"
    if from_date_month == 5:
        from_date_month = "05"
    if from_date_month == 4:
        from_date_month = "04"
    if from_date_month == 3:
        from_date_month = "03"
    if from_date_month == 2:
        from_date_month = "02"
    if from_date_month == 1:
        from_date_month = "01"
    if from_date_month == 0:
        from_date_month = "01"
    from_date = from_date_day + str(from_date_month) + str(from_date_year)
    print(from_date)
    xoa = var.driver.find_element(By.XPATH, fromdate_month_input)
    xoa.send_keys(Keys.CONTROL, "a")
    var.driver.find_element(By.XPATH, fromdate_month_input).send_keys(from_date)





class synthesis_report:     #báo cáo tổng hợp

    def activity_synthesis_group_report(self, code, eventname, result):         #báo cáo doanh nghiệp - Báo cáo tổng hợp hoạt động (theo nhóm)
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.goto_43e02740)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
        var.driver.find_element(By.XPATH, var.managerment_report).click()
        time.sleep(4)
        var.driver.find_element(By.XPATH, var.activity_synthesis_report).click()
        time.sleep(4)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo odanh nghiệp - Báo cáo tổng hợp hoạt động (theo nhóm)",
                                              var.check_activity_synthesis_report, "BÁO CÁO TỔNG HỢP", "_BaoCaoDoanhNghiep_BaoCaoTongHopHoatDong.png")


    def activity_synthesis_group_report_search(self, code, eventname, result):          #báo cáo doanh nghiệp - Báo cáo tổng hợp hoạt động (theo nhóm) - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            write_from_date(var.activity_synthesis_group_report_fromdate_input)
            var.driver.find_element(By.XPATH, var.activity_synthesis_report_search).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.activity_synthesis_report).click()
            time.sleep(4)
            write_from_date(var.activity_synthesis_group_report_fromdate_input)
            var.driver.find_element(By.XPATH, var.activity_synthesis_report_search).click()
        time.sleep(8)
        chucnangkhac.write_result_displayed_try(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo tổng hợp hoạt động (theo nhóm)",
                                                var.check_activity_synthesis_report_search, "_BaoCaoDoanhNghiep_BaoCaoTongHopHoatDong_TimKiem.png")


    def activity_synthesis_group_report_downloadexcel(self, code, eventname, result):           #báo cáo doanh nghiệp - Báo cáo tổng hợp hoạt động (theo nhóm) - check file excel
        var.driver.implicitly_wait(4)
        chucnangkhac.delete_excel()
        try:
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.activity_synthesis_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.activity_synthesis_report_search).click()
            time.sleep(8)

        # del var.driver.requests
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(10)
        var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoTongHopHoatDong.png")
        x2x = XLS2XLSX(var.excelpath + "/ActivitySummaryNew_43E02740.xls")
        x2x.to_xlsx(var.excelpath + "/ActivitySummaryNew_43E02740.xlsx")

        # #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']
        wordbook = openpyxl.load_workbook(var.excelpath+"/ActivitySummaryNew_43E02740.xlsX")
        sheet = wordbook.get_sheet_by_name("BC Tổng hợp")
        activity_synthesis_group_report_time_from = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_time_from).text
        activity_synthesis_group_report_time_from = activity_synthesis_group_report_time_from + ":00"
        activity_synthesis_group_report_time_to = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_time_to).text
        activity_synthesis_group_report_time_to = activity_synthesis_group_report_time_to + ":00"
        activity_synthesis_group_report_from_address = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_from_address).text
        activity_synthesis_group_report_to_address = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_to_address).text
        activity_synthesis_group_report_time_run = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_time_run).text
        activity_synthesis_group_report_time_run = activity_synthesis_group_report_time_run + ":00"

        activity_synthesis_group_report_percent_run = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_percent_run).text
        activity_synthesis_group_report_percent_run = str(''.join(re.findall(r'\d+', activity_synthesis_group_report_percent_run)))

        activity_synthesis_group_report_time_stop = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_time_stop).text
        activity_synthesis_group_report_time_stop = activity_synthesis_group_report_time_stop + ":00"

        activity_synthesis_group_report_km_gps = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_km_gps).text
        activity_synthesis_group_report_km_gps = str(''.join(re.findall(r'\d+', activity_synthesis_group_report_km_gps)))

        activity_synthesis_group_report_km_engine = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_km_engine).text
        activity_synthesis_group_report_fuel_1km = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_fuel_1km).text
        activity_synthesis_group_report_fuel_consume = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_fuel_consume).text
        activity_synthesis_group_report_number_of_stop = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_number_of_stop).text
        activity_synthesis_group_report_turn_on_air_conditional = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_turn_on_air_conditional).text
        activity_synthesis_group_report_max_speed = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_max_speed).text
        activity_synthesis_group_report_average_speed = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_average_speed).text
        activity_synthesis_group_report_violate_continue = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_violate_continue).text
        activity_synthesis_group_report_increase_speed1 = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_increase_speed1).text
        activity_synthesis_group_report_increase_speed2 = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_increase_speed2).text
        activity_synthesis_group_report_violate_too_speed = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_violate_too_speed).text
        activity_synthesis_group_report_detail = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_detail).text
        activity_synthesis_group_report_route1 = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_route1).text
        activity_synthesis_group_report_route2 = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_route2).text
        activity_synthesis_group_report_stop = var.driver.find_element(By.XPATH, var.activity_synthesis_group_report_stop).text


        logging.info("Báo cáo doanh nghiệp - Báo cáo tổng hợp hoạt động (theo nhóm)")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "5"].value)
            print(sheet[column + "5"])
            chucnangkhac.write_result_excelreport_clear_data(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "D5", "Ngày tháng")
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "E5", "Giờ đi", "E8", activity_synthesis_group_report_time_from)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "F5", "Giờ đến", "F8", activity_synthesis_group_report_time_to)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "G5", "Địa điểm bắt đầu", "G8", activity_synthesis_group_report_from_address)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "H5", "Địa điểm kết thúc", "H8", activity_synthesis_group_report_to_address)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "I5", "Thời gian  lăn bánh", "I8", activity_synthesis_group_report_time_run)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "J5", "% Thời gian  lăn bánh", "J8", activity_synthesis_group_report_percent_run)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "K5", "Thời gian  dừng", "K8", activity_synthesis_group_report_time_stop)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "L5", "Km (GPS)", "L8", activity_synthesis_group_report_km_gps)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "M5", "Km cơ", "M8", activity_synthesis_group_report_km_engine)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "N5", "Định mức  nhiên liệu trên 1km", "N8", activity_synthesis_group_report_fuel_1km)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "O5", "Nhiên liệu  tiêu thụ  định mức", "O8", activity_synthesis_group_report_fuel_consume)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "P5", "SL  dừng đỗ", "P8", activity_synthesis_group_report_number_of_stop)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "Q5", "Bật điều hòa (phút)", "Q8", activity_synthesis_group_report_turn_on_air_conditional)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "R5", "Vận tốc cực đại", "R8", activity_synthesis_group_report_max_speed)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "S5", "Vận tốc TB", "S8", activity_synthesis_group_report_average_speed)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "T5", "Vi phạm LX liên tục(Lần)", "T8", activity_synthesis_group_report_violate_continue)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "U5", "Vi phạm tăng tốc độ đột ngột(Lần)", "U8", activity_synthesis_group_report_increase_speed1)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "V5", "Vi phạm giảm tốc độ đột ngột(Lần)", "V8", activity_synthesis_group_report_increase_speed2)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "W5", "Vi phạm quá tốc độ (Lần)", "W8", activity_synthesis_group_report_violate_too_speed)
        chucnangkhac.write_result_excel_checkweb(code, activity_synthesis_group_report_detail, "Chi tiết")
        chucnangkhac.write_result_excel_checkweb(code, activity_synthesis_group_report_route1, "Hành trình")
        chucnangkhac.write_result_excel_checkweb(code, activity_synthesis_group_report_route2, "Lộ trình")
        chucnangkhac.write_result_excel_checkweb(code, activity_synthesis_group_report_stop, "Dừng đỗ")


    def activity_synthesis_group_report_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report1)
        except:
            synthesis_report.activity_synthesis_group_report(self,"", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(2)
        logging.info("----------------------")
        logging.info("Báo cáo danh nghiệp - Báo cáo tổng hợp hoạt động (theo nhóm)")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.activity_synthesis_group_repor_checkbox_daymoth).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoTongHopHoatDong_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoTongHopHoatDong_AnHienCot.png")

        try:
            var.driver.find_element(By.XPATH, var.hide_column_cancel).click()
            time.sleep(1.5)
        except:
            pass







    def detailed_activity_report(self, code, eventname, result):            #báo cáo doanh nghiệp - Báo cáo chi tiết hoạt động
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.detailed_activity_report).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.detailed_activity_report).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo chi tiết hoạt động",
                                              var.check_detailed_activity_report, "BÁO CÁO CHI TIẾT HOẠT ĐỘNG", "_BaoCaoDoanhNghiep_BaoCaoChiTietHoatDong.png")


    def detailed_activity_report_checkbox(self, code, eventname, result, path_checkbox, path_text, path_module, check_result, name_image):          #báo cáo doanh nghiệp - Báo cáo chi tiết hoạt động - gộp số phút hoạt động/Chọn Khung giờ
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, path_checkbox).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.detailed_activity_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, path_checkbox).click()
        time.sleep(1.5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, path_module,
                                              path_text, check_result, name_image)

        if var.driver.find_element(By.XPATH, path_checkbox).is_selected() == True:
            var.driver.find_element(By.XPATH, path_checkbox).click()
        time.sleep(1.5)


    def detailed_activity_report_search(self, code, eventname, result):         #báo cáo doanh nghiệp - Báo cáo chi tiết hoạt động - tìm kiếm
        var.driver.implicitly_wait(10)
        try:
            var.driver.find_element(By.XPATH, var.detailed_activity_report_from_date)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.detailed_activity_report).click()
            time.sleep(5)

        write_from_date(var.activity_synthesis_group_report_fromdate_input)

        var.driver.find_element(By.XPATH, var.detailed_activity_report_search).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo chi tiết hoạt động",
                                              var.check_detailed_activity_report_search, "STT", "_BaoCaoDoanhNghiep_BaoCaoChiTietHoatDong_TimKiem.png")


    def detailed_activity_report_downloadexcel(self, code, eventname, result):          #báo cáo doanh nghiệp - Báo cáo chi tiết hoạt động - check file excel
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        time.sleep(1)
        try:
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.detailed_activity_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.detailed_activity_report_search).click()
            time.sleep(5)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(16)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaochitiethoatdong.xlsx"))

        # #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaochitiethoatdong.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        detailed_activity_report_type_vehicle = var.driver.find_element(By.XPATH, var.detailed_activity_report_type_vehicle).text
        detailed_activity_report_license_plate = var.driver.find_element(By.XPATH, var.detailed_activity_report_license_plate).text
        detailed_activity_report_group = var.driver.find_element(By.XPATH, var.detailed_activity_report_group).text
        detailed_activity_report_group = " "+ detailed_activity_report_group

        detailed_activity_report_day_month = var.driver.find_element(By.XPATH, var.detailed_activity_report_day_month).text
        detailed_activity_report_from_hour = var.driver.find_element(By.XPATH, var.detailed_activity_report_from_hour).text
        detailed_activity_report_to_hour = var.driver.find_element(By.XPATH, var.detailed_activity_report_to_hour).text

        detailed_activity_report_start_liters = var.driver.find_element(By.XPATH, var.detailed_activity_report_start_liters).text #số lít ban đầu
        detailed_activity_report_start_liters = detailed_activity_report_start_liters[0:3]
        detailed_activity_report_start_liters_excel = str(sheet["H6"].value)
        detailed_activity_report_start_liters_excel = detailed_activity_report_start_liters_excel[0:3]

        detailed_activity_report_end_liters = var.driver.find_element(By.XPATH, var.detailed_activity_report_end_liters).text
        detailed_activity_report_end_liters = detailed_activity_report_end_liters[0:3]
        detailed_activity_report_end_liters_excel = str(sheet["I6"].value)
        detailed_activity_report_end_liters_excel = detailed_activity_report_end_liters_excel[0:3]

        detailed_activity_report_time_hd = var.driver.find_element(By.XPATH, var.detailed_activity_report_time_hd).text

        detailed_activity_report_km_gps = var.driver.find_element(By.XPATH, var.detailed_activity_report_km_gps).text
        detailed_activity_report_km_gps = str(''.join(re.findall(r'\d+', detailed_activity_report_km_gps)))
        detailed_activity_report_km_gps_ecxel = str(sheet["K6"].value)
        detailed_activity_report_km_gps_ecxel = str(''.join(re.findall(r'\d+', detailed_activity_report_km_gps_ecxel)))

        detailed_activity_report_km_co = var.driver.find_element(By.XPATH, var.detailed_activity_report_km_co).text
        detailed_activity_report_km_co = str(''.join(re.findall(r'\d+', detailed_activity_report_km_co)))
        detailed_activity_report_km_co_ecxel = str(sheet["L6"].value)
        detailed_activity_report_km_co_ecxel = str(''.join(re.findall(r'\d+', detailed_activity_report_km_co_ecxel)))

        detailed_activity_report_fuel_1km = var.driver.find_element(By.XPATH, var.detailed_activity_report_fuel_1km).text
        detailed_activity_report_fuel1 = var.driver.find_element(By.XPATH, var.detailed_activity_report_fuel1).text #nhiên liệu tiêu thụ định mức

        detailed_activity_report_fuel2 = var.driver.find_element(By.XPATH, var.detailed_activity_report_fuel2).text #nhien liệu tiêu hao
        detailed_activity_report_fuel2 = str(''.join(re.findall(r'\d+', detailed_activity_report_fuel2)))
        detailed_activity_report_fuel2 = detailed_activity_report_fuel2[0:3]
        detailed_activity_report_fuel2_ecxel = str(sheet["O6"].value)
        detailed_activity_report_fuel2_ecxel = str(''.join(re.findall(r'\d+', detailed_activity_report_fuel2_ecxel)))
        detailed_activity_report_fuel2_ecxel = detailed_activity_report_fuel2_ecxel[0:3]

        detailed_activity_report_from_address = var.driver.find_element(By.XPATH, var.detailed_activity_report_from_address).text
        detailed_activity_report_to_address = var.driver.find_element(By.XPATH, var.detailed_activity_report_to_address).text
        detailed_activity_report_note = var.driver.find_element(By.XPATH, var.detailed_activity_report_note).text
        detailed_activity_report_route = var.driver.find_element(By.XPATH, var.detailed_activity_report_route).text

        logging.info("Báo cáo doanh nghiệp - Báo cáo chi tiết họat động")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "5"].value)
            print(sheet[column + "5"])
            chucnangkhac.write_result_excelreport_clear_data(code, sheet, column, 'Data', "5", "A5", "STT")
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "B5", "Loại phương tiện", "B6", detailed_activity_report_type_vehicle)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "C5", "Biển số", "C6", detailed_activity_report_license_plate)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "D5", "Nhóm", "D6", detailed_activity_report_group)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "E5", "Ngày tháng", "E6", detailed_activity_report_day_month)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "F5", "Từ giờ", "F6", detailed_activity_report_from_hour)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "G5", "Đến giờ", "G6", detailed_activity_report_to_hour)

            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "H5", "Số lít bắt đầu")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "I5", "Số lít kết thúc")

            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "J5", "Khoảng thời gian HĐ", "J6", detailed_activity_report_time_hd)
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "K5", "Km GPS")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "L5", "Km cơ")

            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "M5", "Định mức nhiên liệu trên 1km", "M6", detailed_activity_report_fuel_1km)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "N5", "Nhiên liệu tiêu thụ định mức", "N6", detailed_activity_report_fuel1)

            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "O5", "Nhiên liệu tiêu hao")
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "P5", "Địa chỉ đi", "P6", detailed_activity_report_from_address)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "Q5", "Địa chỉ đến", "Q6", detailed_activity_report_to_address)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "R5", "Ghi chú", "R6", detailed_activity_report_note)

        chucnangkhac.write_result_excel_checkweb(code, detailed_activity_report_route, "Lộ trình")
        chucnangkhac.write_result_excelreport2(code, detailed_activity_report_start_liters, detailed_activity_report_start_liters_excel, "Số lít bắt đầu")
        chucnangkhac.write_result_excelreport2(code, detailed_activity_report_end_liters, detailed_activity_report_end_liters_excel, "Số lít kết thúc")
        chucnangkhac.write_result_excelreport2(code, detailed_activity_report_km_gps, detailed_activity_report_km_gps_ecxel, "Km GPS")
        chucnangkhac.write_result_excelreport2(code, detailed_activity_report_km_co, detailed_activity_report_km_co_ecxel, "Km cơ")
        chucnangkhac.write_result_excelreport2(code, detailed_activity_report_fuel2, detailed_activity_report_fuel2_ecxel, "Nhin liệu tiêu hao")


    def detailed_activity_report_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report2)
        except:
            synthesis_report.detailed_activity_report(self,"", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(2)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo chi tiết hoạt động")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.detailed_activity_report_checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoChiTietHoatDong_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoChiTietHoatDong_AnHienCot.png")


        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass













    def report_km_activity_summary(self, code, eventname, result):          #báo cáo doanh nghiệp - Báo cáo tổng hợp km xe hoạt động
        var.driver.implicitly_wait(10)
        try:
            var.driver.find_element(By.XPATH, var.report_km_activity_summary).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.report_km_activity_summary).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo tổng hợp km xe hoạt động",
                                              var.check_report_km_activity_summary, "BÁO CÁO TỔNG HỢP KM XE HOẠT ĐỘNG", "_BaoCaoDoanhNghiep_BaoCaoTongHopKmHoatDong.png")


    def report_km_activity_summary_search(self, code, eventname, result, checkbox, path_check, desire, path_image):         # báo cáo doanh nghiệp - Báo cáo tổng hợp km xe hoạt động - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, checkbox).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_km_activity_summary).click()
            time.sleep(4)
        time.sleep(1)
        write_from_date(var.report_km_activity_summary_from_date)
        var.driver.find_element(By.XPATH, var.report_km_activity_summary_search).click()
        time.sleep(10)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo tổng hợp km xe hoạt động",
                                              path_check, desire, path_image)


    def report_km_activity_summary_downloadexcel1(self, code, eventname, result):           # báo cáo doanh nghiệp - Báo cáo tổng hợp km xe hoạt động - Chi tiết kích xung - checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_km_activity_summary).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail).click()
        time.sleep(0.5)
        var.driver.find_element(By.XPATH, var.report_km_activity_summary_search).click()
        time.sleep(6)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaotonghopkmhoatdong_chitietkichxung.xlsx"))
        # #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaotonghopkmhoatdong_chitietkichxung.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        report_km_activity_summary_detail_stt = var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail_stt).text
        report_km_activity_summary_detail_liscense_plate = var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail_liscense_plate).text
        report_km_activity_summary_detail_group = var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail_group).text
        report_km_activity_summary_detail_group = " " + report_km_activity_summary_detail_group

        report_km_activity_summary_detail_day_moth = var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail_day_moth).text
        report_km_activity_summary_detail_type_vehicle = var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail_type_vehicle).text
        report_km_activity_summary_detail_customer = var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail_customer).text

        report_km_activity_summary_detail_km_gps = var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail_km_gps).text
        report_km_activity_summary_detail_km_gps = str(''.join(re.findall(r'\d+', report_km_activity_summary_detail_km_gps)))
        report_km_activity_summary_detail_km_gps = report_km_activity_summary_detail_km_gps[0:3]
        report_km_activity_summary_detail_km_gps_excel = str(sheet["G6"].value)
        report_km_activity_summary_detail_km_gps_excel = str(''.join(re.findall(r'\d+', report_km_activity_summary_detail_km_gps_excel)))
        report_km_activity_summary_detail_km_gps_excel = report_km_activity_summary_detail_km_gps_excel[0:3]

        report_km_activity_summary_detail_km_co = var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail_km_co).text
        report_km_activity_summary_detail_km_co = str(''.join(re.findall(r'\d+', report_km_activity_summary_detail_km_co)))
        report_km_activity_summary_detail_km_co = report_km_activity_summary_detail_km_co[0:3]
        report_km_activity_summary_detail_km_co_excel = str(sheet["H6"].value)
        report_km_activity_summary_detail_km_co_excel = str(''.join(re.findall(r'\d+', report_km_activity_summary_detail_km_co_excel)))
        report_km_activity_summary_detail_km_co_excel = report_km_activity_summary_detail_km_co_excel[0:3]

        report_km_activity_summary_detail_time_on = var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail_time_on).text
        report_km_activity_summary_detail_time_off = var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail_time_off).text
        report_km_activity_summary_detail_detail = var.driver.find_element(By.XPATH, var.report_km_activity_summary_detail_detail).text

        logging.info("Báo cáo doanh nghiệp - Báo cáo tổng hợp km xe hoạt động - Chi tiết xung")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "5"].value)
            print(sheet[column + "5"])
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "A5", "STT", "A6", report_km_activity_summary_detail_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "B5", "Biển kiểm soát", "B6", report_km_activity_summary_detail_liscense_plate)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "C5", "Nhóm", "C6", report_km_activity_summary_detail_group)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "D5", "Ngày tháng", "D6", report_km_activity_summary_detail_day_moth)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "E5", "Loại phương tiện", "E6", report_km_activity_summary_detail_type_vehicle)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "F5", "Khách hàng", "F6", report_km_activity_summary_detail_customer)
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "G5", "Km GPS")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "H5", "Km cơ")
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "I5", "Thời gian kích xung", "I6", report_km_activity_summary_detail_time_on)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "J5", "Thời gian ngắt xung", "J6", report_km_activity_summary_detail_time_off)

        chucnangkhac.write_result_excel_checkweb(code, report_km_activity_summary_detail_detail, "Chi tiết kích xung")
        chucnangkhac.write_result_excelreport2(code, report_km_activity_summary_detail_km_gps, report_km_activity_summary_detail_km_gps_excel, "Km GPS")
        chucnangkhac.write_result_excelreport2(code, report_km_activity_summary_detail_km_co, report_km_activity_summary_detail_km_co_excel, "Km cơ")


    def report_km_activity_summary_hide_column1(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report3)
        except:
            synthesis_report.report_km_activity_summary(self,"", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(2)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo tổng hợp km xe hoạt động - Chi tiết xung")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_groupvehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoTongHopKmHoatDong_ChiTietXung_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoTongHopKmHoatDong_ChiTietXung_AnHienCot.png")


        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass


    def report_km_activity_summary_downloadexcel2(self, code, eventname, result):           # báo cáo doanh nghiệp - Báo cáo tổng hợp km xe hoạt động - Tổng hợp - checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            var.driver.find_element(By.XPATH, var.report_km_activity_summary_summary).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_km_activity_summary).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_km_activity_summary_summary).click()
        time.sleep(0.5)
        var.driver.find_element(By.XPATH, var.report_km_activity_summary_search).click()
        time.sleep(6)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaotonghopkmhoatdong_tonghop.xlsx"))
        # #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaotonghopkmhoatdong_tonghop.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        report_km_activity_summary_stt = var.driver.find_element(By.XPATH, var.report_km_activity_summary_stt).text
        report_km_activity_summary_liscense_plate = var.driver.find_element(By.XPATH, var.report_km_activity_summary_liscense_plate).text
        report_km_activity_summary_type_vehicle = var.driver.find_element(By.XPATH, var.report_km_activity_summary_type_vehicle).text

        report_km_activity_summary_kmgps = var.driver.find_element(By.XPATH, var.report_km_activity_summary_kmgps).text
        report_km_activity_summary_kmgps = str(''.join(re.findall(r'\d+', report_km_activity_summary_kmgps)))
        report_km_activity_summary_kmgps = report_km_activity_summary_kmgps[0:3]
        report_km_activity_summary_kmgps_ecxel = str(sheet["D6"].value)
        report_km_activity_summary_kmgps_ecxel = str(''.join(re.findall(r'\d+', report_km_activity_summary_kmgps_ecxel)))
        report_km_activity_summary_kmgps_ecxel = report_km_activity_summary_kmgps_ecxel[0:3]

        report_km_activity_summary_kmco = var.driver.find_element(By.XPATH, var.report_km_activity_summary_kmco).text
        report_km_activity_summary_kmco = str(''.join(re.findall(r'\d+', report_km_activity_summary_kmco)))
        report_km_activity_summary_kmco = report_km_activity_summary_kmco[0:3]
        report_km_activity_summary_kmco_ecxel = str(sheet["E6"].value)
        report_km_activity_summary_kmco_ecxel = str(''.join(re.findall(r'\d+', report_km_activity_summary_kmco_ecxel)))
        report_km_activity_summary_kmco_ecxel = report_km_activity_summary_kmco_ecxel[0:3]

        report_km_activity_summary_time_on = var.driver.find_element(By.XPATH, var.report_km_activity_summary_time_on).text
        report_km_activity_summary_time_off = var.driver.find_element(By.XPATH, var.report_km_activity_summary_time_off).text

        logging.info("Báo cáo doanh nghiệp - Báo cáo tổng hợp km xe hoạt động - Tổng hợp")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "5"].value)
            print(sheet[column + "5"])
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "A5", "STT", "A6", report_km_activity_summary_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "B5", "Biển kiểm soát", "B6", report_km_activity_summary_liscense_plate)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "C5", "Loại phương tiện", "C6", report_km_activity_summary_type_vehicle)
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "D5", "Km GPS")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "E5", "Km cơ")

            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "F5", "Thời gian kích xung", "F6", report_km_activity_summary_time_on)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "G5", "Thời gian ngắt xung", "G6", report_km_activity_summary_time_off)
        chucnangkhac.write_result_excelreport2(code, report_km_activity_summary_kmgps, report_km_activity_summary_kmgps_ecxel, "Km gps")
        chucnangkhac.write_result_excelreport2(code, report_km_activity_summary_kmco, report_km_activity_summary_kmco_ecxel, "Km cơ")


    def report_km_activity_summary_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report4)
        except:
            synthesis_report.report_km_activity_summary(self,"", "", "")

        var.driver.find_element(By.XPATH, var.report_km_activity_summary_summary).click()
        time.sleep(1)
        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(2)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo tổng hợp km xe hoạt động - Tổng hợp")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoTongHopKmHoatDong_TongHop_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoTongHopKmHoatDong_TongHop_AnHienCot.png")


        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass











    def report_checkin_checkout(self, code, eventname, result):          #Báo cáo tổng hợp lái xe đăng nhập đăng xuất
        var.driver.implicitly_wait(10)
        try:
            var.driver.find_element(By.XPATH, var.report_checkin_checkout).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.report_checkin_checkout).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo doanh nghiệp - Báo cáo tổng hợp lái xe đăng nhập đăng xuất",
                                              var.check_report_km_activity_summary, "BÁO CÁO TỔNG HỢP LÁI XE ĐĂNG NHẬP ĐĂNG XUẤT", "_BaoCaoDoanhNghiep_BaoCaoTongHopLaiXeDangNhapDangXuat.png")


    def report_checkin_checkout_search(self, code, eventname,result):  #Báo cáo tổng hợp lái xe đăng nhập đăng xuất - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_checkin_checkout).click()
            time.sleep(5)
        write_from_date(var.stop_report_fromdate_input)
        time.sleep(1)
        button = var.driver.find_element(By.XPATH, var.report_search)
        var.driver.execute_script("arguments[0].click();", button)
        time.sleep(5)

        chucnangkhac.write_result_displayed_try(code, eventname, result,
                                                "Báo cáo doanh nghiệp - Báo cáo tổng hợp lái xe đăng nhập đăng xuất",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoDoanhNghiep_BaoCaoTongHopLaiXeDangNhapDangXuat_TimKiem.png")


    def report_checkin_checkout_downloadexcel(self, code, eventname, result):    #Báo cáo tổng hợp lái xe đăng nhập đăng xuất -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_checkin_checkout).click()
            time.sleep(5)
            write_from_date(var.stop_report_fromdate_input)
            time.sleep(1)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(6)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaotonghoplaixedangnhapdangxuat.xlsx"))
        #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaotonghoplaixedangnhapdangxuat.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        report_checkin_checkout_stt = var.driver.find_element(By.XPATH, var.report_checkin_checkout_stt).text
        report_checkin_checkout_laixe = var.driver.find_element(By.XPATH, var.report_checkin_checkout_laixe).text
        report_checkin_checkout_gplx = var.driver.find_element(By.XPATH, var.report_checkin_checkout_gplx).text
        report_checkin_checkout_bienkiemsoat = var.driver.find_element(By.XPATH, var.report_checkin_checkout_bienkiemsoat).text
        report_checkin_checkout_dangnhap = var.driver.find_element(By.XPATH, var.report_checkin_checkout_dangnhap).text
        report_checkin_checkout_dangxuat = var.driver.find_element(By.XPATH, var.report_checkin_checkout_dangxuat).text
        report_checkin_checkout_solandangnhap = var.driver.find_element(By.XPATH, var.report_checkin_checkout_stt).text
        report_checkin_checkout_solandangxuat = var.driver.find_element(By.XPATH, var.report_checkin_checkout_solandangxuat).text
        report_checkin_checkout_tongsolan = var.driver.find_element(By.XPATH, var.report_checkin_checkout_tongsolan).text


        logging.info("Báo cáo doanh nghiệp - Báo cáo tổng hợp lái xe đăng nhập đăng xuất")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "6"].value)
            print(sheet[column + "6"])
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "A6", "STT", "A7", report_checkin_checkout_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "B6", "Lái xe", "B7", report_checkin_checkout_laixe)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "C6", "GPLX", "C7", report_checkin_checkout_gplx)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "D6", "Biển kiểm soát", "D7", report_checkin_checkout_bienkiemsoat)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "E6", "Đăng nhập", "E7", report_checkin_checkout_dangnhap)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "F6", "Đăng xuất", "F7", report_checkin_checkout_dangxuat)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "G6", "Số lần đăng nhập", "G7", report_checkin_checkout_solandangnhap)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "H6", "Số lần đăng xuất", "H7", report_checkin_checkout_solandangxuat)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "I6", "Tổng số lần", "I7", report_checkin_checkout_tongsolan)


    def report_checkin_checkout_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report16)
        except:
            synthesis_report.report_checkin_checkout(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(2)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo doanh nghiệp - Báo cáo tổng hợp lái xe đăng nhập đăng xuất")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoTongHopLaiXeDangNhapDangXuat_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoTongHopLaiXeDangNhapDangXuat_AnHienCot.png")


        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass






class activity_report:      #Báo cáo hoạt động


    def stop_report(self, code, eventname, result):     #Báo cáo dừng đỗ
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.stop_report).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.stop_report).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo dừng đỗ",
                                              var.check_report_km_activity_summary, "BÁO CÁO DỪNG ĐỖ", "_BaoCaoDoanhNghiep_BaoCaoDungDo.png")


    def stop_report_search(self, code, eventname,result):  #Báo cáo dừng đỗ - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.stop_report_fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.stop_report).click()
            time.sleep(5)

        write_from_date(var.stop_report_fromdate_input)
        var.driver.find_element(By.XPATH, var.activity_synthesis_report_search).click()
        time.sleep(5)

        chucnangkhac.write_result_displayed_try(code, eventname, result,
                                                "Báo cáo doanh nghiệp - Báo cáo dừng đỗ",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoDoanhNghiep_BaoCaoDungDo_TimKiem.png")


    def stop_report_downloadexcel1(self, code, eventname, result):    #Báo cáo dừng đỗ -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.stop_report).click()
            time.sleep(5)
            write_from_date(var.stop_report_fromdate_input)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(6)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        try:
            filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
            shutil.move(filename, os.path.join(var.excelpath, r"baocaodungdo.xlsx"))
        except:
            var.driver.find_element(By.XPATH, var.downloadexcel).click()
            time.sleep(15)
            filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
            shutil.move(filename, os.path.join(var.excelpath, r"baocaodungdo.xlsx"))

        # #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

        try:
            wordbook = openpyxl.load_workbook(var.excelpath+"/baocaodungdo.xlsX")
            sheet = wordbook.get_sheet_by_name("Data")
        except:
            var.driver.find_element(By.XPATH, var.downloadexcel).click()
            time.sleep(15)
            filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
            shutil.move(filename, os.path.join(var.excelpath, r"baocaodungdo.xlsx"))
            wordbook = openpyxl.load_workbook(var.excelpath+"/baocaodungdo.xlsX")
            sheet = wordbook.get_sheet_by_name("Data")


        stop_report_stt = var.driver.find_element(By.XPATH, var.stop_report_stt).text
        stop_report_liscense_plate = var.driver.find_element(By.XPATH, var.stop_report_liscense_plate).text
        stop_report_group = var.driver.find_element(By.XPATH, var.stop_report_group).text
        stop_report_group = " " + stop_report_group

        stop_report_name_driver = var.driver.find_element(By.XPATH, var.stop_report_name_driver).text
        stop_report_staff_code = var.driver.find_element(By.XPATH, var.stop_report_staff_code).text
        stop_report_phone_number = var.driver.find_element(By.XPATH, var.stop_report_phone_number).text
        stop_report_time = var.driver.find_element(By.XPATH, var.stop_report_time).text
        stop_report_time_minute = var.driver.find_element(By.XPATH, var.stop_report_time_minute).text
        stop_report_time_stop = var.driver.find_element(By.XPATH, var.stop_report_time_stop).text
        stop_report_turn_on_stop = var.driver.find_element(By.XPATH, var.stop_report_turn_on_stop).text     #Nổ máy khi dừng
        stop_report_turn_on_air_condition_stop = var.driver.find_element(By.XPATH, var.stop_report_turn_on_air_condition_stop).text

        stop_report_fuel = var.driver.find_element(By.XPATH, var.stop_report_fuel).text
        stop_report_fuel = str(''.join(re.findall(r'\d+', stop_report_fuel)))
        stop_report_fuel = stop_report_fuel[0:1]
        stop_report_fuel_excel = str(sheet["L6"].value)
        stop_report_fuel_excel = str(''.join(re.findall(r'\d+', stop_report_fuel_excel)))
        stop_report_fuel_excel = stop_report_fuel_excel[0:1]




        stop_report_temperature = var.driver.find_element(By.XPATH, var.stop_report_temperature).text       #Nhiệt độ
        stop_report_address = var.driver.find_element(By.XPATH, var.stop_report_address).text
        stop_report_titude_latitude_longitude = var.driver.find_element(By.XPATH, var.stop_report_titude_latitude_longitude).text
        stop_report_note = var.driver.find_element(By.XPATH, var.stop_report_note).text
        stop_report_route = var.driver.find_element(By.XPATH, var.stop_report_route).text
        stop_report_video = var.driver.find_element(By.XPATH, var.stop_report_video).get_attribute("src")
        stop_report_video = stop_report_video[-40::]


        logging.info("Báo cáo doanh nghiệp - Báo cáo dừng đỗ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "5"].value)
            print(sheet[column + "5"])

            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "A5", "STT", "A6", stop_report_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "B5", "Biển kiểm soát", "B6", stop_report_liscense_plate)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "C5", "Nhóm", "C6", stop_report_group)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "D5", "Tên lái xe", "D6", stop_report_name_driver)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "E5", "Mã nhân viên", "E6", stop_report_staff_code)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "F5", "Số điện thoại", "F6", stop_report_phone_number)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "G5", "Thời gian", "G6", stop_report_time)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "H5", "Thời gian (phút)", "H6", stop_report_time_minute)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "I5", "TG dừng đỗ(HH:mm:ss)", "I6", stop_report_time_stop)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "J5", "Bật máy khi dừng (phút)", "J6", stop_report_turn_on_stop)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "K5", "Bật điều hòa khi dừng (phút)", "K6",stop_report_turn_on_air_condition_stop)
            # chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "L5", "Nhiên liệu tiêu hao (lít)", "L6", stop_report_fuel)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "M5", "Nhiệt độ", "M6", stop_report_temperature)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "N5", "Địa điểm", "N6", stop_report_address)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "O5", "Kinh độ, vĩ độ", "O6", stop_report_titude_latitude_longitude)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "P5", "Ghi chú", "P6", stop_report_note)
        chucnangkhac.write_result_excel_checkweb(code, stop_report_route, "Lộ trình")
        chucnangkhac.write_result_excel_checkweb(code, stop_report_video, "/icons/VideoCam/icons8-video-call-48.png")
        chucnangkhac.write_result_excelreport2(code, stop_report_fuel, stop_report_fuel_excel, "Nhiên liệu tiêu hao (lít)")



    def stop_report_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report5)
        except:
            activity_report.stop_report(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(2)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo dừng đỗ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_groupvehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoDungDo_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoDungDo_AnHienCot.png")


        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass












    def report_business_trip(self, code, eventname, result):    #Báo cáo chuyến kinh doanh
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.report_business_trip).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.report_business_trip).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo chuyến kinh doanh",
                                              var.check_report_km_activity_summary, "BÁO CÁO CHUYẾN KINH DOANH", "_BaoCaoDoanhNghiep_BaoCaoChuyenKinhDoanh.png")


    def report_business_trip_search(self, code, eventname, result):    #Báo cáo chuyến kinh doanh - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_business_trip).click()
            time.sleep(5)

        write_from_date(var.fromdate_input)
        button = var.driver.find_element(By.XPATH, var.report_search)
        var.driver.execute_script("arguments[0].click();", button)
        time.sleep(5)
        chucnangkhac.write_result_displayed_try(code, eventname, result,
                                                "Báo cáo doanh nghiệp - Báo cáo chuyến kinh doanh",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoDoanhNghiep_BaoCaoChuyenKinhDoanh_TimKiem.png")


    def report_business_trip_downloadexcel(self, code, eventname, result):    #Báo cáo chuyến kinh doanh -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            # var.driver.find_element(By.XPATH, var.report_search).click()
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_business_trip).click()
            time.sleep(5)
            write_from_date(var.fromdate_input)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(6)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaochuyenkinhdoanh.xlsx"))
        #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaochuyenkinhdoanh.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        report_business_trip_stt = var.driver.find_element(By.XPATH, var.report_business_trip_stt).text
        report_business_trip_type_vehicle = var.driver.find_element(By.XPATH, var.report_business_trip_type_vehicle).text
        report_business_trip_liscense_plate = var.driver.find_element(By.XPATH, var.report_business_trip_liscense_plate).text
        report_business_trip_group = var.driver.find_element(By.XPATH, var.report_business_trip_group).text
        report_business_trip_group = " " + report_business_trip_group
        report_business_trip_from_addres = var.driver.find_element(By.XPATH, var.report_business_trip_from_addres).text
        report_business_trip_to_addres = var.driver.find_element(By.XPATH, var.report_business_trip_to_addres).text
        report_business_trip_from_hour = var.driver.find_element(By.XPATH, var.report_business_trip_from_hour).text
        report_business_trip_to_hour = var.driver.find_element(By.XPATH, var.report_business_trip_to_hour).text
        report_business_trip_time_active = var.driver.find_element(By.XPATH, var.report_business_trip_time_active).text

        report_business_trip_km_gps = var.driver.find_element(By.XPATH, var.report_business_trip_km_gps).text
        report_business_trip_km_gps = str(''.join(re.findall(r'\d+', report_business_trip_km_gps)))
        report_business_trip_km_gps = report_business_trip_km_gps[0:3]
        report_business_trip_km_gps_excel = str(sheet["J7"].value)
        report_business_trip_km_gps_excel = str(''.join(re.findall(r'\d+', report_business_trip_km_gps_excel)))
        report_business_trip_km_gps_excel = report_business_trip_km_gps_excel[0:3]

        report_business_trip_v_hanhtrinh = var.driver.find_element(By.XPATH, var.report_business_trip_v_hanhtrinh).text
        report_business_trip_v_hanhtrinh = str(''.join(re.findall(r'\d+', report_business_trip_v_hanhtrinh)))
        report_business_trip_v_hanhtrinh = report_business_trip_v_hanhtrinh[0:3]
        report_business_trip_v_hanhtrinh_excel = str(sheet["K7"].value)
        report_business_trip_v_hanhtrinh_excel = str(''.join(re.findall(r'\d+', report_business_trip_v_hanhtrinh_excel)))
        report_business_trip_v_hanhtrinh_excel = report_business_trip_v_hanhtrinh_excel[0:3]

        report_business_trip_v_average = var.driver.find_element(By.XPATH, var.report_business_trip_v_average).text
        report_business_trip_v_average = str(''.join(re.findall(r'\d+', report_business_trip_v_average)))
        report_business_trip_v_average = report_business_trip_v_average[0:3]
        report_business_trip_v_average_excel = str(sheet["L7"].value)
        report_business_trip_v_average_excel = str(''.join(re.findall(r'\d+', report_business_trip_v_average_excel)))
        report_business_trip_v_average_excel = report_business_trip_v_average_excel[0:3]

        report_business_trip_km_co = var.driver.find_element(By.XPATH, var.report_business_trip_km_co).text
        report_business_trip_fuel_start = var.driver.find_element(By.XPATH, var.report_business_trip_fuel_start).text
        report_business_trip_fuel_start = str(''.join(re.findall(r'\d+', report_business_trip_fuel_start)))
        report_business_trip_fuel_start = report_business_trip_fuel_start[0:3]
        rreport_business_trip_fuel_start_excel = str(sheet["N7"].value)
        rreport_business_trip_fuel_start_excel = str(''.join(re.findall(r'\d+', rreport_business_trip_fuel_start_excel)))
        rreport_business_trip_fuel_start_excel = rreport_business_trip_fuel_start_excel[0:3]

        report_business_trip_fuel_end = var.driver.find_element(By.XPATH, var.report_business_trip_fuel_end).text
        report_business_trip_fuel_end = str(''.join(re.findall(r'\d+', report_business_trip_fuel_end)))
        report_business_trip_fuel_end = report_business_trip_fuel_end[0:3]
        report_business_trip_fuel_end_excel = str(sheet["O7"].value)
        report_business_trip_fuel_end_excel = str(''.join(re.findall(r'\d+', report_business_trip_fuel_end_excel)))
        report_business_trip_fuel_end_excel = report_business_trip_fuel_end_excel[0:3]

        report_business_trip_fuel1 = var.driver.find_element(By.XPATH, var.report_business_trip_fuel1).text     #nhiên liệu tiêu thụ
        report_business_trip_fuel1 = str(''.join(re.findall(r'\d+', report_business_trip_fuel1)))
        report_business_trip_fuel1 = report_business_trip_fuel1[0:3]
        report_business_trip_fuel1_excel = str(sheet["P7"].value)
        report_business_trip_fuel1_excel = str(''.join(re.findall(r'\d+', report_business_trip_fuel1_excel)))
        report_business_trip_fuel1_excel = report_business_trip_fuel1_excel[0:3]

        report_business_trip_fuel2 = var.driver.find_element(By.XPATH, var.report_business_trip_fuel2).text     #nhiên liệu 1km
        report_business_trip_fuel3 = var.driver.find_element(By.XPATH, var.report_business_trip_fuel3).text     #Nhiên liệu tiêu thụ
        report_business_trip_time_in_place_from = var.driver.find_element(By.XPATH, var.report_business_trip_time_in_place_from).text
        report_business_trip_time_in_place_to = var.driver.find_element(By.XPATH, var.report_business_trip_time_in_place_to).text
        report_business_trip_time_stop = var.driver.find_element(By.XPATH, var.report_business_trip_time_stop).text
        report_business_trip_turn_air_conditional = var.driver.find_element(By.XPATH, var.report_business_trip_turn_air_conditional).text
        report_business_trip_route = var.driver.find_element(By.XPATH, var.report_business_trip_route).text

        logging.info("Báo cáo doanh nghiệp - Báo cáo chuyến kinh doanh")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "6"].value)
            print(sheet[column + "6"])
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "A6", "STT", "A7", report_business_trip_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "B6", "Loại phương tiện", "B7", report_business_trip_type_vehicle)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "C6", "Biển số", "C7", report_business_trip_liscense_plate)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "D6", "Nhóm", "D7", report_business_trip_group)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "E6", "Địa chỉ xuất phát", "E7", report_business_trip_from_addres)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "F6", "Địa chỉ đến", "F7", report_business_trip_to_addres)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "G6", "Giờ đi", "G7", report_business_trip_from_hour)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "H6", "Giờ đến", "H7", report_business_trip_to_hour)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "I6", "TG hoạt động (hh:mm)", "I7", report_business_trip_time_active)

            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "J6", "Km GPS")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "K6", "Vận tốc hành trình")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "L6", "Vận tốc trung bình")
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "M6", "Km Cơ", "M7", report_business_trip_km_co)
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "N6", "NL bắt đầu chuyến (lít)")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "O6", "NL kết thúc chuyến (lít)")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "P6", "NL tiêu thụ (lít)")

            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "Q6", "Định mức nhiên liệu trên 1km", "Q7", report_business_trip_fuel2)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "R6", "Nhiên liệu tiêu thụ định mức", "R7", report_business_trip_fuel3)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "S6", "TG trong điểm đi (hh:mm)", "S7", report_business_trip_time_in_place_from)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "T6", "TG trong điểm đến (hh:mm)", "T7", report_business_trip_time_in_place_to)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "U6", "TG dừng đỗ (hh:mm)", "U7", report_business_trip_time_stop)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "V6", "TG bật điều hòa (hh:mm)", "V7", report_business_trip_turn_air_conditional)

        chucnangkhac.write_result_excel_checkweb(code, report_business_trip_route, "Lộ trình")
        chucnangkhac.write_result_excelreport2(code, report_business_trip_km_gps, report_business_trip_km_gps_excel, "Km GPS")
        chucnangkhac.write_result_excelreport2(code, report_business_trip_v_hanhtrinh, report_business_trip_v_hanhtrinh_excel, "Vận tốc hành trình")
        chucnangkhac.write_result_excelreport2(code, report_business_trip_v_average, report_business_trip_v_average_excel, "Vận tốc trung bình")
        chucnangkhac.write_result_excelreport2(code, report_business_trip_fuel_start, rreport_business_trip_fuel_start_excel, "NL bắt đầu chuyến (lít)")
        chucnangkhac.write_result_excelreport2(code, report_business_trip_fuel_end, report_business_trip_fuel_end_excel, "NL kết thúc chuyến (lít)")
        chucnangkhac.write_result_excelreport2(code, report_business_trip_fuel1, report_business_trip_fuel1_excel, "NL tiêu thụ (lít)")


    def report_business_trip_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report6)
        except:
            activity_report.report_business_trip(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(2)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo chuyến kinh doanh")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoChuyenKinhDoanh_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoChuyenKinhDoanh_AnHienCot.png")
        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass








    def station_report(self, code, eventname, result):      #Báo cáo ra  vào trạm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.station_report).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.station_report).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo ra vào trạm",
                                              var.check_report_km_activity_summary, "BÁO CÁO RA VÀO TRẠM", "_BaoCaoDoanhNghiep_BaoCaoRaVaoTram.png")


    def station_report_search(self, code, eventname, result):    #Báo cáo ra  vào trạm - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.station_report).click()
            time.sleep(5)

        write_from_date(var.fromdate_input)
        button = var.driver.find_element(By.XPATH, var.report_search)
        var.driver.execute_script("arguments[0].click();", button)
        time.sleep(5)
        chucnangkhac.write_result_displayed_try(code, eventname, result,
                                                "Báo cáo doanh nghiệp - Báo cáo ra vào trạm",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoDoanhNghiep_BaoCaoRaVaoTram_TimKiem.png")


    def station_report_downloadexcel(self, code, eventname, result):    #Báo cáo ra vào trạm -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            # var.driver.find_element(By.XPATH, var.report_search).click()
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.station_report).click()
            time.sleep(5)
            write_from_date(var.fromdate_input)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(6)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaoravaotram.xlsx"))
        #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaoravaotram.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        station_report_stt = var.driver.find_element(By.XPATH, var.station_report_stt).text
        station_report_type_vehicle = var.driver.find_element(By.XPATH, var.station_report_type_vehicle).text
        station_report_liscense_plate = var.driver.find_element(By.XPATH, var.station_report_liscense_plate).text
        station_report_group = var.driver.find_element(By.XPATH, var.station_report_group).text
        station_report_group = " " + station_report_group
        station_report_in_station = var.driver.find_element(By.XPATH, var.station_report_in_station).text
        station_report_out_station = var.driver.find_element(By.XPATH, var.station_report_out_station).text
        station_report_name_station = var.driver.find_element(By.XPATH, var.station_report_name_station).text
        station_report_time_station = var.driver.find_element(By.XPATH, var.station_report_time_station).text
        station_report_route = var.driver.find_element(By.XPATH, var.station_report_route).text
        station_report_linkvideo = var.driver.find_element(By.XPATH, var.station_report_linkvideo).get_attribute('src')
        station_report_linkvideo = str(station_report_linkvideo[-38::])

        logging.info("Báo cáo doanh nghiệp - Báo cáo ra vào trạm")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "7", "A7", "STT", "A8", station_report_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "7", "B7", "Loại phương tiện", "B8", station_report_type_vehicle)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "7", "C7", "Biển số", "C8", station_report_liscense_plate)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "7", "D7", "Nhóm", "D8", station_report_group)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "7", "E7", "Thời điểm vào trạm", "E8", station_report_in_station)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "7", "F7", "Thời điểm ra trạm", "F8", station_report_out_station)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "7", "G7", "Tên trạm", "G8", station_report_name_station)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "7", "H7", "Thời gian trong trạm (Phút)", "H8", station_report_time_station)
        chucnangkhac.write_result_excel_checkweb(code, station_report_route, "Lộ trình")
        chucnangkhac.write_result_excel_checkweb(code, station_report_linkvideo, "/Images/TrackingOnline/camera-blue.png")


    def station_report_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report7)
        except:
            activity_report.station_report(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(2)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo ra vào trạm")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoRaVaoTram_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoRaVaoTram_AnHienCot .png")
        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass










    def report_air_conditioner_summaries(self, code, eventname, result):        #Báo cáo tổng hợp bật điều hòa
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo tổng hợp bật điều hòa",
                                              var.check_report_km_activity_summary, "BÁO CÁO TỔNG HỢP BẬT ĐIỀU HÒA", "_BaoCaoDoanhNghiep_BaoCaoTongHopDieuHoa.png")


    def report_air_conditioner_summaries_search(self, code, eventname, result):    #Báo cáo tổng hợp bật điều hòa - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries).click()
            time.sleep(5)

        write_from_date(var.fromdate_input)
        button = var.driver.find_element(By.XPATH, var.report_search)
        var.driver.execute_script("arguments[0].click();", button)
        time.sleep(5)
        chucnangkhac.write_result_displayed_try(code, eventname, result,
                                                "Báo cáo doanh nghiệp - Báo cáo tổng hợp điều hòa",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoDoanhNghiep_BaoCaoTongHopDieuHoa_TimKiem.png")


    def report_air_conditioner_summaries_downloadexcel(self, code, eventname, result):    #Báo cáo tổng hợp bật điều hòa -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries).click()
            time.sleep(5)
            write_from_date(var.fromdate_input)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(7)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaotonghopbattatdieuhoa.xlsx"))
        #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaotonghopbattatdieuhoa.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        report_air_conditioner_summaries_stt = var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries_stt).text
        report_air_conditioner_summaries_liscense_plate = var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries_liscense_plate).text
        report_air_conditioner_summaries_from_day = var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries_from_day).text
        report_air_conditioner_summaries_from_day = "00:00 " + report_air_conditioner_summaries_from_day

        report_air_conditioner_summaries_to_day = var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries_to_day).text
        report_air_conditioner_summaries_to_day = "23:59 " + report_air_conditioner_summaries_to_day

        report_air_conditioner_summaries_summary_on_time = var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries_summary_on_time).text
        report_air_conditioner_summaries_number_stop_on_air_conditonal = var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries_number_stop_on_air_conditonal).text
        report_air_conditioner_summaries_number_stop_lit_air_conditonal = var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries_number_stop_lit_air_conditonal).text
        report_air_conditioner_summaries_number_stop_lit_air_conditonal = str(''.join(re.findall(r'\d+', report_air_conditioner_summaries_number_stop_lit_air_conditonal)))
        report_air_conditioner_summaries_number_stop_lit_air_conditonal = report_air_conditioner_summaries_number_stop_lit_air_conditonal[0:2]

        report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel = sheet["G7"].value
        print("e0: ".format(report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel))

        report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel = round(report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel, 1)
        print("e1: ".format(report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel))

        report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel = str(report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel)
        report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel = str(''.join(re.findall(r'\d+', report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel)))
        print("e2: ".format(report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel))
        report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel = report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel[0:2]



        report_air_conditioner_summaries_detail = var.driver.find_element(By.XPATH, var.report_air_conditioner_summaries_detail).text


        logging.info("Báo cáo doanh nghiệp - Báo cáo tổng hợp bật điều hòa")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "6"].value)
            print(sheet[column + "6"])
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "A6", "STT", "A7", report_air_conditioner_summaries_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "B6", "Biển kiểm soát", "B7", report_air_conditioner_summaries_liscense_plate)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "C6", "Từ ngày", "C7", report_air_conditioner_summaries_from_day)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "D6", "Đến ngày", "D7", report_air_conditioner_summaries_to_day)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "E6", "Tổng thời gian bật (phút)", "E7", report_air_conditioner_summaries_summary_on_time)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "F6", "Số phút dừng đỗ bật điều hòa", "F7", report_air_conditioner_summaries_number_stop_on_air_conditonal)
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "G6", "Số lít dừng đỗ bật điều hòa")
        chucnangkhac.write_result_excel_checkweb(code, report_air_conditioner_summaries_detail, "Chi tiết")
        chucnangkhac.write_result_excelreport2(code, report_air_conditioner_summaries_number_stop_lit_air_conditonal, report_air_conditioner_summaries_number_stop_lit_air_conditonal_excel, "Số lít dừng đỗ bật điều hòa")


    def report_air_conditioner_summaries_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report8)
        except:
            activity_report.report_air_conditioner_summaries(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(2)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo tổng hợp bật điều hòa")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoRaVaoTram_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoRaVaoTram_AnHienCot .png")
        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass









    def machine_report(self, code, eventname, result):        #Báo cáo động cơ
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.machine_report).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk3'], var.data['login']['conhom_quantri_mk3'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.machine_report).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo động cơ",
                                              var.check_report_km_activity_summary, "BÁO CÁO ĐỘNG CƠ", "_BaoCaoDoanhNghiep_BaoCaoDongCo.png")


    def machine_report_search(self, code, eventname, result):    #Báo cáo động cơ - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk3'], var.data['login']['conhom_quantri_mk3'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.machine_report).click()
            time.sleep(5)

        write_from_date(var.fromdate_input)
        button = var.driver.find_element(By.XPATH, var.report_search)
        var.driver.execute_script("arguments[0].click();", button)
        time.sleep(5)
        try:
            print("a1")
            var.driver.find_element(By.XPATH, "//*[@id='ctl00_ctl00_MainContent_Content_ScrollPanel']/div/table/tbody/tr[2]/td[2]")
            print("a2")
        except:
            write_from_date_month(var.fromdate_input)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(5)

        chucnangkhac.write_result_displayed_try(code, eventname, result,
                                                "Báo cáo doanh nghiệp - Báo cáo động cơ",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoDoanhNghiep_BaoCaoDongCo_TimKiem.png")


    def machine_report_downloadexcel(self, code, eventname, result):    #Báo cáo động cơ -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            # var.driver.find_element(By.XPATH, var.report_search).click()
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk3'], var.data['login']['conhom_quantri_mk3'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.machine_report).click()
            time.sleep(5)
            write_from_date(var.fromdate_input)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(6)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaodongco.xlsx"))
        #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaodongco.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")


        machine_report_stt = var.driver.find_element(By.XPATH, var.machine_report_stt).text
        machine_report_liscense_plate = var.driver.find_element(By.XPATH, var.machine_report_liscense_plate).text
        machine_report_group = var.driver.find_element(By.XPATH, var.machine_report_group).text
        machine_report_group = " " + machine_report_group
        machine_report_start_time = var.driver.find_element(By.XPATH, var.machine_report_start_time).text
        machine_report_time = var.driver.find_element(By.XPATH, var.machine_report_time).text
        # machine_report_time = str(machine_report_time)
        #
        # machine_report_time_excel = str(sheet["E7"].value)
        # machine_report_time_excel = str(''.join(re.findall(r'\d+', machine_report_time_excel)))

        machine_report_to_time = var.driver.find_element(By.XPATH, var.machine_report_to_time).text
        machine_report_start_address = var.driver.find_element(By.XPATH, var.machine_report_start_address).text
        machine_report_end_address = var.driver.find_element(By.XPATH, var.machine_report_end_address).text
        machine_report_route = var.driver.find_element(By.XPATH, var.machine_report_route).text

        logging.info("Báo cáo doanh nghiệp - Báo cáo động cơ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "6"].value)
            print(sheet[column + "6"])
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "A6", "STT", "A7", machine_report_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "B6", "Biển kiểm soát", "B7", machine_report_liscense_plate)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "C6", "Nhóm", "C7", machine_report_group)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "D6", "Giờ bắt đầu", "D7", machine_report_start_time)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "E6", "Thời gian (hh:mm)", "E7", machine_report_time)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "F6", "Giờ đến", "F7", machine_report_to_time)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "G6", "Địa điểm bắt đầu", "G7", machine_report_start_address)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "H6", "Địa điểm kết thúc", "H7", machine_report_end_address)
        chucnangkhac.write_result_excel_checkweb(code, machine_report_route, "Lộ trình")
        # chucnangkhac.write_result_excelreport2(code, machine_report_time, machine_report_time_excel, "Thời gian (hh:mm")


    def machine_report_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report9)
        except:
            activity_report.machine_report(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(2)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo động cơ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoDongCo_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoDongCo_AnHienCot .png")
        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass







    def report_speed_over(self, code, eventname, result):        #Báo cáo quá tốc độ
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.report_speed_over).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.report_speed_over).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo quá tốc độ",
                                              var.check_report_km_activity_summary, "BÁO CÁO QUÁ TỐC ĐỘ", "_BaoCaoDoanhNghiep_BaoCaoQuaTocDo.png")


    def report_speed_over_search(self, code, eventname, result):    #Báo cáo quá tốc độ - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_speed_over).click()
            time.sleep(5)
        # write_from_date_month(var.fromdate_input)   #hàm nhập trừ 1 tháng so với hiện tại

        JS_ADD_TEXT_TO_INPUT = """
          elm = arguments[0], txt = arguments[1];
          elm.value += txt;
          elm.dispatchEvent(new Event('change'));
          """

        var.driver.find_element(By.XPATH, var.fromdate_input).clear()
        elm = var.driver.find_element(By.XPATH, var.fromdate_input)
        var.driver.execute_script(JS_ADD_TEXT_TO_INPUT, elm, giamsat.data['baocao']['quatocdo_tungay'])
        time.sleep(1)

        var.driver.find_element(By.XPATH, var.todate_input).clear()
        elm = var.driver.find_element(By.XPATH, var.todate_input)
        var.driver.execute_script(JS_ADD_TEXT_TO_INPUT, elm, giamsat.data['baocao']['quatocdo_denngay'])
        time.sleep(1)

        button = var.driver.find_element(By.XPATH, var.report_search)
        var.driver.execute_script("arguments[0].click();", button)
        time.sleep(10)
        print("aaaa")
        chucnangkhac.write_result_displayed_try(code, eventname, result,"Báo cáo quá tốc độ",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoDoanhNghiep_BaoCaoQuaTocDo_TimKiem.png")


    def report_speed_over_downloadexcel(self, code, eventname, result):    #Báo cáo quá tốc độ -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()

        try:
            # var.driver.find_element(By.XPATH, var.report_search).click()
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_speed_over).click()
            time.sleep(5)

            write_from_date_month(var.fromdate_input)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(14)
            print("n1")

        try:
            var.driver.find_element(By.XPATH, var.nodata).is_displayed()
            print("n2")
        except:
            print("n3")
            var.driver.find_element(By.XPATH, var.downloadexcel).click()
            time.sleep(15)
            filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
            shutil.move(filename, os.path.join(var.excelpath, r"baocaoquatocdo.xlsx"))

            #Đọc check file excel
            bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
            wordbook = openpyxl.load_workbook(var.excelpath+"/baocaoquatocdo.xlsX")
            sheet = wordbook.get_sheet_by_name("Data")

            report_speed_over_stt = var.driver.find_element(By.XPATH, var.report_speed_over_stt).text
            report_speed_over_bienkiemsoat = var.driver.find_element(By.XPATH, var.report_speed_over_bienkiemsoat).text
            report_speed_over_thoidiem = var.driver.find_element(By.XPATH, var.report_speed_over_thoidiem).text
            report_speed_over_thoigian = var.driver.find_element(By.XPATH, var.report_speed_over_thoigian).text
            report_speed_over_quangduong = var.driver.find_element(By.XPATH, var.report_speed_over_quangduong).text
            report_speed_over_tocdocucdai = var.driver.find_element(By.XPATH, var.report_speed_over_tocdocucdai).text
            report_speed_over_diadiembatdau = var.driver.find_element(By.XPATH, var.report_speed_over_diadiembatdau).text
            report_speed_over_diadiemketthuc = var.driver.find_element(By.XPATH, var.report_speed_over_diadiemketthuc).text
            report_speed_over_ghichu = var.driver.find_element(By.XPATH, var.report_speed_over_ghichu).text
            report_speed_over_bando = var.driver.find_element(By.XPATH, var.report_speed_over_bando).text


            logging.info("Báo cáo doanh nghiệp - Báo cáo quá tốc độ")
            logging.info("Mã - " + code)
            logging.info("Tên sự kiện - " + eventname)
            logging.info("Kết quả - " + result)
            for column in bangchucai:
                print(sheet[column + "6"].value)
                print(sheet[column + "6"])
                chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "A6", "STT", "A7", report_speed_over_stt)
                chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "B6", "Biển kiểm soát", "B7", report_speed_over_bienkiemsoat)
                chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "C6", "Thời điểm", "C7", report_speed_over_thoidiem)
                chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "D6", "Thời gian (s)", "D7", report_speed_over_thoigian)
                chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "E6", "Quãng đường (m)", "E7", report_speed_over_quangduong)
                chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "F6", "Tốc độ cực đại ≥", "F7", report_speed_over_tocdocucdai)
                chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "G6", "Địa điểm bắt đầu", "G7", report_speed_over_diadiembatdau)
                chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "H6", "Địa điểm kết thúc", "H7", report_speed_over_diadiemketthuc)
                chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "I6", "Ghi chú", "I7", report_speed_over_ghichu)
            chucnangkhac.write_result_excel_checkweb(code, report_speed_over_bando, "Lộ trình")


    def report_speed_over_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report15)
        except:
            activity_report.report_speed_over(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(2)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo quá tốc độ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoQuaTocDo_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoQuaTocDo_AnHienCot.png")
        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass








class report_schedule:  #Báo cáo lịch trình
    def position_history(self, code, eventname, result):     #Báo cáo hành trình
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.position_history).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.position_history).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo hành trình",
                                              var.check_report_km_activity_summary, "BÁO CÁO HÀNH TRÌNH", "_BaoCaoDoanhNghiep_BaoCaoHanhTrinh.png")


    def position_history_time_slot(self, code, eventname, result):     #Báo cáo hành trình
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.position_history_time_slot).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.position_history).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.position_history_time_slot).click()
        time.sleep(1.5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo hành trình",
                                              var.check_position_history_time_slot, "Khung giờ xuất báo cáo", "_BaoCaoDoanhNghiep_BaoCaoHanhTrinh_KhungGio.png")

        if var.driver.find_element(By.XPATH, var.position_history_time_slot).is_selected() == True:
            var.driver.find_element(By.XPATH, var.position_history_time_slot).click()


    def position_history_search(self, code, eventname, result):     #Báo cáo hành trình
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.position_history_choose_car).click()
            time.sleep(1)
            var.driver.find_element(By.XPATH, var.position_history_choose_car3rd).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.position_history).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.position_history_choose_car).click()
            time.sleep(1)
            var.driver.find_element(By.XPATH, var.position_history_choose_car3rd).click()
        time.sleep(1)
        button = var.driver.find_element(By.XPATH, var.report_search)
        var.driver.execute_script("arguments[0].click();", button)
        time.sleep(4)
        chucnangkhac.write_result_displayed_try(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo hành trình",
                                                var.check_position_history_search, "_BaoCaoDoanhNghiep_BaoCaoHanhTrinh_TimKiem.png")


    def position_history_downloadexcel1(self, code, eventname, result):    #Báo cáo hành trình -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            # var.driver.find_element(By.XPATH, var.check_position_history_search)
            var.driver.find_element(By.XPATH, var.check_report_search1)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.position_history).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.position_history_choose_car).click()
            time.sleep(1)
            var.driver.find_element(By.XPATH, var.position_history_choose_car3rd).click()
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(8)
        time.sleep(5)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaohanhtrinh.xls"))


        logging.info("Báo cáo doanh nghiệp - Báo cáo hành trình")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        logging.info("True")
        chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")



        # x2x = XLS2XLSX(var.excelpath + "/baocaohanhtrinh.xls")
        # x2x.to_xlsx(var.excelpath + "/baocaohanhtrinh.xlsx")
        #
        # # #Đọc check file excel

        # wordbook = openpyxl.load_workbook(var.excelpath+"/baocaohanhtrinh.xlsX")
        # sheet = wordbook.get_sheet_by_name("baocaohanhtrinh")
        # position_history_vehicle = var.driver.find_element(By.XPATH, var.position_history_vehicle).text
        # print("web: " + position_history_vehicle)
        # position_history_vehicle_excel = str(sheet["A4"].value)
        # print("excel: " + position_history_vehicle_excel)

        # logging.info("Báo cáo doanh nghiệp - Báo cáo hành trình")
        # logging.info("Mã - " + code)
        # logging.info("Tên sự kiện - " + eventname)
        # logging.info("Kết quả - " + result)
        # chucnangkhac.write_result_excelreport2(code, position_history_vehicle, position_history_vehicle_excel, "Phương tiện")


    def position_history_hide_map(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report10)
        except:
            report_schedule.position_history(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_map).click()
        time.sleep(3)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo hành trình")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.position_history_hide_map).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoDongCo_AnHienBanDo.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoDongCo_AnHienBanDo .png")
        time.sleep(1.5)






class fuel_report:
    def fuel_consumption_summary_report(self, code, eventname, result):     #Báo cáo tổng hợp tiêu hao nhiên liệu
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo tổng hợp tiêu hao nhiên liệu",
                                              var.check_report_km_activity_summary, "BÁO CÁO TỔNG HỢP TIÊU HAO NHIÊN LIỆU", "_BaoCaoDoanhNghiep_BaoCaoTongHopTieuHaoNhienLieu.png")


    def fuel_consumption_summary_report_search(self, code, eventname,result):  #Báo cáo tổng hợp tiêu hao nhiên liệu - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report).click()
            time.sleep(5)
        # write_from_date(var.stop_report_fromdate_input)
        # time.sleep(1)
        var.driver.find_element(By.XPATH, var.activity_synthesis_report_search).click()
        time.sleep(12)

        chucnangkhac.write_result_displayed_try(code, eventname, result,
                                                "Báo cáo doanh nghiệp - Báo cáo tổng hợp tiêu hao nhiên liệu",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoDoanhNghiep_BaoCaoTongHopTieuHaoNhienLieu_TimKiem.png")


    def fuel_consumption_summary_report_downloadexcel(self, code, eventname, result):    #Báo cáo tổng hợp tiêu hao nhiên liệu -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            # var.driver.find_element(By.XPATH, var.report_search).click()
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report).click()
            time.sleep(5)
            write_from_date(var.stop_report_fromdate_input)
            time.sleep(1)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            # var.driver.find_element(By.XPATH, var.report_search).click()
            time.sleep(12)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        try:
            filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
            shutil.move(filename, os.path.join(var.excelpath, r"baocaotonghoptieuhaonhienlieu.xlsx"))
        except:
            var.driver.find_element(By.XPATH, var.downloadexcel).click()
            time.sleep(15)
            filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
            shutil.move(filename, os.path.join(var.excelpath, r"baocaotonghoptieuhaonhienlieu.xlsx"))

        #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaotonghoptieuhaonhienlieu.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        fuel_consumption_summary_report_stt = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_stt).text
        fuel_consumption_summary_report_group = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_group).text
        fuel_consumption_summary_report_group = fuel_consumption_summary_report_group + " "
        fuel_consumption_summary_report_liscense_plate = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_liscense_plate).text
        fuel_consumption_summary_report_type_vehicle = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_type_vehicle).text
        fuel_consumption_summary_report_time_run = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_time_run).text

        # fuel_consumption_summary_report_time_stop = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_time_stop).text

        fuel_consumption_summary_report_number_stop = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_number_stop).text
        fuel_consumption_summary_report_number_of_suction = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_number_of_suction).text   #số lần hút

        fuel_consumption_summary_report_so_lit_dau_ky = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_so_lit_dau_ky).text
        fuel_consumption_summary_report_so_lit_dau_ky = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_so_lit_dau_ky)))
        fuel_consumption_summary_report_so_lit_dau_ky = fuel_consumption_summary_report_so_lit_dau_ky[0:4]
        fuel_consumption_summary_report_so_lit_dau_ky_excel = str(sheet["H7"].value)
        fuel_consumption_summary_report_so_lit_dau_ky_excel = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_so_lit_dau_ky_excel)))
        fuel_consumption_summary_report_so_lit_dau_ky_excel = fuel_consumption_summary_report_so_lit_dau_ky_excel[0:4]

        fuel_consumption_summary_report_so_lit_do = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_so_lit_do).text
        fuel_consumption_summary_report_so_lit_do = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_so_lit_do)))
        fuel_consumption_summary_report_so_lit_do = fuel_consumption_summary_report_so_lit_do[0:4]
        fuel_consumption_summary_report_so_lit_do_excel = str(sheet["I7"].value)
        fuel_consumption_summary_report_so_lit_do_excel = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_so_lit_do_excel)))
        fuel_consumption_summary_report_so_lit_do_excel = fuel_consumption_summary_report_so_lit_do_excel[0:4]

        fuel_consumption_summary_report_so_lit_hut = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_so_lit_hut).text
        fuel_consumption_summary_report_so_lit_hut = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_so_lit_hut)))
        fuel_consumption_summary_report_so_lit_hut = fuel_consumption_summary_report_so_lit_hut[0:4]
        fuel_consumption_summary_report_so_lit_hut_excel = str(sheet["J7"].value)
        fuel_consumption_summary_report_so_lit_hut_excel = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_so_lit_hut_excel)))
        fuel_consumption_summary_report_so_lit_hut_excel = fuel_consumption_summary_report_so_lit_hut_excel[0:4]

        fuel_consumption_summary_report_so_lit_cuoi_ky = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_so_lit_cuoi_ky).text
        fuel_consumption_summary_report_so_lit_cuoi_ky = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_so_lit_cuoi_ky)))
        fuel_consumption_summary_report_so_lit_cuoi_ky = fuel_consumption_summary_report_so_lit_cuoi_ky[0:4]
        fuel_consumption_summary_report_so_lit_cuoi_ky_excel = str(sheet["K7"].value)
        fuel_consumption_summary_report_so_lit_cuoi_ky_excel = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_so_lit_cuoi_ky_excel)))
        fuel_consumption_summary_report_so_lit_cuoi_ky_excel = fuel_consumption_summary_report_so_lit_cuoi_ky_excel[0:4]


        fuel_consumption_summary_report_nl_tieu_thu_dinh_muc = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_nl_tieu_thu_dinh_muc).text
        fuel_consumption_summary_report_nl_tieu_thu_thuc_te = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_nl_tieu_thu_thuc_te).text
        fuel_consumption_summary_report_nl_tieu_thu_thuc_te = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_nl_tieu_thu_thuc_te)))
        fuel_consumption_summary_report_nl_tieu_thu_thuc_te = fuel_consumption_summary_report_nl_tieu_thu_thuc_te[0:4]
        fuel_consumption_summary_report_nl_tieu_thu_thuc_te_excel = str(sheet["M7"].value)
        fuel_consumption_summary_report_nl_tieu_thu_thuc_te_excel = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_nl_tieu_thu_thuc_te_excel)))
        fuel_consumption_summary_report_nl_tieu_thu_thuc_te_excel = fuel_consumption_summary_report_nl_tieu_thu_thuc_te_excel[0:4]

        fuel_consumption_summary_report_dinh_muc_quy_dinh = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_dinh_muc_quy_dinh).text

        fuel_consumption_summary_report_dinh_muc_thuc_te = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_dinh_muc_thuc_te).text
        fuel_consumption_summary_report_dinh_muc_thuc_te = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_dinh_muc_thuc_te)))
        fuel_consumption_summary_report_dinh_muc_thuc_te = fuel_consumption_summary_report_dinh_muc_thuc_te[0:3]
        fuel_consumption_summary_report_dinh_muc_thuc_te_excel = str(sheet["O7"].value)
        fuel_consumption_summary_report_dinh_muc_thuc_te_excel = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_dinh_muc_thuc_te_excel)))
        fuel_consumption_summary_report_dinh_muc_thuc_te_excel = fuel_consumption_summary_report_dinh_muc_thuc_te_excel[0:3]


        fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may).text
        fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may)))
        fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may = fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may[0:2]
        fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may_excel = str(sheet["P7"].value)
        fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may_excel = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may_excel)))
        fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may_excel = fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may_excel[0:2]


        # fuel_consumption_summary_report_dinh_muc_thuc_te_dung_do = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_dinh_muc_thuc_te_dung_do).text

        fuel_consumption_summary_report_tong_km = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_tong_km).text
        fuel_consumption_summary_report_tong_km = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_tong_km)))
        fuel_consumption_summary_report_tong_km = fuel_consumption_summary_report_tong_km[0:2]
        fuel_consumption_summary_report_tong_km_excel = str(sheet["Q7"].value)
        fuel_consumption_summary_report_tong_km_excel = str(''.join(re.findall(r'\d+', fuel_consumption_summary_report_tong_km_excel)))
        fuel_consumption_summary_report_tong_km_excel = fuel_consumption_summary_report_tong_km_excel[0:2]


        fuel_consumption_summary_report_km_co = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_km_co).text
        fuel_consumption_summary_report_chi_tiet = var.driver.find_element(By.XPATH, var.fuel_consumption_summary_report_chi_tiet).text



        logging.info("Báo cáo doanh nghiệp - Báo cáo tổng hợp tiêu hao nhiên liệu")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "6"].value)
            print(sheet[column + "6"])
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "A6", "STT", "A7", fuel_consumption_summary_report_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "B6", "Nhóm xe", "B7", fuel_consumption_summary_report_group)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "C6", "Biển kiểm soát", "C7", fuel_consumption_summary_report_liscense_plate)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "D6", "Loại phương tiện", "D7", fuel_consumption_summary_report_type_vehicle)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "E6", "Thời gian bật máy", "E7", fuel_consumption_summary_report_time_run)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "F6", "Số lần đổ", "F7", fuel_consumption_summary_report_number_stop)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "G6", "Số lần hút", "G7", fuel_consumption_summary_report_number_of_suction)
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "H6", "Số lít đầu kỳ")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "I6", "Số lít đổ")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "J6", "Số lít hút")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "K6", "Số lít cuối kỳ")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "M6", "Nhiên liệu tiêu thụ thực tế")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "O6", "Định mức thực tế")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "Q6", "Tổng km")
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "L6", "Nhiên liệu tiêu thụ định mức", "L7", fuel_consumption_summary_report_nl_tieu_thu_dinh_muc)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "N6", "Định mức quy định", "N7", fuel_consumption_summary_report_dinh_muc_quy_dinh)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "R6", "Km cơ", "R7", fuel_consumption_summary_report_km_co)
        chucnangkhac.write_result_excel_checkweb(code, fuel_consumption_summary_report_chi_tiet, "Chi tiết")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_summary_report_so_lit_dau_ky, fuel_consumption_summary_report_so_lit_dau_ky_excel, "Số lít đầu kỳ")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_summary_report_so_lit_do, fuel_consumption_summary_report_so_lit_do_excel, "Số lít đổ")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_summary_report_so_lit_hut, fuel_consumption_summary_report_so_lit_hut_excel, "Số lít hút")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_summary_report_so_lit_cuoi_ky, fuel_consumption_summary_report_so_lit_cuoi_ky_excel, "Số lít cuối kỳ")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_summary_report_nl_tieu_thu_thuc_te, fuel_consumption_summary_report_nl_tieu_thu_thuc_te_excel, "Nhiên liệu tiêu thụ thực tế")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_summary_report_dinh_muc_thuc_te, fuel_consumption_summary_report_dinh_muc_thuc_te_excel, "Định mức thực tế")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may, fuel_consumption_summary_report_dinh_muc_thuc_te_bat_may_excel, "Định mức thực tế bật máy")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_summary_report_tong_km, fuel_consumption_summary_report_tong_km_excel, "Tổng km")



    def fuel_consumption_summary_report_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report11)
        except:
            fuel_report.fuel_consumption_summary_report(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(3)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo tổng hợp tiêu hao nhiên liệu")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "__BaoCaoDoanhNghiep_BaoCaoTongHopTieuHaoNhienLieu_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "__BaoCaoDoanhNghiep_BaoCaoTongHopTieuHaoNhienLieu_AnHienCot .png")
        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass









    def fuel_consumption_daily_report(self, code, eventname, result):     #Báo cáo tiêu hao nhiên liệu
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo tiêu hao nhiên liệu",
                                              var.check_report_km_activity_summary, "BÁO CÁO TIÊU HAO NHIÊN LIỆU", "_BaoCaoDoanhNghiep_BaoCaoTieuHaoNhienLieu.png")


    def fuel_consumption_daily_report_search(self, code, eventname,result):  #Báo cáo tiêu hao nhiên liệu - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report).click()
            time.sleep(5)
        write_from_date(var.stop_report_fromdate_input)
        time.sleep(1)
        button = var.driver.find_element(By.XPATH, var.report_search)
        var.driver.execute_script("arguments[0].click();", button)
        time.sleep(5)

        chucnangkhac.write_result_displayed_try(code, eventname, result,
                                                "Báo cáo doanh nghiệp - Báo cáo tiêu hao nhiên liệu",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoDoanhNghiep_BaoCaoTieuHaoNhienLieu_TimKiem.png")



    def fuel_consumption_daily_report_downloadexcel(self, code, eventname, result):    #Báo cáo tiêu hao nhiên liệu -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            # var.driver.find_element(By.XPATH, var.report_search).click()
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report).click()
            time.sleep(5)
            write_from_date(var.stop_report_fromdate_input)
            time.sleep(1)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(6)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaotieuhaonhienlieu.xlsx"))
        #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaotieuhaonhienlieu.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        fuel_consumption_daily_report_stt = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_stt).text
        fuel_consumption_daily_report_ngaythang = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_ngaythang).text
        fuel_consumption_daily_report_bienkiemsoat = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_bienkiemsoat).text
        fuel_consumption_daily_report_loaiphuongtien = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_loaiphuongtien).text

        fuel_consumption_daily_report_batdaudichuyen = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_batdaudichuyen).text
        fuel_consumption_daily_report_batdaudichuyen = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_batdaudichuyen)))
        fuel_consumption_daily_report_batdaudichuyen = fuel_consumption_daily_report_batdaudichuyen[0:4]
        fuel_consumption_daily_report_batdaudichuyen_excel = str(sheet["E6"].value)
        fuel_consumption_daily_report_batdaudichuyen_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_batdaudichuyen_excel)))
        fuel_consumption_daily_report_batdaudichuyen_excel = fuel_consumption_daily_report_batdaudichuyen_excel[0:4]

        fuel_consumption_daily_report_ketthucdichuyen = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_ketthucdichuyen).text
        fuel_consumption_daily_report_ketthucdichuyen = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_ketthucdichuyen)))
        fuel_consumption_daily_report_ketthucdichuyen = fuel_consumption_daily_report_ketthucdichuyen[0:4]
        fuel_consumption_daily_report_ketthucdichuyen_excel = str(sheet["F6"].value)
        fuel_consumption_daily_report_ketthucdichuyen_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_ketthucdichuyen_excel)))
        fuel_consumption_daily_report_ketthucdichuyen_excel = fuel_consumption_daily_report_ketthucdichuyen_excel[0:4]


        fuel_consumption_daily_report_thoigianbatmay = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_thoigianbatmay).text
        fuel_consumption_daily_report_thoigianbatmay = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_thoigianbatmay)))
        fuel_consumption_daily_report_thoigianbatmay = fuel_consumption_daily_report_thoigianbatmay[0:4]
        fuel_consumption_daily_report_thoigianbatmay_excel = str(sheet["G6"].value)
        fuel_consumption_daily_report_thoigianbatmay_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_thoigianbatmay_excel)))
        fuel_consumption_daily_report_thoigianbatmay_excel = fuel_consumption_daily_report_thoigianbatmay_excel[0:4]


        fuel_consumption_daily_report_thoigianlanbanh = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_thoigianlanbanh).text
        fuel_consumption_daily_report_thoigianlanbanh = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_thoigianlanbanh)))
        fuel_consumption_daily_report_thoigianlanbanh = fuel_consumption_daily_report_thoigianlanbanh[0:5]
        fuel_consumption_daily_report_thoigianlanbanh_excel = str(sheet["H6"].value)
        fuel_consumption_daily_report_thoigianlanbanh_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_thoigianlanbanh_excel)))
        fuel_consumption_daily_report_thoigianlanbanh_excel = fuel_consumption_daily_report_thoigianlanbanh_excel[0:5]

        fuel_consumption_daily_report_thoigiandungdonomay = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_thoigiandungdonomay).text
        fuel_consumption_daily_report_solando = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_solando).text
        fuel_consumption_daily_report_solanhut = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_solanhut).text

        fuel_consumption_daily_report_solitdaungay = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_solitdaungay).text
        fuel_consumption_daily_report_solitdaungay = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_solitdaungay)))
        fuel_consumption_daily_report_solitdaungay = fuel_consumption_daily_report_solitdaungay[0:4]
        fuel_consumption_daily_report_solitdaungay_excel = str(sheet["L6"].value)
        fuel_consumption_daily_report_solitdaungay_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_solitdaungay_excel)))
        fuel_consumption_daily_report_solitdaungay_excel = fuel_consumption_daily_report_solitdaungay_excel[0:4]





        fuel_consumption_daily_report_solitdo = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_solitdo).text
        fuel_consumption_daily_report_solitdo = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_solitdo)))
        fuel_consumption_daily_report_solitdo = fuel_consumption_daily_report_solitdo[0:4]
        fuel_consumption_daily_report_solitdo_excel = str(sheet["M6"].value)
        fuel_consumption_daily_report_solitdo_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_solitdo_excel)))
        fuel_consumption_daily_report_solitdo_excel = fuel_consumption_daily_report_solitdo_excel[0:4]








        fuel_consumption_daily_report_solithut = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_solithut).text
        fuel_consumption_daily_report_solithut = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_solithut)))
        fuel_consumption_daily_report_solithut = fuel_consumption_daily_report_solithut[0:4]
        fuel_consumption_daily_report_solithut_excel = str(sheet["N6"].value)
        fuel_consumption_daily_report_solithut_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_solithut_excel)))
        fuel_consumption_daily_report_solithut_excel = fuel_consumption_daily_report_solithut_excel[0:4]



        fuel_consumption_daily_report_solitton = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_solitton).text
        fuel_consumption_daily_report_solitton = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_solitton)))
        fuel_consumption_daily_report_solitton = fuel_consumption_daily_report_solitton[0:4]
        fuel_consumption_daily_report_solitton_excel = str(sheet["O6"].value)
        fuel_consumption_daily_report_solitton_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_solitton_excel)))
        fuel_consumption_daily_report_solitton_excel = fuel_consumption_daily_report_solitton_excel[0:4]

        fuel_consumption_daily_report_nhienlieutieuhao = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_nhienlieutieuhao).text
        fuel_consumption_daily_report_nhienlieutieuhao = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_nhienlieutieuhao)))
        fuel_consumption_daily_report_nhienlieutieuhao = fuel_consumption_daily_report_nhienlieutieuhao[0:4]
        fuel_consumption_daily_report_nhienlieutieuhao_excel = str(sheet["P6"].value)
        fuel_consumption_daily_report_nhienlieutieuhao_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_nhienlieutieuhao_excel)))
        fuel_consumption_daily_report_nhienlieutieuhao_excel = fuel_consumption_daily_report_nhienlieutieuhao_excel[0:4]

        fuel_consumption_daily_report_dinhmucquydinh = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_dinhmucquydinh).text

        fuel_consumption_daily_report_dinhmucthuctetheokm = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_dinhmucthuctetheokm).text
        fuel_consumption_daily_report_dinhmucthuctetheokm = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_dinhmucthuctetheokm)))
        fuel_consumption_daily_report_dinhmucthuctetheokm = fuel_consumption_daily_report_dinhmucthuctetheokm[0:3]
        fuel_consumption_daily_report_dinhmucthuctetheokm_excel = str(sheet["R6"].value)
        fuel_consumption_daily_report_dinhmucthuctetheokm_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_dinhmucthuctetheokm_excel)))
        fuel_consumption_daily_report_dinhmucthuctetheokm_excel = fuel_consumption_daily_report_dinhmucthuctetheokm_excel[0:3]

        # fuel_consumption_daily_report_dinhmucthuctedungdonomay = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_dinhmucthuctedungdonomay).text
        fuel_consumption_daily_report_dinhmucthuctebatmay = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_dinhmucthuctebatmay).text
        fuel_consumption_daily_report_dinhmucthuctebatmay = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_dinhmucthuctebatmay)))
        fuel_consumption_daily_report_dinhmucthuctebatmay = fuel_consumption_daily_report_dinhmucthuctebatmay[0:3]
        fuel_consumption_daily_report_dinhmucthuctebatmay_excel = str(sheet["S6"].value)
        fuel_consumption_daily_report_dinhmucthuctebatmay_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_dinhmucthuctebatmay_excel)))
        fuel_consumption_daily_report_dinhmucthuctebatmay_excel = fuel_consumption_daily_report_dinhmucthuctebatmay_excel[0:3]


        fuel_consumption_daily_report_sotien = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_sotien).text
        if fuel_consumption_daily_report_sotien == "":
            fuel_consumption_daily_report_sotien = "0"

        fuel_consumption_daily_report_km_gps = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_km_gps).text
        fuel_consumption_daily_report_km_gps = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_km_gps)))
        fuel_consumption_daily_report_km_gps = fuel_consumption_daily_report_km_gps[0:3]
        fuel_consumption_daily_report_km_gps_excel = str(sheet["U6"].value)
        fuel_consumption_daily_report_km_gps_excel = str(''.join(re.findall(r'\d+', fuel_consumption_daily_report_km_gps_excel)))
        fuel_consumption_daily_report_km_gps_excel = fuel_consumption_daily_report_km_gps_excel[0:3]



        fuel_consumption_daily_report_bieudo = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_bieudo).get_attribute("src")
        print("src1a:" + fuel_consumption_daily_report_bieudo)
        print("src1b:" + fuel_consumption_daily_report_bieudo[-24::])

        fuel_consumption_daily_report_lotrinh = var.driver.find_element(By.XPATH, var.fuel_consumption_daily_report_lotrinh).text

        logging.info("Báo cáo doanh nghiệp - Báo cáo tiêu hao nhiên liệu")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "5"].value)
            print(sheet[column + "5"])
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "A5", "STT", "A6", fuel_consumption_daily_report_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "B5", "Ngày tháng", "B6", fuel_consumption_daily_report_ngaythang)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "C5", "Biển kiểm soát", "C6", fuel_consumption_daily_report_bienkiemsoat)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "D5", "Loại phương tiện", "D6", fuel_consumption_daily_report_loaiphuongtien)

            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "E5", "Bắt đầu di chuyển")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "F5", "Kết thúc di chuyển")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "G5", "Thời gian bật máy")

            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "H5", "Thời gian lăn bánh")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "L5", "Số lít đầu ngày")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "O5", "Số lít tồn")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "P5", "Nhiên liệu tiêu hao")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "R5", "Định mức thực tế theo km")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "S5", "Định mức thực tế bật máy")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "5", "U5", "Km GPS")


            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "I5", "Thời gian dừng đỗ bật máy", "I6", fuel_consumption_daily_report_thoigiandungdonomay)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "J5", "Số lần đổ", "J6", fuel_consumption_daily_report_solando)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "K5", "Số lần hút", "K6", fuel_consumption_daily_report_solanhut)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "M5", "Số lít đổ", "M6", fuel_consumption_daily_report_solitdo)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "N5", "Số lít hút", "N6", fuel_consumption_daily_report_solithut)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "Q5", "Định mức quy định", "Q6", fuel_consumption_daily_report_dinhmucquydinh)
            # chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "S5", "Định mức thực tế dừng đỗ bật máy", "S6", fuel_consumption_daily_report_dinhmucthuctedungdonomay)
            # chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "S5", "Định mức thực tế bật máy", "S6", fuel_consumption_daily_report_dinhmucthuctebatmay)

            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "T5", "Số tiền", "T6", fuel_consumption_daily_report_sotien)

        chucnangkhac.write_result_excel_checkweb(code, fuel_consumption_daily_report_bieudo[-24::], "/Images/s_icon_graph.jpg")
        chucnangkhac.write_result_excel_checkweb(code, fuel_consumption_daily_report_lotrinh, "Lộ trình")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_daily_report_batdaudichuyen, fuel_consumption_daily_report_batdaudichuyen_excel, "Bắt đầu di chuyển")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_daily_report_ketthucdichuyen, fuel_consumption_daily_report_ketthucdichuyen_excel, "Kết thúc di chuyển")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_daily_report_thoigianlanbanh, fuel_consumption_daily_report_thoigianlanbanh_excel, "Thời gian lăn bánh")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_daily_report_solitdaungay, fuel_consumption_daily_report_solitdaungay_excel, "Số lít đầu ngày")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_daily_report_solitdo, fuel_consumption_daily_report_solitdo_excel, "Số lít đổ")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_daily_report_solithut, fuel_consumption_daily_report_solithut_excel, "Số lít hút")



        chucnangkhac.write_result_excelreport2(code, fuel_consumption_daily_report_solitton, fuel_consumption_daily_report_solitton_excel, "Số lít tồn")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_daily_report_nhienlieutieuhao, fuel_consumption_daily_report_nhienlieutieuhao_excel, "Nhiên liệu tiêu hao")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_daily_report_dinhmucthuctetheokm, fuel_consumption_daily_report_dinhmucthuctetheokm_excel, "Định mức thực tế theo km")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_daily_report_dinhmucthuctebatmay, fuel_consumption_daily_report_dinhmucthuctebatmay_excel, "Định mức thực tế bật máy")
        chucnangkhac.write_result_excelreport2(code, fuel_consumption_daily_report_km_gps, fuel_consumption_daily_report_km_gps_excel, "Km GPS")



    def fuel_consumption_daily_report_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report12)
        except:
            fuel_report.fuel_consumption_daily_report(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(3)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo tiêu hao nhiên liệu")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoTieuHaoNhienLieu_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoTieuHaoNhienLieu_AnHienCot .png")
        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass









    def report_pour_fuel(self, code, eventname, result):     #Báo cáo đổ hút nhiên liệu
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.report_pour_fuel).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.report_pour_fuel).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo đổ hút nhiên liệu",
                                              var.check_report_km_activity_summary, "BÁO CÁO ĐỔ HÚT NHIÊN LIỆU", "_BaoCaoDoanhNghiep_BaoCaoDoHutNhienLieu.png")


    def report_pour_fuel_search(self, code, eventname,result):  #Báo cáo đổ hút nhiên liệu - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_pour_fuel).click()
            time.sleep(5)
        write_from_date(var.stop_report_fromdate_input)
        time.sleep(1)
        button = var.driver.find_element(By.XPATH, var.report_search)
        var.driver.execute_script("arguments[0].click();", button)
        time.sleep(5)

        chucnangkhac.write_result_displayed_try(code, eventname, result,
                                                "Báo cáo doanh nghiệp - Báo cáo đổ hút nhiên liệu",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoDoanhNghiep_BaoCaoDoHutNhienLieu_TimKiem.png")


    def report_pour_fuel_downloadexcel(self, code, eventname, result):    #Báo cáo đổ hút nhiên liệu -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            # var.driver.find_element(By.XPATH, var.report_search).click()
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.report_pour_fuel).click()
            time.sleep(5)
            write_from_date(var.stop_report_fromdate_input)
            time.sleep(1)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(6)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaodohutnhienlieu.xlsx"))

        #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaodohutnhienlieu.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        report_pour_fuel_stt = var.driver.find_element(By.XPATH, var.report_pour_fuel_stt).text
        report_pour_fuel_bienkiemsoat = var.driver.find_element(By.XPATH, var.report_pour_fuel_bienkiemsoat).text
        report_pour_fuel_thoigian = var.driver.find_element(By.XPATH, var.report_pour_fuel_thoigian).text

        report_pour_fuel_solittruoc = var.driver.find_element(By.XPATH, var.report_pour_fuel_solittruoc).text
        report_pour_fuel_solittruoc = str(''.join(re.findall(r'\d+', report_pour_fuel_solittruoc)))
        report_pour_fuel_solittruoc = report_pour_fuel_solittruoc[0:4]
        report_pour_fuel_solittruoc_excel = str(sheet["D7"].value)
        report_pour_fuel_solittruoc_excel = str(''.join(re.findall(r'\d+', report_pour_fuel_solittruoc_excel)))
        report_pour_fuel_solittruoc_excel = report_pour_fuel_solittruoc_excel[0:4]

        report_pour_fuel_solit = var.driver.find_element(By.XPATH, var.report_pour_fuel_solit).text
        report_pour_fuel_solit = str(''.join(re.findall(r'\d+', report_pour_fuel_solit)))
        report_pour_fuel_solit = report_pour_fuel_solit[0:4]
        report_pour_fuel_solit_excel = str(sheet["E7"].value)
        report_pour_fuel_solit_excel = str(''.join(re.findall(r'\d+', report_pour_fuel_solit_excel)))
        report_pour_fuel_solit_excel = report_pour_fuel_solit_excel[0:4]

        report_pour_fuel_solitsau = var.driver.find_element(By.XPATH, var.report_pour_fuel_solitsau).text
        report_pour_fuel_solitsau = str(''.join(re.findall(r'\d+', report_pour_fuel_solitsau)))
        report_pour_fuel_solitsau = report_pour_fuel_solitsau[0:4]
        report_pour_fuel_solitsau_excel = str(sheet["F7"].value)
        report_pour_fuel_solitsau_excel = str(''.join(re.findall(r'\d+', report_pour_fuel_solitsau_excel)))
        report_pour_fuel_solitsau_excel = report_pour_fuel_solitsau_excel[0:4]

        report_pour_fuel_sotien = var.driver.find_element(By.XPATH, var.report_pour_fuel_sotien).text
        report_pour_fuel_diachi = var.driver.find_element(By.XPATH, var.report_pour_fuel_diachi).text
        report_pour_fuel_ghichu = var.driver.find_element(By.XPATH, var.report_pour_fuel_ghichu).text
        report_pour_fuel_lotrinh = var.driver.find_element(By.XPATH, var.report_pour_fuel_lotrinh).text
        report_pour_fuel_bieudo = var.driver.find_element(By.XPATH, var.report_pour_fuel_bieudo).get_attribute("src")
        print("src2a:" + report_pour_fuel_bieudo)
        print("src2b:" + report_pour_fuel_bieudo[-24::])


        report_pour_fuel_xoa = var.driver.find_element(By.XPATH, var.report_pour_fuel_xoa).text


        logging.info("Báo cáo doanh nghiệp - Báo cáo đổ hút nhiên liệu")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "6"].value)
            print(sheet[column + "6"])
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "A6", "STT", "A7", report_pour_fuel_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "B6", "Biển kiểm soát", "B7", report_pour_fuel_bienkiemsoat)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "C6", "Thời gian", "C7", report_pour_fuel_thoigian)

            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "D6", "Số lít trước")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "E6", "Số lít")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "F6", "Số lít sau")
            chucnangkhac.write_result_excelreport(code, sheet, column, 'Data', "6", "I6", "Ghi chú")

            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "G6", "Số tiền", "G7", report_pour_fuel_sotien)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "H6", "Địa chỉ", "H7", report_pour_fuel_diachi)
        chucnangkhac.write_result_excel_checkweb(code, report_pour_fuel_lotrinh, "Lộ trình")
        chucnangkhac.write_result_excel_checkweb(code, report_pour_fuel_bieudo[-24::], "/Images/s_icon_graph.jpg")
        chucnangkhac.write_result_excel_checkweb(code, report_pour_fuel_xoa, "Xóa")
        chucnangkhac.write_result_excelreport2(code, report_pour_fuel_solittruoc, report_pour_fuel_solittruoc_excel, "Số lít trước")
        chucnangkhac.write_result_excelreport2(code, report_pour_fuel_solit, report_pour_fuel_solit_excel, "Số lít")
        chucnangkhac.write_result_excelreport2(code, report_pour_fuel_solitsau, report_pour_fuel_solitsau_excel, "Số lít sau")


    def report_pour_fuel_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report13)
        except:
            fuel_report.report_pour_fuel(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(3)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo đổ hút nhiên liệu")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoDoHutNhienLieu_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoDoHutNhienLieu_AnHienCot.png")
        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass





class system_report:

    def device_singnal_report(self, code, eventname, result):     #Báo cáo mất tín hiệu
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.device_singnal_report).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.device_singnal_report).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo doanh nghiệp - Báo cáo mất tín hiệu",
                                              var.check_report_km_activity_summary, "BÁO CÁO MẤT TÍN HIỆU", "_BaoCaoDoanhNghiep_BaoCaoMatTinHieu.png")


    def device_singnal_report_search(self, code, eventname,result):  #Báo cáo mất tín hiệu - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.device_singnal_report).click()
            time.sleep(5)
        write_from_date(var.stop_report_fromdate_input)
        time.sleep(1)
        button = var.driver.find_element(By.XPATH, var.report_search)
        var.driver.execute_script("arguments[0].click();", button)
        time.sleep(5)

        chucnangkhac.write_result_displayed_try(code, eventname, result,
                                                "Báo cáo doanh nghiệp - Báo cáo mất tín hiệu",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoDoanhNghiep_BaoCaoMatTinHieu_TimKiem.png")



    def device_singnal_report_downloadexcel(self, code, eventname, result):    #Báo cáo mất tín hiệu -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            # var.driver.find_element(By.XPATH, var.report_search).click()
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.managerment_report).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.device_singnal_report).click()
            time.sleep(5)
            write_from_date(var.stop_report_fromdate_input)
            time.sleep(1)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(6)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"baocaomatinhieu.xlsx"))

        #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        wordbook = openpyxl.load_workbook(var.excelpath+"/baocaomatinhieu.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        device_singnal_report_stt = var.driver.find_element(By.XPATH, var.device_singnal_report_stt).text
        device_singnal_report_bienkiemsoat = var.driver.find_element(By.XPATH, var.device_singnal_report_bienkiemsoat).text
        device_singnal_report_nhom = var.driver.find_element(By.XPATH, var.device_singnal_report_nhom).text
        device_singnal_report_nhom = " " + device_singnal_report_nhom
        device_singnal_report_thoidiembatdau = var.driver.find_element(By.XPATH, var.device_singnal_report_thoidiembatdau).text
        device_singnal_report_thoidiemketthuc = var.driver.find_element(By.XPATH, var.device_singnal_report_thoidiemketthuc).text
        device_singnal_report_thoigian = var.driver.find_element(By.XPATH, var.device_singnal_report_thoigian).text
        device_singnal_report_diadiembatdau = var.driver.find_element(By.XPATH, var.device_singnal_report_diadiembatdau).text
        device_singnal_report_diadiemketthuc = var.driver.find_element(By.XPATH, var.device_singnal_report_diadiemketthuc).text
        device_singnal_report_trangthai = var.driver.find_element(By.XPATH, var.device_singnal_report_trangthai).text
        device_singnal_report_lotrinh = var.driver.find_element(By.XPATH, var.device_singnal_report_lotrinh).text

        logging.info("Báo cáo doanh nghiệp - Báo cáo mất tín hiệu")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "5"].value)
            print(sheet[column + "5"])
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "A5", "STT", "A6", device_singnal_report_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "B5", "Biển kiểm soát", "B6", device_singnal_report_bienkiemsoat)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "C5", "Nhóm", "C6", device_singnal_report_nhom)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "D5", "Thời điểm bắt đầu", "D6", device_singnal_report_thoidiembatdau)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "E5", "Thời điểm kết thúc", "E6", device_singnal_report_thoidiemketthuc)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "F5", "Thời gian", "F6", device_singnal_report_thoigian)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "G5", "Địa điểm bắt đầu", "G6", device_singnal_report_diadiembatdau)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "H5", "Địa điểm kết thúc", "H6", device_singnal_report_diadiemketthuc)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "5", "I5", "Trạng thái", "I6", device_singnal_report_trangthai)
        chucnangkhac.write_result_excel_checkweb(code, device_singnal_report_lotrinh, "Lộ trình")


    def device_singnal_report_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report14)
        except:
            system_report.device_singnal_report(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(3)
        logging.info("----------------------")
        logging.info("Báo cáo doanh nghiệp - Báo cáo mất tín hiệu")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoDoanhNghiep_BaoCaoMatTinHieu_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoDoanhNghiep_BaoCaoMatTinHieu_AnHienCot.png")
        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass






class report_BGT:

    def speed_over_report(self, code, eventname, result):       #Quá tốc độ giới hạn
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.report_BGT).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.speed_over_report).click()
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.report_BGT).click()
            time.sleep(5)
            var.driver.find_element(By.XPATH, var.speed_over_report).click()
        time.sleep(5)
        chucnangkhac.write_result_text_try_if(code, eventname, result, "Báo cáo BGT -   Quá tốc độ giới hạn",
                                              var.check_report_km_activity_summary, "QUÁ TỐC ĐỘ GIỚI HẠN", "_BaoCaoBGT_QuaTocDoGioiHan.png")


    def speed_over_report_search(self, code, eventname,result):  #Quá tốc độ giới hạn - tìm kiếm
        var.driver.implicitly_wait(5)
        try:
            var.driver.find_element(By.XPATH, var.fromdate_input)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.report_BGT).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.speed_over_report).click()
            time.sleep(5)

        var.driver.find_element(By.XPATH, var.fromdate_input).click()
        xoa = var.driver.find_element(By.XPATH, var.fromdate_input)
        xoa.send_keys(Keys.CONTROL, "a")
        var.driver.find_element(By.XPATH, var.fromdate_input).send_keys("01/05/2024")
        time.sleep(1)

        var.driver.find_element(By.XPATH, var.todate_input).click()
        time.sleep(1)
        xoa = var.driver.find_element(By.XPATH, var.todate_input)
        xoa.send_keys(Keys.CONTROL, "a")
        var.driver.find_element(By.XPATH, var.todate_input).send_keys("10/06/2024")


        # write_from_date_month(var.fromdate_input)
        time.sleep(1)
        button = var.driver.find_element(By.XPATH, var.report_search)
        var.driver.execute_script("arguments[0].click();", button)
        time.sleep(5)

        chucnangkhac.write_result_displayed_try(code, eventname, result,
                                                "Báo cáo BGT -   Quá tốc độ giới hạn",
                                                var.check_activity_synthesis_report_search,
                                                "_BaoCaoBGT_QuaTocDoGioiHan_TimKiem.png")


    def speed_over_report_downloadexcel(self, code, eventname, result):    #Quá tốc độ giới hạn -  checkdownload
        var.driver.implicitly_wait(5)
        chucnangkhac.delete_excel()
        try:
            var.driver.find_element(By.XPATH, var.check_report_search)
        except:
            login.login.login_v2(self, var.data['login']['conhom_quantri_tk'], var.data['login']['conhom_quantri_mk'])
            var.driver.find_element(By.XPATH, var.report_BGT).click()
            time.sleep(4)
            var.driver.find_element(By.XPATH, var.speed_over_report).click()
            time.sleep(5)
            write_from_date_month(var.stop_report_fromdate_input)
            time.sleep(1)
            button = var.driver.find_element(By.XPATH, var.report_search)
            var.driver.execute_script("arguments[0].click();", button)
            time.sleep(6)
        var.driver.find_element(By.XPATH, var.downloadexcel).click()
        time.sleep(15)
        filename = max([var.excelpath + "\\" + f for f in os.listdir(var.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var.excelpath, r"quatocdogioihan.xlsx"))

        #Đọc check file excel
        bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        wordbook = openpyxl.load_workbook(var.excelpath+"/quatocdogioihan.xlsX")
        sheet = wordbook.get_sheet_by_name("Data")

        speed_over_report_stt = var.driver.find_element(By.XPATH, var.speed_over_report_stt).text
        speed_over_report_biensoxe = var.driver.find_element(By.XPATH, var.speed_over_report_biensoxe).text
        speed_over_report_hotenlaixe = var.driver.find_element(By.XPATH, var.speed_over_report_hotenlaixe).text
        speed_over_report_sogiaypheplaixe = var.driver.find_element(By.XPATH, var.speed_over_report_sogiaypheplaixe).text
        speed_over_report_loaihinhvandong = var.driver.find_element(By.XPATH, var.speed_over_report_loaihinhvandong).text
        speed_over_report_thoidiem = var.driver.find_element(By.XPATH, var.speed_over_report_thoidiem).text
        speed_over_report_tocdotrungbinhkhiquagioiohan = var.driver.find_element(By.XPATH, var.speed_over_report_tocdotrungbinhkhiquagioiohan).text
        speed_over_report_tocdogioihan = var.driver.find_element(By.XPATH, var.speed_over_report_tocdogioihan).text
        speed_over_report_tocdoquagioiohan = var.driver.find_element(By.XPATH, var.speed_over_report_tocdoquagioiohan).text
        speed_over_report_diadiemquatocdogioihan = var.driver.find_element(By.XPATH, var.speed_over_report_diadiemquatocdogioihan).text
        speed_over_report_ghichu = var.driver.find_element(By.XPATH, var.speed_over_report_ghichu).text

        logging.info("Báo cáo BGT -   Quá tốc độ giới hạn")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        for column in bangchucai:
            print(sheet[column + "6"].value)
            print(sheet[column + "6"])
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "A6", "TT", "A7", speed_over_report_stt)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "B6", "Biến số xe", "B7", speed_over_report_biensoxe)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "C6", "Họ tên lái xe", "C7", speed_over_report_hotenlaixe)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "D6", "Số giấy phép lái xe", "D7", speed_over_report_sogiaypheplaixe)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "E6", "Loại hình hoạt động", "E7", speed_over_report_loaihinhvandong)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "F6", "Thời điểm", "F7", speed_over_report_thoidiem)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "G6", "Tốc độ trung bình khi quá tốc độ giới hạn (km/h)", "G7", speed_over_report_tocdotrungbinhkhiquagioiohan)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "H6", "Tốc độ giới hạn (km/h)", "H7", speed_over_report_tocdogioihan)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "I6", "Tọa độ quá tốc độ giới hạn", "I7", speed_over_report_tocdoquagioiohan)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "J6", "Địa điểm quá tốc độ giới hạn", "J7", speed_over_report_diadiemquatocdogioihan)
            chucnangkhac.write_result_excelreport1(code, sheet, column, 'Data', "6", "K6", "Ghi chú", "K7", speed_over_report_ghichu)


    def speed_over_report_hide_column(self, code, eventname, result):
        var.driver.implicitly_wait(4)
        try:
            var.driver.find_element(By.XPATH, var.check_activity_synthesis_report17)
        except:
            report_BGT.speed_over_report(self, "", "", "")

        var.driver.find_element(By.XPATH, var.icon_hide_column).click()
        time.sleep(3)
        logging.info("----------------------")
        logging.info("Báo cáo BGT -   Quá tốc độ giới hạn")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            var.driver.find_element(By.XPATH, var.checkbox_typevehicle).click()
            logging.info("True")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Pass")
            time.sleep(1)
        except:
            logging.info("False")
            var.driver.save_screenshot(var.imagepath + code + "_BaoCaoBGT_QuaTocDoGioiHan_AnHienCot.png")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 7, "Fail")
            chucnangkhac.writeData(var.checklistpath, "Checklist", code, 13, code + "_BaoCaoBGT_QuaTocDoGioiHan_AnHienCot.png")
        try:
            button = var.driver.find_element(By.XPATH, var.hide_column_cancel1)
            var.driver.execute_script("arguments[0].click();", button)
            print("đã vào 1")
            time.sleep(1.5)
        except:
            pass

